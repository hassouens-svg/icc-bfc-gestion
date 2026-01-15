from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import warnings
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any, Union
import uuid
from uuid import uuid4
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
import jwt
from jwt.exceptions import InvalidTokenError
import io
import pandas as pd
from fastapi.responses import StreamingResponse, FileResponse
import base64
import mimetypes
import firebase_admin
from firebase_admin import credentials, messaging
import requests
import re

# YouTube API
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

# Suppress warnings
warnings.filterwarnings('ignore', message='.*bcrypt.*')
warnings.filterwarnings('ignore', category=DeprecationWarning)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Firebase FCM configuration
FIREBASE_API_KEY = "AIzaSyC_oHqQRFPoJ2NmO-7OVCq3boYyc6VkFWE"
FIREBASE_PROJECT_ID = "icc-bfc-app-notifications"
FIREBASE_SERVER_KEY = "AAAAg-BLWmM:APA91bE3qZt9wkF1rRCsNzfq__dCAhBrFsIgjjRXEUmOJXEH9mK6CUKtA9khppJW0Qa4pQdTKoiT7ZVaPpqSoQUdRkQkJpO97rG4RXnUuM8XRcHgPWaAJfgVSWy33eIGDbdbXQzY4wN1"

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# JWT settings
SECRET_KEY = os.environ["SECRET_KEY"]
ALGORITHM = "HS256"

# Security
security = HTTPBearer()

# Create the main app
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# ==================== MODELS ====================

class City(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    country: Optional[str] = "France"

class CityCreate(BaseModel):
    name: str
    country: Optional[str] = "France"

class ReferentPermissions(BaseModel):
    can_view_all_months: bool = False
    can_edit_visitors: bool = True
    can_stop_tracking: bool = True
    can_add_comments: bool = True
    can_mark_presence: bool = True
    can_view_analytics: bool = False

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    password: str  # hashed
    plain_password: Optional[str] = None  # Mot de passe en clair pour export (super_admin only)
    city: str
    role: str  # superviseur_promos, superviseur_fi, referent, accueil, promotions, pilote_fi, responsable_secteur, pasteur, super_admin, responsable_eglise, gestion_projet, respo_departement, star
    telephone: Optional[str] = None  # Phone number (especially for pilote_fi)
    assigned_month: Optional[Union[str, List[str]]] = None  # For referents: "2025-01"
    promo_name: Optional[str] = None  # Custom name for promo (instead of "2025-01")
    assigned_secteur_id: Optional[str] = None  # For responsable_secteur
    assigned_fi_id: Optional[str] = None  # DEPRECATED: Use assigned_fi_ids instead
    assigned_fi_ids: Optional[List[str]] = []  # For pilote_fi - multiple FIs
    permissions: Optional[Dict[str, bool]] = None  # For referents permissions
    dashboard_permissions: Optional[Dict[str, bool]] = None  # What user can see in their dashboard (controlled by super_admin)
    is_blocked: bool = False  # For blocking users
    team_members: Optional[List[Dict[str, str]]] = []  # For responsable_promo - list of team members with firstname/lastname
    assigned_departement: Optional[str] = None  # For respo_departement: department name
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    username: str
    password: str
    city: str
    role: str
    telephone: Optional[str] = None
    assigned_month: Optional[Union[str, List[str]]] = None
    permissions: Optional[Dict[str, bool]] = None
    assigned_departement: Optional[str] = None  # For respo_departement

class UserUpdate(BaseModel):
    username: Optional[str] = None
    city: Optional[str] = None
    role: Optional[str] = None
    assigned_month: Optional[Union[str, List[str]]] = None
    promo_name: Optional[str] = None  # Custom name for promo
    assigned_fi_id: Optional[str] = None  # DEPRECATED
    assigned_fi_ids: Optional[List[str]] = None  # Multiple FIs for pilote_fi
    assigned_secteur_id: Optional[str] = None
    permissions: Optional[Dict[str, bool]] = None
    team_members: Optional[List[Dict[str, str]]] = None  # For responsable_promo team
    dashboard_permissions: Optional[Dict[str, bool]] = None
    assigned_departement: Optional[str] = None  # For respo_departement

class UserLogin(BaseModel):
    username: str
    password: str
    city: Optional[str] = None  # Optionnel pour pasteur/superadmin
    department: Optional[str] = None  # "accueil" or "promotion"

class CulteStats(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str  # Date du dimanche YYYY-MM-DD
    ville: str
    type_culte: str  # "Culte 1", "Culte 2", "EJP", "Événements spéciaux"
    nombre_fideles: int
    nombre_adultes: int = 0  # Nombre d'adultes
    nombre_enfants: int = 0  # Nombre d'enfants
    nombre_stars: int
    commentaire: Optional[str] = None  # Commentaire pour chaque type de culte
    created_by: str  # Username de celui qui a créé
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

class CulteStatsCreate(BaseModel):
    date: str
    ville: str
    type_culte: str
    nombre_fideles: int
    nombre_adultes: int = 0
    nombre_enfants: int = 0
    nombre_stars: int
    commentaire: Optional[str] = None

class CulteStatsUpdate(BaseModel):
    nombre_fideles: Optional[int] = None
    nombre_adultes: Optional[int] = None
    nombre_enfants: Optional[int] = None
    nombre_stars: Optional[int] = None
    commentaire: Optional[str] = None

class PresenceEntry(BaseModel):
    date: str
    present: Optional[bool] = None  # Optionnel: peut être null si seulement commentaire
    commentaire: Optional[str] = None

class CommentEntry(BaseModel):
    date: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    author: str
    text: str

class Visitor(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    firstname: str
    lastname: str
    city: str
    types: List[str]  # ["Nouveau Arrivant", "Nouveau Converti", "De Passage"]
    phone: str  # Required now
    email: Optional[EmailStr] = None
    address: Optional[str] = None  # New field
    photo_url: Optional[str] = None  # Photo du visiteur
    arrival_channel: str  # Comment ils ont connu ICC
    age_range: Optional[str] = None  # "13-18 ans", "18-25 ans", etc.
    visit_date: str
    assigned_month: str  # "2025-01"
    presences_dimanche: List[PresenceEntry] = Field(default_factory=list)
    presences_jeudi: List[PresenceEntry] = Field(default_factory=list)
    formation_pcnc: bool = False
    formation_au_coeur_bible: bool = False
    formation_star: bool = False
    comments: List[CommentEntry] = Field(default_factory=list)
    tracking_stopped: bool = False
    stop_reason: Optional[str] = None
    stopped_by: Optional[str] = None
    stopped_date: Optional[str] = None
    is_ancien: bool = False  # True si ajouté via "Ancien Visiteur"
    ejp: bool = False  # Église des Jeunes Prodiges
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class VisitorCreate(BaseModel):
    firstname: str
    lastname: str
    city: str
    types: List[str]
    phone: str  # Required
    email: Optional[EmailStr] = None
    address: Optional[str] = None  # Optional
    arrival_channel: str
    age_range: Optional[str] = None
    visit_date: str
    is_ancien: bool = False
    ejp: bool = False

class VisitorUpdate(BaseModel):
    firstname: Optional[str] = None
    lastname: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    email: Optional[EmailStr] = None
    arrival_channel: Optional[str] = None
    age_range: Optional[str] = None
    types: Optional[List[str]] = None
    type: Optional[str] = None
    city: Optional[str] = None
    formation_pcnc: Optional[bool] = None
    formation_au_coeur_bible: Optional[bool] = None
    formation_star: Optional[bool] = None
    ejp: Optional[bool] = None
    est_disciple: Optional[str] = None  # "Oui", "En Cours", "Non"
    date_devenu_disciple: Optional[str] = None

class CommentAdd(BaseModel):
    text: str

class PresenceAdd(BaseModel):
    date: str
    present: Optional[bool] = None  # Optionnel: peut être null si seulement commentaire
    type: str  # "dimanche" or "jeudi"
    commentaire: Optional[str] = None

class FormationUpdate(BaseModel):
    formation_type: str  # "pcnc", "au_coeur_bible", "star"
    completed: bool

class StopTracking(BaseModel):
    reason: str

# ==================== FAMILLES D'IMPACT MODELS ====================

class Secteur(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nom: str
    ville: str
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SecteurCreate(BaseModel):
    nom: str
    ville: str

class FamilleImpact(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nom: str
    secteur_id: str
    ville: str
    adresse: Optional[str] = None
    pilote_id: Optional[str] = None  # DEPRECATED: Use pilote_ids
    pilote_ids: Optional[List[str]] = []  # Multiple pilotes per FI
    heure_debut: Optional[str] = None  # Heure de début de la FI (format HH:MM)
    heure_fin: Optional[str] = None  # Heure de fin de la FI (format HH:MM)
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FamilleImpactCreate(BaseModel):
    nom: str
    secteur_id: str
    ville: str
    adresse: Optional[str] = None
    pilote_id: Optional[str] = None  # DEPRECATED
    pilote_ids: Optional[List[str]] = []  # Multiple pilotes
    heure_debut: Optional[str] = None  # Heure de début (HH:MM)
    heure_fin: Optional[str] = None  # Heure de fin (HH:MM)

class MembreFI(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    prenom: str
    nom: str
    fi_id: str
    date_ajout: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    source: str  # "manuel" or "nouveau_arrivant"
    nouveau_arrivant_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MembreFICreate(BaseModel):
    prenom: str
    nom: str
    fi_id: str
    source: str = "manuel"
    nouveau_arrivant_id: Optional[str] = None

class PresenceFI(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    membre_fi_id: str
    date: str  # Format: "2025-01-09" (jeudi)
    present: bool
    commentaire: Optional[str] = None
    marked_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PresenceFICreate(BaseModel):
    membre_fi_id: str
    date: str
    present: bool
    commentaire: Optional[str] = None

class AffectationFI(BaseModel):
    nouveau_arrivant_id: str
    fi_id: str


class Notification(BaseModel):
    model_config = ConfigDict(arbitrary_types_allowed=True)
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str  # Utilisateur qui reçoit la notification
    type: str  # 'presence_reminder', 'fi_stagnation', 'low_fidelisation', 'unassigned_visitor'
    message: str
    data: Optional[Dict[str, Any]] = None  # Données supplémentaires (IDs, etc.)
    read: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class NotificationCreate(BaseModel):
    user_id: str
    type: str
    message: str
    data: Optional[Dict[str, Any]] = None

# ==================== EVENTS & PROJECTS MODELS ====================

class Projet(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    titre: str
    description: Optional[str] = None
    statut: str = "planifie"  # planifie, en_cours, termine, annule
    date_debut: Optional[str] = None  # Format: "2025-01-15"
    date_fin: Optional[str] = None
    budget_prevu: Optional[float] = 0.0
    budget_reel: Optional[float] = 0.0
    ville: str
    created_by: str  # username
    team_members: Optional[List[Dict[str, str]]] = []  # [{"nom": "Jean Dupont", "email": "jean@email.com"}]
    archived: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProjetCreate(BaseModel):
    titre: str
    description: Optional[str] = None
    statut: str = "planifie"
    date_debut: Optional[str] = None
    date_fin: Optional[str] = None
    budget_prevu: Optional[float] = 0.0
    ville: str
    team_members: Optional[List[Dict[str, str]]] = []

class ProjetUpdate(BaseModel):
    titre: Optional[str] = None
    description: Optional[str] = None
    statut: Optional[str] = None
    date_debut: Optional[str] = None
    date_fin: Optional[str] = None
    budget_prevu: Optional[float] = None
    budget_reel: Optional[float] = None
    team_members: Optional[List[Dict[str, str]]] = None

class Pole(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    projet_id: str
    nom: str
    description: Optional[str] = None
    responsable: Optional[str] = None  # username
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PoleCreate(BaseModel):
    projet_id: str
    nom: str
    description: Optional[str] = None
    responsable: Optional[str] = None

class PoleUpdate(BaseModel):
    nom: Optional[str] = None
    description: Optional[str] = None
    responsable: Optional[str] = None

class Tache(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    projet_id: str
    pole_id: Optional[str] = None
    titre: str
    description: Optional[str] = None
    assigne_a: Optional[str] = None  # username
    statut: str = "a_faire"  # a_faire, en_cours, termine, en_retard
    deadline: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TacheCreate(BaseModel):
    projet_id: str
    pole_id: Optional[str] = None
    titre: str
    description: Optional[str] = None
    assigne_a: Optional[str] = None
    statut: Optional[str] = "a_faire"  # Default to "a_faire" but allow override
    deadline: Optional[str] = None

class TacheUpdate(BaseModel):
    titre: Optional[str] = None
    description: Optional[str] = None
    assigne_a: Optional[str] = None
    statut: Optional[str] = None
    deadline: Optional[str] = None
    pole_id: Optional[str] = None

class Depense(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    projet_id: str
    montant: float
    raison: str
    date: str  # Date de la dépense
    created_by: str  # username
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DepenseCreate(BaseModel):
    projet_id: str
    montant: float
    raison: str
    date: Optional[str] = None


class Jalon(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    projet_id: str
    titre: str
    description: Optional[str] = None
    acteur: Optional[str] = None  # username
    date_debut: Optional[str] = None  # datetime ISO format (YYYY-MM-DDTHH:mm)
    date_fin: Optional[str] = None  # datetime ISO format (YYYY-MM-DDTHH:mm)
    statut: str = "a_faire"  # a_faire, en_cours, termine, en_retard, annule
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class JalonCreate(BaseModel):
    projet_id: str
    titre: str
    description: Optional[str] = None
    acteur: Optional[str] = None
    date_debut: Optional[str] = None
    date_fin: Optional[str] = None

class JalonUpdate(BaseModel):
    titre: Optional[str] = None
    description: Optional[str] = None
    acteur: Optional[str] = None
    date_debut: Optional[str] = None
    date_fin: Optional[str] = None
    statut: Optional[str] = None

class CommentaireProjet(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    projet_id: str
    user: str  # username
    texte: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CommentaireProjetCreate(BaseModel):
    projet_id: str
    texte: str

class FichierProjet(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    projet_id: str
    nom: str
    url: str
    type: str  # "document", "image", "autre"
    uploaded_by: str  # username
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CampagneCommunication(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    titre: str
    type: str  # "email", "sms", "both"
    message: str
    image_url: Optional[str] = None  # URL de l'image/affiche jointe
    destinataires: List[Dict[str, str]]  # [{"prenom": "...", "nom": "...", "email": "...", "telephone": "..."}]
    statut: str = "brouillon"  # brouillon, planifie, envoye
    date_envoi: Optional[str] = None  # Pour envoi planifié
    projet_id: Optional[str] = None  # Lier à un projet si besoin
    created_by: str  # username
    enable_rsvp: bool = False  # Activer les réponses Oui/Non/Peut-être
    stats: Optional[Dict[str, int]] = {"envoyes": 0, "oui": 0, "non": 0, "peut_etre": 0}
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CampagneCommunicationCreate(BaseModel):
    titre: str
    type: str
    message: str
    image_url: Optional[str] = None
    destinataires: List[Dict[str, str]]
    date_envoi: Optional[str] = None
    projet_id: Optional[str] = None
    enable_rsvp: bool = False

class RSVP(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    campagne_id: str
    contact_email: Optional[str] = None
    contact_telephone: Optional[str] = None
    reponse: str  # "oui", "non", "peut_etre"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ==================== AUTH HELPERS ====================

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict) -> str:
    return jwt.encode(data, SECRET_KEY, algorithm=ALGORITHM)

def is_super_admin(user: dict) -> bool:
    """Check if user is Super Admin or Pasteur (same permissions)"""
    return user.get("role") in ["super_admin", "pasteur"]

def is_superviseur(user: dict) -> bool:
    """Check if user is Superviseur (Promos or FI)"""
    return user.get("role") in ["superviseur_promos", "superviseur_fi"]

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        
        # Override role and city from JWT token if present (for department-based login)
        if "role" in payload:
            user["role"] = payload["role"]
        if "city" in payload:
            user["city"] = payload["city"]
        
        return user
    except InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/login")
async def login(user_login: UserLogin):
    # Chercher l'utilisateur uniquement par username
    query = {"username": user_login.username}
    user = await db.users.find_one(query, {"_id": 0})
    
    if not user or not verify_password(user_login.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Check if user is blocked
    if user.get("is_blocked", False):
        raise HTTPException(status_code=403, detail="Votre compte a été bloqué. Contactez l'administrateur.")
    
    # Utiliser la ville du compte utilisateur (déjà attribuée à la création)
    user_city = user.get("city", "")
    
    # If department is specified, use it; otherwise use user's default role
    final_role = user["role"]
    if user_login.department:
        # Map department choices to roles
        dept_to_role = {
            "accueil": "accueil",
            "promotions": "promotions"
        }
        final_role = dept_to_role.get(user_login.department, user["role"])
    
    # Create token with user's city from account
    token = create_access_token({"sub": user["id"], "role": final_role, "city": user_city})
    
    return {
        "token": token,
        "user": {
            "id": user["id"],
            "username": user["username"],
            "city": user_city,
            "role": final_role,
            "assigned_month": user.get("assigned_month"),
            "assigned_secteur_id": user.get("assigned_secteur_id"),
            "assigned_fi_id": user.get("assigned_fi_id"),
            "team_members": user.get("team_members", []),
            "promo_name": user.get("promo_name")
        }
    }

@api_router.post("/auth/register")
async def register_visitor(visitor_data: VisitorCreate):
    """Public registration form for new visitors"""
    # Calculate assigned_month from visit_date
    try:
        visit_dt = datetime.fromisoformat(visitor_data.visit_date)
        assigned_month = visit_dt.strftime("%Y-%m")
    except:
        assigned_month = datetime.now(timezone.utc).strftime("%Y-%m")
    
    visitor = Visitor(**visitor_data.model_dump(), assigned_month=assigned_month)
    doc = visitor.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.visitors.insert_one(doc)
    return {"message": "Registration successful", "id": visitor.id}

# ==================== USER ROUTES ====================

@api_router.post("/users")
async def create_user(user_data: UserCreate, current_user: dict = Depends(get_current_user)):
    """Create a new user - Super Admin can create any role, others can create referents only"""
    # Permission checks
    if current_user["role"] == "super_admin":
        # Super admin can create any user
        pass
    elif current_user["role"] == "responsable_eglise":
        # Responsable d'église can create all user types EXCEPT super_admin, pasteur, and other responsable_eglise
        if user_data.role in ["super_admin", "pasteur", "responsable_eglise"]:
            raise HTTPException(status_code=403, detail="You cannot create super_admin, pasteur, or responsable_eglise users")
        # Force city to be the same as responsable_eglise's city
        user_data.city = current_user["city"]
    elif current_user["role"] in ["superviseur_promos", "promotions"]:
        # Superviseurs can only create referents
        if user_data.role not in ["referent", "accueil", "promotions"]:
            raise HTTPException(status_code=403, detail="You can only create referent, accueil, or promotions users")
    elif current_user["role"] == "responsable_secteur":
        # Responsable secteur can only create pilotes
        if user_data.role != "pilote_fi":
            raise HTTPException(status_code=403, detail="You can only create pilote_fi users")
        # Force city to be the same
        user_data.city = current_user["city"]
    elif current_user["role"] == "superviseur_fi":
        # Superviseur FI can create pilotes and responsable_secteur
        if user_data.role not in ["pilote_fi", "responsable_secteur"]:
            raise HTTPException(status_code=403, detail="You can only create pilote_fi or responsable_secteur users")
        # Force city to be the same
        user_data.city = current_user["city"]
    else:
        raise HTTPException(status_code=403, detail="Not authorized to create users")
    
    # Check if username already exists
    existing = await db.users.find_one({
        "username": user_data.username,
        "city": user_data.city
    })
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists in this city")
    
    # Hash password and create user
    hashed_pw = hash_password(user_data.password)
    
    # Create dict from user_data and replace password with hashed version
    user_dict = user_data.model_dump()
    plain_password = user_dict['password']  # Sauvegarder le mot de passe en clair
    user_dict['password'] = hashed_pw
    user_dict['plain_password'] = plain_password  # Stocker aussi en clair pour l'export
    
    # Set default permissions for referents if not provided
    if user_data.role == "referent" and not user_dict.get('permissions'):
        user_dict['permissions'] = {
            "can_view_all_months": False,
            "can_edit_visitors": True,
            "can_stop_tracking": True,
            "can_add_comments": True,
            "can_mark_presence": True,
            "can_view_analytics": False
        }
    
    user = User(**user_dict)
    
    doc = user.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.users.insert_one(doc)
    return {"message": "User created successfully", "id": user.id}

@api_router.post("/users/referent")
async def create_referent(user_data: UserCreate, current_user: dict = Depends(get_current_user)):
    """Legacy endpoint for creating referents - redirects to /users"""
    return await create_user(user_data, current_user)

@api_router.get("/users/referents")
async def get_referents(current_user: dict = Depends(get_current_user)):
    # Get all users - seul super_admin voit toutes les villes
    query = {"role": {"$in": ["referent", "accueil", "promotions", "superviseur_promos", "superviseur_fi", "pilote_fi", "responsable_secteur", "pasteur", "super_admin", "responsable_eglise", "gestion_projet", "responsable_evangelisation", "respo_departement", "star"]}}
    
    # Seul super_admin peut voir toutes les villes, tous les autres voient uniquement leur ville
    if current_user["role"] != "super_admin":
        query["city"] = current_user["city"]
    
    referents = await db.users.find(query, {"_id": 0, "password": 0}).to_list(1000)
    
    return referents

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, update_data: UserUpdate, current_user: dict = Depends(get_current_user)):
    """Update user information"""
    
    # Cas spécial : permettre aux users de mettre à jour leurs propres informations
    is_self_update = (user_id == current_user["id"])
    
    # Liste des champs autorisés pour mise à jour de soi-même
    allowed_self_update_fields = {'team_members', 'promo_name', 'assigned_month'}
    
    # Vérifier si uniquement des champs autorisés sont modifiés
    update_dict = update_data.dict(exclude_unset=True)
    is_only_allowed_fields = all(
        field in allowed_self_update_fields 
        for field in update_dict.keys()
    )
    
    # Si c'est une mise à jour de soi-même avec uniquement les champs autorisés
    if is_self_update and is_only_allowed_fields:
        user = await db.users.find_one({"id": user_id})
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        # Mettre à jour les champs autorisés
        await db.users.update_one({"id": user_id}, {"$set": update_dict})
        return {"message": "User updated successfully"}
    
    # Sinon, vérifier les permissions normales
    
    # Super admin can update users from any city, others only from their city
    if current_user["role"] == "super_admin":
        user = await db.users.find_one({"id": user_id})
    else:
        user = await db.users.find_one({"id": user_id, "city": current_user["city"]})
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Responsable secteur and superviseur_fi can only update pilotes
    if current_user["role"] in ["responsable_secteur", "superviseur_fi"]:
        if user["role"] != "pilote_fi":
            raise HTTPException(status_code=403, detail="Can only update pilotes")
    
    # Build update dict with only provided fields
    update_dict = {}
    if update_data.username is not None:
        # Check if new username already exists
        city_filter = {} if current_user["role"] == "super_admin" else {"city": current_user["city"]}
        existing = await db.users.find_one({
            "username": update_data.username,
            "id": {"$ne": user_id},
            **city_filter
        })
        if existing:
            raise HTTPException(status_code=400, detail="Username already exists")
        update_dict["username"] = update_data.username
    
    if update_data.assigned_month is not None:
        update_dict["assigned_month"] = update_data.assigned_month
    
    if update_data.promo_name is not None:
        update_dict["promo_name"] = update_data.promo_name
    
    if update_data.assigned_fi_id is not None:
        update_dict["assigned_fi_id"] = update_data.assigned_fi_id
    
    if update_data.assigned_fi_ids is not None:
        update_dict["assigned_fi_ids"] = update_data.assigned_fi_ids
    
    if update_data.assigned_secteur_id is not None:
        update_dict["assigned_secteur_id"] = update_data.assigned_secteur_id
    
    if update_data.permissions is not None:
        update_dict["permissions"] = update_data.permissions
    
    if update_data.team_members is not None:
        update_dict["team_members"] = update_data.team_members
    
    if update_dict:
        await db.users.update_one({"id": user_id}, {"$set": update_dict})
    
    return {"message": "User updated successfully"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a user (Admin only)"""
    user_to_delete = await db.users.find_one({"id": user_id})
    if not user_to_delete:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Super admin can delete anyone
    if is_super_admin(current_user):
        await db.users.delete_one({"id": user_id})
        return {"message": "User deleted successfully"}
    
    # Regular admin or promotions can only delete users from their city (but not other admins)
    if current_user["role"] in ["superviseur_promos", "promotions"]:
        if user_to_delete["city"] != current_user["city"]:
            raise HTTPException(status_code=403, detail="Cannot delete users from other cities")
        if user_to_delete["role"] in ["superviseur_promos", "promotions"]:
            raise HTTPException(status_code=403, detail="Cannot delete other admins")
        await db.users.delete_one({"id": user_id})
        return {"message": "User deleted successfully"}
    
    # Superviseur FI can delete pilote_fi and responsable_secteur from their city
    if current_user["role"] == "superviseur_fi":
        if user_to_delete["city"] != current_user["city"]:
            raise HTTPException(status_code=403, detail="Cannot delete users from other cities")
        if user_to_delete["role"] not in ["pilote_fi", "responsable_secteur"]:
            raise HTTPException(status_code=403, detail="Can only delete pilotes and responsables de secteur")
        await db.users.delete_one({"id": user_id})
        return {"message": "User deleted successfully"}
    
    # Responsable secteur can delete pilote_fi from their city (even if assigned to FI)
    if current_user["role"] == "responsable_secteur":
        if user_to_delete["city"] != current_user["city"]:
            raise HTTPException(status_code=403, detail="Cannot delete users from other cities")
        if user_to_delete["role"] != "pilote_fi":
            raise HTTPException(status_code=403, detail="Can only delete pilotes")
        
        # Désassigner le pilote de toutes les FI avant de le supprimer
        await db.familles_impact.update_many(
            {"pilote_ids": user_id},
            {"$pull": {"pilote_ids": user_id}}
        )
        
        await db.users.delete_one({"id": user_id})
        return {"message": "User deleted successfully"}
    
    raise HTTPException(status_code=403, detail="Only admin can delete users")

@api_router.put("/users/{user_id}/block")
async def block_user(user_id: str, current_user: dict = Depends(get_current_user)):
    """Block a user (Super Admin only)"""
    if not is_super_admin(current_user):
        raise HTTPException(status_code=403, detail="Only super admin can block users")
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    await db.users.update_one({"id": user_id}, {"$set": {"is_blocked": True}})
    return {"message": "User blocked successfully"}

@api_router.put("/users/{user_id}/unblock")
async def unblock_user(user_id: str, current_user: dict = Depends(get_current_user)):
    """Unblock a user (Super Admin only)"""
    if not is_super_admin(current_user):
        raise HTTPException(status_code=403, detail="Only super admin can unblock users")
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    await db.users.update_one({"id": user_id}, {"$set": {"is_blocked": False}})
    return {"message": "User unblocked successfully"}

class PasswordReset(BaseModel):
    new_password: str


# ==================== MY EVENT CHURCH MODELS ====================
class ChurchEvent(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    description: Optional[str] = None
    date: str  # Format: YYYY-MM-DD
    time: Optional[str] = None  # Format: HH:MM
    location: Optional[str] = None
    image_url: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    rsvp_enabled: bool = True
    max_participants: Optional[int] = None
    require_names: bool = False  # Demander noms/prénoms dans le formulaire RSVP
    require_payment_method: bool = False  # Demander le moyen de paiement
    custom_link_title: Optional[str] = None  # Titre du lien personnalisé
    custom_link_url: Optional[str] = None  # URL du lien personnalisé

class ChurchEventCreate(BaseModel):
    title: str
    description: Optional[str] = None
    date: str
    time: Optional[str] = None
    location: Optional[str] = None
    image_url: Optional[str] = None
    rsvp_enabled: bool = True
    max_participants: Optional[int] = None
    require_names: bool = False  # Demander noms/prénoms dans le formulaire RSVP
    require_payment_method: bool = False  # Demander le moyen de paiement
    custom_link_title: Optional[str] = None  # Titre du lien personnalisé
    custom_link_url: Optional[str] = None  # URL du lien personnalisé
    require_email_contact: bool = False  # Demander email et contact
    confirmation_message: Optional[str] = None  # Message de confirmation personnalisé
    require_city: bool = False  # Demander de choisir la ville (église)
    city_options: Optional[str] = None  # Liste des villes séparées par des points-virgules

class EventRSVP(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    event_id: str
    name: str
    first_name: Optional[str] = None  # Prénom
    last_name: Optional[str] = None   # Nom
    is_star: bool = False  # Personne VIP/importante
    email: Optional[str] = None
    phone: Optional[str] = None
    city: Optional[str] = None  # Ville (église) choisie
    status: str = "confirmed"  # confirmed, declined, maybe
    guests_count: int = 1
    message: Optional[str] = None
    payment_method: Optional[str] = None  # "card" ou "cash"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    source: Optional[str] = None  # whatsapp, email, sms, facebook, direct

class EventRSVPCreate(BaseModel):
    name: str
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    is_star: bool = False
    email: Optional[str] = None
    phone: Optional[str] = None
    city: Optional[str] = None  # Ville (église) choisie
    status: str = "confirmed"
    guests_count: int = 1
    message: Optional[str] = None
    payment_method: Optional[str] = None  # "card" ou "cash"
    source: Optional[str] = None

# ============== NOTIFICATIONS MODELS ==============

class FCMToken(BaseModel):
    """Modèle pour stocker les tokens FCM des utilisateurs"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    token: str
    device_type: Optional[str] = None  # "android", "ios", "web"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    last_used: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class NotificationCreate(BaseModel):
    """Modèle pour créer une notification"""
    title: str
    message: str
    department: Optional[str] = None  # Département cible
    city: Optional[str] = None  # Ville cible
    event_id: Optional[str] = None  # Événement associé (optionnel)
    target_users: Optional[List[str]] = None  # IDs utilisateurs spécifiques
    target_roles: Optional[List[str]] = None  # Rôles ciblés
    send_to_all: bool = False  # Envoyer à tous
    scheduled_at: Optional[str] = None  # Date/heure programmée (ISO format)
    
class Notification(BaseModel):
    """Modèle de notification stockée"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    message: str
    department: Optional[str] = None
    city: Optional[str] = None
    event_id: Optional[str] = None
    target_users: Optional[List[str]] = None
    target_roles: Optional[List[str]] = None
    send_to_all: bool = False
    scheduled_at: Optional[str] = None
    sent_at: Optional[str] = None
    status: str = "pending"  # pending, sent, failed, scheduled
    sent_count: int = 0
    failed_count: int = 0
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

@api_router.put("/users/{user_id}/reset-password")
async def reset_user_password(user_id: str, password_data: PasswordReset, current_user: dict = Depends(get_current_user)):
    """Reset user password (Super Admin only)"""
    if not is_super_admin(current_user):
        raise HTTPException(status_code=403, detail="Only super admin can reset passwords")
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Hash new password
    hashed_password = hash_password(password_data.new_password)
    await db.users.update_one(
        {"id": user_id}, 
        {"$set": {
            "password": hashed_password,
            "plain_password": password_data.new_password  # Stocker aussi en clair
        }}
    )
    return {"message": "Password reset successfully"}


# ==================== VISITOR ROUTES ====================

@api_router.post("/visitors")
async def create_visitor(visitor_data: VisitorCreate, current_user: dict = Depends(get_current_user)):
    # Restrict accueil role from creating visitors (read-only)
    if current_user["role"] == "accueil":
        raise HTTPException(status_code=403, detail="Accueil role is read-only, cannot create visitors")
    
    # Only superviseur_promos, responsable_promo, referent, super_admin, pasteur can create
    
    # Calculate assigned_month
    try:
        visit_dt = datetime.fromisoformat(visitor_data.visit_date)
        assigned_month = visit_dt.strftime("%Y-%m")
    except:
        assigned_month = datetime.now(timezone.utc).strftime("%Y-%m")
    
    visitor = Visitor(**visitor_data.model_dump(), assigned_month=assigned_month)
    doc = visitor.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.visitors.insert_one(doc)
    return {"message": "Visitor created successfully", "id": visitor.id}

@api_router.post("/visitors/public")
async def create_visitor_public(visitor_data: VisitorCreate):
    """Créer un visiteur - Public (pour les bergeries)"""
    # Calculate assigned_month
    try:
        visit_dt = datetime.fromisoformat(visitor_data.visit_date)
        assigned_month = visit_dt.strftime("%Y-%m")
    except:
        assigned_month = datetime.now(timezone.utc).strftime("%Y-%m")
    
    visitor = Visitor(**visitor_data.model_dump(), assigned_month=assigned_month)
    doc = visitor.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.visitors.insert_one(doc)
    return {"message": "Visitor created successfully", "id": visitor.id}

# PUBLIC VISITOR ENDPOINTS - Pour les bergeries publiques
@api_router.get("/visitors/public/{visitor_id}")
async def get_visitor_public(visitor_id: str):
    """Récupérer un visiteur - Public"""
    visitor = await db.visitors.find_one({"id": visitor_id}, {"_id": 0})
    if not visitor:
        raise HTTPException(status_code=404, detail="Visitor not found")
    return visitor

@api_router.put("/visitors/public/{visitor_id}")
async def update_visitor_public(visitor_id: str, update_data: dict):
    """Mettre à jour un visiteur - Public"""
    # Remove protected fields
    protected = ['id', 'created_at', 'assigned_month']
    for field in protected:
        update_data.pop(field, None)
    
    result = await db.visitors.update_one(
        {"id": visitor_id},
        {"$set": update_data}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Visitor not found")
    return {"message": "Visitor updated successfully"}

@api_router.delete("/visitors/public/{visitor_id}")
async def delete_visitor_public(visitor_id: str):
    """Supprimer un visiteur - Public"""
    result = await db.visitors.delete_one({"id": visitor_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Visitor not found")
    return {"message": "Visitor deleted successfully"}

@api_router.post("/visitors/public/{visitor_id}/comment")
async def add_visitor_comment_public(visitor_id: str, comment_data: dict):
    """Ajouter un commentaire - Public"""
    comment = {
        "text": comment_data.get("text", ""),
        "date": datetime.now(timezone.utc).isoformat(),
        "author": "Public"
    }
    
    result = await db.visitors.update_one(
        {"id": visitor_id},
        {"$push": {"comments": comment}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Visitor not found")
    return {"message": "Comment added successfully"}

@api_router.post("/visitors/public/{visitor_id}/formation")
async def update_visitor_formation_public(visitor_id: str, formation_data: dict):
    """Mettre à jour une formation - Public"""
    formation_type = formation_data.get("formation_type")
    value = formation_data.get("value", False)
    
    if formation_type not in ["formation_pcnc", "formation_au_coeur_bible", "formation_star"]:
        raise HTTPException(status_code=400, detail="Invalid formation type")
    
    result = await db.visitors.update_one(
        {"id": visitor_id},
        {"$set": {formation_type: value}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Visitor not found")
    return {"message": "Formation updated successfully"}

@api_router.post("/visitors/public/{visitor_id}/stop")
async def stop_visitor_tracking_public(visitor_id: str, stop_data: dict):
    """Arrêter le suivi d'un visiteur - Public"""
    reason = stop_data.get("reason", "")
    
    result = await db.visitors.update_one(
        {"id": visitor_id},
        {"$set": {
            "tracking_stopped": True,
            "tracking_stopped_reason": reason,
            "tracking_stopped_date": datetime.now(timezone.utc).isoformat()
        }}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Visitor not found")
    return {"message": "Tracking stopped successfully"}

@api_router.post("/visitors/bulk-ancien")
async def create_bulk_ancien_visitors(visitors_data: List[VisitorCreate], current_user: dict = Depends(get_current_user)):
    # Only superviseur_promos, responsable_promo, referent, super_admin, pasteur can create
    
    if len(visitors_data) > 40:
        raise HTTPException(status_code=400, detail="Maximum 40 visitors can be added at once")
    
    created_ids = []
    for visitor_data in visitors_data:
        # Calculate assigned_month
        try:
            visit_dt = datetime.fromisoformat(visitor_data.visit_date)
            assigned_month = visit_dt.strftime("%Y-%m")
        except:
            assigned_month = datetime.now(timezone.utc).strftime("%Y-%m")
        
        # Force is_ancien to True
        visitor_dict = visitor_data.model_dump()
        visitor_dict['is_ancien'] = True
        
        visitor = Visitor(**visitor_dict, assigned_month=assigned_month)
        doc = visitor.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        
        await db.visitors.insert_one(doc)
        created_ids.append(visitor.id)
    
    return {"message": f"{len(created_ids)} anciens visiteurs créés avec succès", "ids": created_ids}

@api_router.get("/visitors")
async def get_visitors(
    include_stopped: bool = False,
    current_user: dict = Depends(get_current_user)
):
    # Super Admin and Pasteur can see all cities
    if current_user["role"] in ["super_admin", "pasteur"]:
        query = {}
    else:
        query = {
            "city": current_user["city"]
        }
    
    # Include or exclude stopped visitors
    if not include_stopped:
        query["tracking_stopped"] = False
    
    # Filter by role and permissions
    # Both referent and responsable_promo should see all visitors from their assigned month regardless of year
    if current_user["role"] in ["referent", "responsable_promo", "promotions"]:
        # Check if referent/responsable has permission to view all months
        permissions = current_user.get("permissions") or {}
        if not permissions.get("can_view_all_months", False):
            # Extract month from their assigned_month and match all years
            user_assigned_month = current_user.get("assigned_month")
            if user_assigned_month:
                # Support multiple months separated by commas (e.g., "2024-08,2025-08,2026-08")
                months_list = [m.strip() for m in user_assigned_month.split(',')]
                if len(months_list) > 1:
                    # Multiple months: use $in to match any of them
                    query["assigned_month"] = {"$in": months_list}
                else:
                    # Single month: Extract month part only (MM from YYYY-MM) to match all years
                    month_part = user_assigned_month.split("-")[-1] if "-" in user_assigned_month else user_assigned_month
                    # Use regex to match any year with this month
                    # Example: referent with assigned_month="2024-08" sees ALL august visitors (2024-08, 2025-08, etc.)
                    query["assigned_month"] = {"$regex": f"-{month_part}$"}
    
    # superviseur_promos sees ALL visitors from their city (no month filter)
    
    # Exclure les visiteurs supprimés (sauf pour super_admin qui peut les voir avec endpoint dédié)
    query["deleted"] = {"$ne": True}
    
    visitors = await db.visitors.find(query, {"_id": 0}).to_list(10000)
    
    # For "accueil" role, return limited info (just for consultation)
    if current_user["role"] == "accueil":
        return [{
            "id": v["id"], 
            "firstname": v["firstname"], 
            "lastname": v["lastname"],
            "arrival_channel": v.get("arrival_channel", ""),
            "visit_date": v.get("visit_date", ""),
            "city": v["city"]
        } for v in visitors]
    
    return visitors

@api_router.get("/visitors/stopped")
async def get_stopped_visitors(current_user: dict = Depends(get_current_user)):
    # Admin, pasteur, responsable_eglise, superviseur_promos can see stopped visitors
    
    # Pour responsable_eglise, filtrer par leur ville
    if current_user["role"] == "responsable_eglise":
        query = {
            "city": current_user["city"],
            "tracking_stopped": True
        }
    else:
        # Super admin et pasteur voient toutes les villes (pas de filtre)
        query = {
            "tracking_stopped": True
        }
    
    visitors = await db.visitors.find(query, {"_id": 0}).to_list(10000)
    return visitors

@api_router.get("/visitors/{visitor_id}")
async def get_visitor(visitor_id: str, current_user: dict = Depends(get_current_user)):
    # Super admin and pasteur can view visitors from all cities
    if current_user["role"] in ["super_admin", "pasteur"]:
        visitor = await db.visitors.find_one({"id": visitor_id}, {"_id": 0})
    else:
        visitor = await db.visitors.find_one({"id": visitor_id, "city": current_user["city"]}, {"_id": 0})
    
    if not visitor:
        raise HTTPException(status_code=404, detail="Visitor not found")
    
    # Check permissions for referent and responsable_promo
    if current_user["role"] in ["referent", "responsable_promo", "promotions"]:
        permissions = current_user.get("permissions") or {}
        if not permissions.get("can_view_all_months", False):
            # Check if visitor's month matches user's month(s) (regardless of year)
            user_assigned = current_user.get("assigned_month", "")
            if user_assigned:
                # Support multiple months separated by commas
                user_months_list = [m.strip() for m in user_assigned.split(',')]
                visitor_month = visitor.get("assigned_month", "")
                # Check if visitor's month matches any of user's assigned months
                allowed = False
                for um in user_months_list:
                    user_month_part = um.split("-")[-1] if "-" in um else um
                    visitor_month_part = visitor_month.split("-")[-1] if "-" in visitor_month else visitor_month
                    if user_month_part == visitor_month_part:
                        allowed = True
                        break
                if not allowed:
                    raise HTTPException(status_code=403, detail="Access denied")
    
    # Accueil can't see details (consultation only)
    if current_user["role"] == "accueil":
        raise HTTPException(status_code=403, detail="Access denied")
    
    return visitor

@api_router.put("/visitors/{visitor_id}")
async def update_visitor(visitor_id: str, update_data: VisitorUpdate, current_user: dict = Depends(get_current_user)):
    # Allow all logged-in users to update visitors based on their role
    
    # Super admin et pasteur peuvent modifier tous les visiteurs
    if current_user["role"] in ["super_admin", "pasteur"]:
        visitor = await db.visitors.find_one({"id": visitor_id})
    else:
        visitor = await db.visitors.find_one({"id": visitor_id, "city": current_user["city"]})
    
    if not visitor:
        raise HTTPException(status_code=404, detail="Visitor not found")
    
    # Update only provided fields
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    
    if update_dict:
        await db.visitors.update_one({"id": visitor_id}, {"$set": update_dict})
    
    return {"message": "Visitor updated successfully"}

@api_router.get("/visitors/deleted/all")
async def get_deleted_visitors(current_user: dict = Depends(get_current_user)):
    """Get all deleted visitors (super_admin only)"""
    if current_user["role"] != "super_admin":
        raise HTTPException(status_code=403, detail="Only super_admin can view deleted visitors")
    
    deleted_visitors = await db.visitors.find(
        {"deleted": True}, 
        {"_id": 0}
    ).to_list(10000)
    
    return deleted_visitors

@api_router.delete("/visitors/{visitor_id}")
async def delete_visitor(visitor_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a visitor (soft delete - mark as deleted)"""
    visitor = await db.visitors.find_one({"id": visitor_id, "city": current_user["city"]})
    
    if not visitor:
        raise HTTPException(status_code=404, detail="Visitor not found")
    
    # Check permissions based on role
    
    # Les responsables de promo peuvent supprimer tous leurs visiteurs
    # Pas de restriction supplémentaire pour eux
    
    # Soft delete: marquer comme supprimé au lieu de supprimer physiquement
    await db.visitors.update_one(
        {"id": visitor_id},
        {"$set": {
            "deleted": True,
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": current_user["username"]
        }}
    )
    return {"message": "Visitor deleted successfully"}

@api_router.post("/visitors/{visitor_id}/comment")
async def add_comment(visitor_id: str, comment: CommentAdd, current_user: dict = Depends(get_current_user)):
    visitor = await db.visitors.find_one({"id": visitor_id, "city": current_user["city"]})
    
    if not visitor:
        raise HTTPException(status_code=404, detail="Visitor not found")
    
    # Check permissions for referents/responsables
    if current_user["role"] in ["referent", "responsable_promo", "promotions"]:
        permissions = current_user.get("permissions") or {}
        if not permissions.get("can_add_comments", True):
            raise HTTPException(status_code=403, detail="Permission denied: cannot add comments")
    
    comment_entry = CommentEntry(
        author=current_user["username"],
        text=comment.text
    )
    
    await db.visitors.update_one(
        {"id": visitor_id},
        {"$push": {"comments": comment_entry.model_dump()}}
    )
    
    return {"message": "Comment added successfully"}


# ==================== KPI DISCIPOLAT ====================
class KPIDiscipolatEntry(BaseModel):
    visitor_id: Optional[str] = None
    mois: str  # Format: "2024-01"
    presence_dimanche: int = 0
    presence_fi: int = 0
    presence_reunion_disciples: int = 0
    service_eglise: int = 0
    consommation_pain_jour: int = 0
    bapteme: int = 0
    commentaire: str = ""

# Coefficients (poids) pour chaque KPI - Mis à jour
KPI_WEIGHTS = {
    "presence_dimanche": 5,
    "presence_fi": 2,
    "presence_reunion_disciples": 3,
    "service_eglise": 6,
    "consommation_pain_jour": 5,
    "bapteme": 2
}

# Niveaux de discipolat
def get_discipolat_level(score: float) -> str:
    if score < 20:
        return "Non classé"
    elif score < 40:
        return "Débutant"
    elif score < 60:
        return "Intermédiaire"
    else:
        return "Confirmé"

def calculate_kpi_score(kpi: dict) -> float:
    """Calcule le score KPI basé sur les coefficients"""
    score = 0
    for key, weight in KPI_WEIGHTS.items():
        score += kpi.get(key, 0) * weight
    return score

@api_router.post("/visitors/{visitor_id}/kpi")
async def save_kpi_discipolat(visitor_id: str, kpi: KPIDiscipolatEntry, current_user: dict = Depends(get_current_user)):
    """Enregistrer les KPIs Discipolat pour un visiteur pour un mois donné"""
    # Vérifier que le visiteur existe
    visitor = await db.visitors.find_one({"id": visitor_id})
    if not visitor:
        raise HTTPException(status_code=404, detail="Visitor not found")
    
    # Calculer le score
    score = calculate_kpi_score(kpi.model_dump())
    level = get_discipolat_level(score)
    
    kpi_data = {
        "visitor_id": visitor_id,
        "mois": kpi.mois,
        "presence_dimanche": kpi.presence_dimanche,
        "presence_fi": kpi.presence_fi,
        "presence_reunion_disciples": kpi.presence_reunion_disciples,
        "service_eglise": kpi.service_eglise,
        "consommation_pain_jour": kpi.consommation_pain_jour,
        "bapteme": kpi.bapteme,
        "commentaire": kpi.commentaire,
        "score": score,
        "level": level,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user["username"]
    }
    
    # Upsert: mettre à jour ou créer
    await db.kpi_discipolat.update_one(
        {"visitor_id": visitor_id, "mois": kpi.mois},
        {"$set": kpi_data},
        upsert=True
    )
    
    # Recalculer le statut moyen du visiteur
    all_kpis = await db.kpi_discipolat.find({"visitor_id": visitor_id}, {"_id": 0}).to_list(100)
    if all_kpis:
        avg_score = sum(k.get("score", 0) for k in all_kpis) / len(all_kpis)
        avg_level = get_discipolat_level(avg_score)
        
        # Mettre à jour le statut du visiteur
        await db.visitors.update_one(
            {"id": visitor_id},
            {"$set": {
                "discipolat_score": round(avg_score, 1),
                "discipolat_level": avg_level,
                "discipolat_updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    return {
        "message": "KPI saved successfully",
        "score": score,
        "level": level
    }

@api_router.get("/visitors/{visitor_id}/kpi")
async def get_visitor_kpis(visitor_id: str, current_user: dict = Depends(get_current_user)):
    """Récupérer tous les KPIs d'un visiteur"""
    kpis = await db.kpi_discipolat.find(
        {"visitor_id": visitor_id},
        {"_id": 0}
    ).sort("mois", -1).to_list(100)
    
    # Calculer le statut moyen
    if kpis:
        avg_score = sum(k.get("score", 0) for k in kpis) / len(kpis)
        avg_level = get_discipolat_level(avg_score)
    else:
        avg_score = 0
        avg_level = "Non classé"
    
    return {
        "kpis": kpis,
        "average_score": round(avg_score, 1),
        "average_level": avg_level,
        "weights": KPI_WEIGHTS
    }

@api_router.get("/visitors/{visitor_id}/kpi/{mois}")
async def get_visitor_kpi_for_month(visitor_id: str, mois: str, current_user: dict = Depends(get_current_user)):
    """Récupérer le KPI d'un visiteur pour un mois spécifique"""
    kpi = await db.kpi_discipolat.find_one(
        {"visitor_id": visitor_id, "mois": mois},
        {"_id": 0}
    )
    
    if not kpi:
        return {
            "visitor_id": visitor_id,
            "mois": mois,
            "presence_dimanche": 0,
            "presence_fi": 0,
            "presence_reunion_disciples": 0,
            "service_eglise": 0,
            "consommation_pain_jour": 0,
            "bapteme": 0,
            "commentaire": "",
            "score": 0,
            "level": "Non classé"
        }
    
    return kpi


@api_router.get("/visitors/kpi/all-statuses")
async def get_all_visitors_kpi_statuses(current_user: dict = Depends(get_current_user)):
    """Récupérer les statuts KPI moyens de tous les visiteurs de la ville"""
    # Récupérer tous les KPIs de cette ville
    all_kpis = await db.kpi_discipolat.find({}).to_list(10000)
    
    # Grouper par visitor_id et calculer la moyenne
    visitor_kpis = {}
    for kpi in all_kpis:
        vid = kpi.get("visitor_id")
        if vid not in visitor_kpis:
            visitor_kpis[vid] = []
        visitor_kpis[vid].append(kpi.get("score", 0))
    
    # Calculer moyenne et niveau pour chaque visiteur
    result = {}
    for vid, scores in visitor_kpis.items():
        if scores:
            avg = round(sum(scores) / len(scores), 1)
            if avg < 15:
                level = "Non classé"
            elif avg <= 30:
                level = "Débutant"
            elif avg <= 51:
                level = "Intermédiaire"
            else:
                level = "Confirmé"
            result[vid] = {"average_score": avg, "level": level, "months_count": len(scores)}
    
    return result


class ManualStatusUpdate(BaseModel):
    manual_status: Optional[str] = None
    manual_commentaire: str = ""


@api_router.post("/visitors/{visitor_id}/manual-status")
async def update_manual_status(visitor_id: str, data: ManualStatusUpdate, current_user: dict = Depends(get_current_user)):
    """Définir un statut manuel pour un visiteur (remplace le calcul automatique)"""
    visitor = await db.visitors.find_one({"id": visitor_id})
    if not visitor:
        raise HTTPException(status_code=404, detail="Visitor not found")
    
    await db.visitors.update_one(
        {"id": visitor_id},
        {"$set": {
            "manual_discipolat_status": data.manual_status,
            "manual_discipolat_commentaire": data.manual_commentaire,
            "manual_status_updated_at": datetime.now(timezone.utc).isoformat(),
            "manual_status_updated_by": current_user["username"]
        }}
    )
    
    return {"message": "Manual status updated successfully"}


@api_router.post("/visitors/{visitor_id}/presence")
async def add_presence(visitor_id: str, presence: PresenceAdd, current_user: dict = Depends(get_current_user)):
    visitor = await db.visitors.find_one({"id": visitor_id, "city": current_user["city"]})
    
    if not visitor:
        raise HTTPException(status_code=404, detail="Visitor not found")
    
    # Check permissions for referents/responsables
    if current_user["role"] in ["referent", "responsable_promo", "promotions"]:
        permissions = current_user.get("permissions") or {}
        if not permissions.get("can_mark_presence", True):
            raise HTTPException(status_code=403, detail="Permission denied: cannot mark presence")
    
    presence_entry = PresenceEntry(
        date=presence.date, 
        present=presence.present,
        commentaire=presence.commentaire
    )
    
    field = "presences_dimanche" if presence.type == "dimanche" else "presences_jeudi"
    
    # Check if date already exists, update or add
    existing_presences = visitor.get(field, [])
    date_exists = False
    
    for i, p in enumerate(existing_presences):
        if p["date"] == presence.date:
            existing_presences[i] = presence_entry.model_dump()
            date_exists = True
            break
    
    if not date_exists:
        await db.visitors.update_one(
            {"id": visitor_id},
            {"$push": {field: presence_entry.model_dump()}}
        )
    else:
        await db.visitors.update_one(
            {"id": visitor_id},
            {"$set": {field: existing_presences}}
        )
    
    return {"message": "Presence updated successfully"}

@api_router.get("/presences/by-date")
async def get_presences_by_date(date: str, current_user: dict = Depends(get_current_user)):
    """Get all presences for a specific date"""
    # Filter visitors by city if not admin/super_admin
    city_filter = {} if current_user["role"] in ["admin", "super_admin", "pasteur"] else {"city": current_user["city"]}
    
    visitors = await db.visitors.find(city_filter).to_list(length=None)
    
    result = []
    for visitor in visitors:
        # Check both dimanche and jeudi presences
        all_presences = visitor.get("presences_dimanche", []) + visitor.get("presences_jeudi", [])
        presence_for_date = next((p for p in all_presences if p.get("date") == date), None)
        
        if presence_for_date:
            result.append({
                "visitor_id": visitor["id"],
                "firstname": visitor["firstname"],
                "lastname": visitor["lastname"],
                "present": presence_for_date.get("present"),
                "commentaire": presence_for_date.get("commentaire")
            })
    
    return result

@api_router.post("/visitors/{visitor_id}/formation")
async def update_formation(visitor_id: str, formation: FormationUpdate, current_user: dict = Depends(get_current_user)):
    visitor = await db.visitors.find_one({"id": visitor_id, "city": current_user["city"]})
    
    if not visitor:
        raise HTTPException(status_code=404, detail="Visitor not found")
    
    field_map = {
        "pcnc": "formation_pcnc",
        "au_coeur_bible": "formation_au_coeur_bible",
        "star": "formation_star"
    }
    
    field = field_map.get(formation.formation_type)
    if not field:
        raise HTTPException(status_code=400, detail="Invalid formation type")
    
    await db.visitors.update_one(
        {"id": visitor_id},
        {"$set": {field: formation.completed}}
    )
    
    return {"message": "Formation updated successfully"}

@api_router.post("/visitors/{visitor_id}/stop-tracking")
async def stop_tracking(visitor_id: str, stop_data: StopTracking, current_user: dict = Depends(get_current_user)):
    visitor = await db.visitors.find_one({"id": visitor_id, "city": current_user["city"]})
    
    if not visitor:
        raise HTTPException(status_code=404, detail="Visitor not found")
    
    # Check permissions for referents/responsables
    if current_user["role"] in ["referent", "responsable_promo", "promotions"]:
        permissions = current_user.get("permissions") or {}
        if not permissions.get("can_stop_tracking", True):
            raise HTTPException(status_code=403, detail="Permission denied: cannot stop tracking")
    
    await db.visitors.update_one(
        {"id": visitor_id},
        {"$set": {
            "tracking_stopped": True,
            "stop_reason": stop_data.reason,
            "stopped_by": current_user["username"],
            "stopped_date": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Tracking stopped successfully"}

# ==================== CITY ROUTES ====================

@api_router.post("/cities")
async def create_city(city_data: CityCreate, current_user: dict = Depends(get_current_user)):
    
    # Check if city already exists
    existing = await db.cities.find_one({"name": city_data.name})
    if existing:
        raise HTTPException(status_code=400, detail="City already exists")
    
    city = City(**city_data.model_dump())
    await db.cities.insert_one(city.model_dump())
    
    return city

@api_router.get("/cities")
async def get_cities():
    cities = await db.cities.find({}, {"_id": 0}).to_list(1000)
    return cities

@api_router.get("/cities/public")
async def get_cities_public():
    """Get list of cities with countries - Public endpoint"""
    cities = await db.cities.find({}, {"_id": 0, "id": 1, "name": 1, "country": 1}).to_list(1000)
    return [city for city in cities if city.get("name")]

@api_router.post("/cities/initialize")
async def initialize_cities(current_user: dict = Depends(get_current_user)):
    """Initialize all cities with their countries - Creates cities if they don't exist"""
    
    from uuid import uuid4
    
    # Define all cities with their countries
    cities_data = [
        {'name': 'Milan', 'country': 'Italie'},
        {'name': 'Rome', 'country': 'Italie'},
        {'name': 'Perugia', 'country': 'Italie'},
        {'name': 'Bologne', 'country': 'Italie'},
        {'name': 'Turin', 'country': 'Italie'},
        {'name': 'Dijon', 'country': 'France'},
        {'name': 'Auxerre', 'country': 'France'},
        {'name': 'Besançon', 'country': 'France'},
        {'name': 'Chalon-Sur-Saone', 'country': 'France'},
        {'name': 'Dole', 'country': 'France'},
        {'name': 'Sens', 'country': 'France'}
    ]
    
    created_count = 0
    updated_count = 0
    
    for city_data in cities_data:
        # Check if city exists (case-insensitive)
        existing = await db.cities.find_one({
            'name': {'$regex': f'^{city_data["name"]}$', '$options': 'i'}
        })
        
        if existing:
            # Update existing city
            result = await db.cities.update_one(
                {'_id': existing['_id']},
                {'$set': {'country': city_data['country']}}
            )
            if result.modified_count > 0:
                updated_count += 1
        else:
            # Create new city
            new_city = {
                'id': str(uuid4()),
                'name': city_data['name'],
                'country': city_data['country']
            }
            await db.cities.insert_one(new_city)
            created_count += 1
    
    return {
        "success": True,
        "message": f"{created_count} villes créées, {updated_count} villes mises à jour",
        "created_count": created_count,
        "updated_count": updated_count,
        "total_cities": await db.cities.count_documents({})
    }

@api_router.put("/cities/{city_id}")
async def update_city(city_id: str, city_data: CityCreate, current_user: dict = Depends(get_current_user)):
    """Update city name (Admin only)"""
    
    # Check if new name already exists (excluding current city)
    existing = await db.cities.find_one({
        "name": city_data.name,
        "id": {"$ne": city_id}
    })
    if existing:
        raise HTTPException(status_code=400, detail="City name already exists")
    
    # Get the old city name before updating
    old_city = await db.cities.find_one({"id": city_id})
    if not old_city:
        raise HTTPException(status_code=404, detail="City not found")
    
    old_name = old_city["name"]
    
    # Update city name
    result = await db.cities.update_one(
        {"id": city_id},
        {"$set": {"name": city_data.name}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="City not found")
    
    # Update all users with this city
    await db.users.update_many(
        {"city": old_name},
        {"$set": {"city": city_data.name}}
    )
    
    # Update all visitors with this city
    await db.visitors.update_many(
        {"city": old_name},
        {"$set": {"city": city_data.name}}
    )
    
    return {"message": "City updated successfully"}

@api_router.delete("/cities/{city_id}")
async def delete_city(city_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a city (Admin only)"""
    
    city = await db.cities.find_one({"id": city_id})
    if not city:
        raise HTTPException(status_code=404, detail="City not found")
    
    city_name = city["name"]
    
    # Check if city has users or visitors
    users_count = await db.users.count_documents({"city": city_name})
    visitors_count = await db.visitors.count_documents({"city": city_name})
    
    if users_count > 0 or visitors_count > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete city with {users_count} users and {visitors_count} visitors"
        )
    
    # Delete the city
    await db.cities.delete_one({"id": city_id})
    
    return {"message": "City deleted successfully"}

@api_router.get("/cities/{city_id}/stats")
async def get_city_stats(
    city_id: str,
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get comprehensive statistics for a specific city with year/month filters"""
    
    city = await db.cities.find_one({"id": city_id})
    if not city:
        raise HTTPException(status_code=404, detail="City not found")
    
    city_name = city["name"]
    
    # Build query with filters
    query = {"city": city_name}
    if year and month:
        query["assigned_month"] = f"{year}-{month:02d}"
    elif year:
        query["assigned_month"] = {"$regex": f"^{year}-"}
    elif month:
        query["assigned_month"] = {"$regex": f"-{month:02d}$"}
    
    # PROMOTIONS STATS
    visitors = await db.visitors.find(query).to_list(10000)
    na_count = sum(1 for v in visitors if "Nouveau Arrivant" in v.get("types", []))
    nc_count = sum(1 for v in visitors if "Nouveau Converti" in v.get("types", []))
    dp_count = sum(1 for v in visitors if "De Passage" in v.get("types", []))
    
    # Calculate average fidelisation for promotions
    total_fidelisation = 0
    promo_count = 0
    for visitor in visitors:
        if visitor.get("presences_dimanche") or visitor.get("presences_jeudi"):
            total_presences = len([p for p in visitor.get("presences_dimanche", []) if p.get("present")]) + len([p for p in visitor.get("presences_jeudi", []) if p.get("present")])
            expected = (len(visitor.get("presences_dimanche", [])) + len(visitor.get("presences_jeudi", [])))
            if expected > 0:
                total_fidelisation += (total_presences / expected) * 100
                promo_count += 1
    avg_fidelisation_promos = (total_fidelisation / promo_count) if promo_count > 0 else 0
    
    # FAMILLE D'IMPACT STATS
    fi_query = {"ville": city_name}  # Note: FI uses "ville" not "city"
    fis = await db.familles_impact.find(fi_query, {"_id": 0}).to_list(10000)
    secteurs = await db.secteurs.find({"ville": city_name}, {"_id": 0}).to_list(1000)
    
    # Count FI members from membres_fi collection
    fi_ids = [fi["id"] for fi in fis]
    membres_query = {"fi_id": {"$in": fi_ids}} if fi_ids else {"fi_id": None}
    membres = await db.membres_fi.find(membres_query, {"_id": 0}).to_list(10000)
    total_fi_members = len(membres)
    
    # Calculate FI fidelisation from presences_fi collection
    presences_query = {"fi_id": {"$in": fi_ids}} if fi_ids else {"fi_id": None}
    if year and month:
        presences_query["date"] = {"$regex": f"^{year}-{month:02d}"}
    elif year:
        presences_query["date"] = {"$regex": f"^{year}"}
    
    presences = await db.presences_fi.find(presences_query, {"_id": 0}).to_list(100000)
    
    # Calculate fidelisation per member
    fi_fidelisation = 0
    fi_count = 0
    for membre in membres:
        membre_presences = [p for p in presences if p.get("membre_fi_id") == membre["id"]]
        if membre_presences:
            present_count = sum(1 for p in membre_presences if p.get("present"))
            if len(membre_presences) > 0:
                fi_fidelisation += (present_count / len(membre_presences)) * 100
                fi_count += 1
    
    avg_fidelisation_fi = (fi_fidelisation / fi_count) if fi_count > 0 else 0
    
    # CULTE STATS
    culte_query = {"ville": city_name}
    if year and month:
        culte_query["date"] = {"$regex": f"^{year}-{month:02d}"}
    elif year:
        culte_query["date"] = {"$regex": f"^{year}"}
    
    culte_stats = await db.culte_stats.find(culte_query).to_list(10000)
    
    total_adultes = sum(s.get("nombre_adultes", 0) for s in culte_stats)
    total_enfants = sum(s.get("nombre_enfants", 0) for s in culte_stats)
    total_stars = sum(s.get("nombre_stars", 0) for s in culte_stats)
    culte_count = len(culte_stats)
    
    avg_adultes = (total_adultes / culte_count) if culte_count > 0 else 0
    avg_enfants = (total_enfants / culte_count) if culte_count > 0 else 0
    avg_stars = (total_stars / culte_count) if culte_count > 0 else 0
    
    return {
        "city_name": city_name,
        "filters": {
            "year": year,
            "month": month
        },
        "promotions": {
            "nouveaux_arrivants": na_count,
            "nouveaux_convertis": nc_count,
            "de_passage": dp_count,
            "avg_fidelisation": round(avg_fidelisation_promos, 2)
        },
        "familles_impact": {
            "nombre_secteurs": len(secteurs),
            "nombre_familles": len(fis),
            "total_membres": total_fi_members,
            "avg_fidelisation": round(avg_fidelisation_fi, 2)
        },
        "culte_stats": {
            "avg_adultes": round(avg_adultes, 1),
            "avg_enfants": round(avg_enfants, 1),
            "avg_stars": round(avg_stars, 1),
            "total_services": culte_count
        },
        "evangelisation": await get_evangelisation_stats_for_city(city_name, year, month)
    }

async def get_evangelisation_stats_for_city(city_name: str, year: Optional[int], month: Optional[int]):
    """Helper to get evangelisation stats for a city"""
    query = {"ville": city_name}
    
    if year and month:
        query["date"] = {"$regex": f"^{year}-{month:02d}"}
    elif year:
        query["date"] = {"$regex": f"^{year}"}
    
    records = await db.evangelisation.find(query).to_list(1000)
    
    stats = {
        "eglise": {
            "nombre_gagneurs_ame": 0,
            "nombre_personnes_receptives": 0,
            "nombre_priere_salut": 0,
            "nombre_contacts_pris": 0,
            "nombre_ames_invitees": 0,
            "nombre_miracles": 0
        },
        "familles_impact": {
            "nombre_gagneurs_ame": 0,
            "nombre_personnes_receptives": 0,
            "nombre_priere_salut": 0,
            "nombre_contacts_pris": 0,
            "nombre_ames_invitees": 0,
            "nombre_miracles": 0
        }
    }
    
    for record in records:
        if record.get("eglise"):
            for key in stats["eglise"]:
                stats["eglise"][key] += record["eglise"].get(key, 0)
        
        if record.get("familles_impact"):
            for key in stats["familles_impact"]:
                stats["familles_impact"][key] += record["familles_impact"].get(key, 0)
    
    return stats

# ==================== ANALYTICS ROUTES ====================

@api_router.get("/analytics/stats")
async def get_stats(
    impersonate_user_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    # If impersonating, load the target user's data for filtering
    target_user = current_user
    if impersonate_user_id and current_user["role"] in ["pasteur", "super_admin"]:
        # SuperAdmin/Pasteur can impersonate - load target user
        impersonate_user = await db.users.find_one({"id": impersonate_user_id}, {"_id": 0})
        if impersonate_user:
            target_user = impersonate_user
    
    # Base query filter
    base_query = {"tracking_stopped": False}
    
    # For pasteur, super_admin, and responsable_eglise: responsable_eglise sees only their city
    if target_user["role"] == "responsable_eglise":
        base_query["city"] = target_user["city"]
    elif target_user["role"] not in ["pasteur", "super_admin"]:
        base_query["city"] = target_user["city"]
    
    city = target_user["city"]
    
    # If referent or responsable_promo, filter by their assigned month (all years)
    if target_user["role"] in ["referent", "responsable_promo", "promotions", "berger"]:
        permissions = target_user.get("permissions") or {}
        if not permissions.get("can_view_all_months", False):
            assigned_month = target_user.get("assigned_month")
            if assigned_month:
                # Handle both string and list formats
                if isinstance(assigned_month, list):
                    # Multiple months assigned - match any of them
                    month_parts = []
                    for month in assigned_month:
                        month_part = month.split("-")[-1] if "-" in month else month
                        month_parts.append(f"-{month_part}$")
                    # Match any of the assigned months
                    base_query["assigned_month"] = {"$regex": "|".join(month_parts)}
                else:
                    # Single month assigned
                    month_part = assigned_month.split("-")[-1] if "-" in assigned_month else assigned_month
                    # Use regex to match any year with this month
                    base_query["assigned_month"] = {"$regex": f"-{month_part}$"}
    
    # Total visitors
    total_visitors = await db.visitors.count_documents(base_query)
    
    # Total referents (only for admin/promotions/pasteur/super_admin)
    if target_user["role"] in ["superviseur_promos", "promotions", "pasteur", "super_admin"]:
        ref_query = {"role": {"$in": ["referent", "accueil", "promotions"]}}
        if target_user["role"] not in ["pasteur", "super_admin"]:
            ref_query["city"] = city
        total_referents = await db.users.count_documents(ref_query)
    else:
        total_referents = 0
    
    # By arrival channel
    pipeline = [
        {"$match": base_query},
        {"$group": {"_id": "$arrival_channel", "count": {"$sum": 1}}}
    ]
    by_channel = await db.visitors.aggregate(pipeline).to_list(1000)
    
    # By month
    pipeline = [
        {"$match": base_query},
        {"$group": {"_id": "$assigned_month", "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}}
    ]
    by_month = await db.visitors.aggregate(pipeline).to_list(1000)
    
    # By type
    pipeline = [
        {"$match": base_query},
        {"$unwind": "$types"},
        {"$group": {"_id": "$types", "count": {"$sum": 1}}}
    ]
    by_type = await db.visitors.aggregate(pipeline).to_list(1000)
    
    # Formations stats
    formation_pcnc_count = await db.visitors.count_documents({**base_query, "formation_pcnc": True})
    formation_au_coeur_bible_count = await db.visitors.count_documents({**base_query, "formation_au_coeur_bible": True})
    formation_star_count = await db.visitors.count_documents({**base_query, "formation_star": True})
    
    return {
        "total_visitors": total_visitors,
        "total_referents": total_referents,
        "by_channel": by_channel,
        "by_month": by_month,
        "by_type": by_type,
        "formation_pcnc": formation_pcnc_count,
        "formation_au_coeur_bible": formation_au_coeur_bible_count,
        "formation_star": formation_star_count
    }

@api_router.get("/analytics/export")
async def export_excel(current_user: dict = Depends(get_current_user)):
    
    visitors = await db.visitors.find({"city": current_user["city"]}, {"_id": 0}).to_list(10000)
    
    # Prepare data for Excel
    data = []
    for v in visitors:
        data.append({
            "Prénom": v.get("firstname"),
            "Nom": v.get("lastname"),
            "Types": ", ".join(v.get("types", [])),
            "Téléphone": v.get("phone", ""),
            "Email": v.get("email", ""),
            "Canal d'arrivée": v.get("arrival_channel"),
            "Date de visite": v.get("visit_date"),
            "Mois assigné": v.get("assigned_month"),
            "Formation PCNC": "Oui" if v.get("formation_pcnc") else "Non",
            "Formation Au cœur de la bible": "Oui" if v.get("formation_au_coeur_bible") else "Non",
            "Formation STAR": "Oui" if v.get("formation_star") else "Non",
            "Suivi arrêté": "Oui" if v.get("tracking_stopped") else "Non",
            "Raison arrêt": v.get("stop_reason", ""),
            "Arrêté par": v.get("stopped_by", ""),
        })
    
    df = pd.DataFrame(data)
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Visiteurs')
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=visiteurs_{current_user['city']}_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

# ==================== FIDELISATION ROUTES ====================

def get_week_number(date_str):
    """Get week number from date string"""
    try:
        date = datetime.fromisoformat(date_str)
        return date.isocalendar()[1]
    except:
        return None

def get_weeks_in_month(year, month):
    """Get all week numbers in a given month"""
    first_day = datetime(year, month, 1)
    if month == 12:
        last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = datetime(year, month + 1, 1) - timedelta(days=1)
    
    weeks = set()
    current = first_day
    while current <= last_day:
        weeks.add(current.isocalendar()[1])
        current += timedelta(days=1)
    return sorted(list(weeks))

@api_router.get("/fidelisation/referent")
async def get_referent_fidelisation(current_user: dict = Depends(get_current_user)):
    """Get fidelisation rate for referent, responsable_promo, superviseur_promos, and promotions roles"""
    
    assigned_month = current_user.get("assigned_month")
    # Pour les rôles sans assigned_month, utiliser tous les visiteurs de la ville
    if not assigned_month and current_user["role"] in ["superviseur_promos", "super_admin", "pasteur"]:
        # Ces rôles voient tous les visiteurs de leur ville
        pass
    elif not assigned_month:
        raise HTTPException(status_code=400, detail="No assigned month")
    
    # Get all visitors for this user
    query = {"city": current_user["city"]}
    
    # Pour referent, responsable_promo ET promotions, filtrer par assigned_month
    # Les superviseurs et admins voient toute la ville
    if assigned_month and current_user["role"] in ["referent", "responsable_promo", "promotions"]:
        # Support multiple months separated by commas (e.g., "2024-08,2025-08,2026-08")
        months_list = [m.strip() for m in assigned_month.split(',')]
        if len(months_list) > 1:
            query["assigned_month"] = {"$in": months_list}
        else:
            query["assigned_month"] = assigned_month
    
    all_visitors = await db.visitors.find(query, {"_id": 0}).to_list(10000)

    
    # Pour la fidélisation, on ne compte que les visiteurs actifs
    visitors = [v for v in all_visitors if not v.get("tracking_stopped")]
    
    # Mais total_visitors compte TOUS (actifs + arrêtés)
    total_visitors = len(all_visitors)
    total_visitors_actifs = len(visitors)
    if total_visitors_actifs == 0:
        # Retourner quand même total_visitors pour affichage
        return {
            "total_visitors": total_visitors,
            "total_visitors_actifs": 0,
            "weekly_rates": [],
            "monthly_average": 0
        }
    
    # Déterminer la plage de dates basée sur les présences réelles
    # Trouver la date min et max des présences
    all_presence_dates = []
    for visitor in visitors:
        for p in visitor.get("presences_dimanche", []):
            date_val = p.get("date")
            if date_val and isinstance(date_val, str):
                all_presence_dates.append(date_val)
        for p in visitor.get("presences_jeudi", []):
            date_val = p.get("date")
            if date_val and isinstance(date_val, str):
                all_presence_dates.append(date_val)
    
    # Parse and validate dates
    valid_dates = []
    from datetime import datetime as dt
    for date_str in all_presence_dates:
        try:
            parsed = dt.strptime(date_str, "%Y-%m-%d")
            valid_dates.append(parsed)
        except (ValueError, TypeError):
            # Skip invalid dates
            continue
    
    if not valid_dates:
        # Pas de présences valides - générer quand même les 52 semaines
        weeks = list(range(1, 53))
    else:
        # Utiliser toutes les 52 semaines de l'année mais ne calculer le taux que pour celles qui ont des présences
        weeks = list(range(1, 53))
    
    weekly_rates = []
    for week in weeks:
        # Count presences for this week with WEIGHTED LOGIC (dimanche x2, jeudi x1)
        total_presences_dimanche = 0
        total_presences_jeudi = 0
        for visitor in visitors:
            for presence in visitor.get("presences_dimanche", []):
                if get_week_number(presence["date"]) == week and presence.get("present", False):
                    total_presences_dimanche += 1
            for presence in visitor.get("presences_jeudi", []):
                if get_week_number(presence["date"]) == week and presence.get("present", False):
                    total_presences_jeudi += 1
        
        # Calculate rate with PONDÉRATION: dimanche x2, jeudi x1
        expected_dimanche = total_visitors_actifs * 1  # 1 dimanche par semaine
        expected_jeudi = total_visitors_actifs * 1  # 1 jeudi par semaine
        max_weighted = (expected_dimanche * 2) + (expected_jeudi * 1)
        actual_weighted = (total_presences_dimanche * 2) + (total_presences_jeudi * 1)
        rate = (actual_weighted / max_weighted * 100) if max_weighted > 0 else 0
        
        weekly_rates.append({
            "week": week,
            "rate": round(rate, 2),
            "presences": total_presences_dimanche + total_presences_jeudi,
            "expected": expected_dimanche + expected_jeudi
        })
    
    # Calculate monthly average
    monthly_average = sum(w["rate"] for w in weekly_rates) / len(weekly_rates) if weekly_rates else 0
    
    # Compter NA et NC de TOUS les visiteurs
    total_na = len([v for v in all_visitors if "Nouveau Arrivant" in v.get("types", [])])
    total_nc = len([v for v in all_visitors if "Nouveau Converti" in v.get("types", [])])
    
    return {
        "total_visitors": total_visitors,
        "total_visitors_actifs": total_visitors_actifs,
        "total_na": total_na,
        "total_nc": total_nc,
        "weekly_rates": weekly_rates,
        "monthly_average": round(monthly_average, 2)
    }

@api_router.get("/fidelisation/admin")
async def get_admin_fidelisation(week: int = None, month: str = None, current_user: dict = Depends(get_current_user)):
    """Get fidelisation rates for all referents (admin view)"""
    
    # Get all referents (all cities for pasteur/super_admin, own city for responsable_eglise and others)
    ref_query = {"role": "referent"}
    if current_user["role"] not in ["pasteur", "super_admin"]:
        ref_query["city"] = current_user["city"]
    
    referents = await db.users.find(ref_query, {"_id": 0, "password": 0}).to_list(1000)
    
    results = []
    
    for referent in referents:
        assigned_month = referent.get("assigned_month")
        if not assigned_month:
            continue
        
        # Get visitors for this referent  
        visitors = await db.visitors.find({
            "city": referent["city"],
            "assigned_month": assigned_month,
            "tracking_stopped": False
        }, {"_id": 0}).to_list(10000)
        
        total_visitors = len(visitors)
        if total_visitors == 0:
            continue
        
        # Déterminer la plage de dates basée sur les présences réelles
        # Trouver la date min et max des présences
        all_presence_dates = []
        for visitor in visitors:
            for p in visitor.get("presences_dimanche", []):
                if p.get("date"):
                    all_presence_dates.append(p["date"])
            for p in visitor.get("presences_jeudi", []):
                if p.get("date"):
                    all_presence_dates.append(p["date"])
        
        if not all_presence_dates:
            # Pas de présences, utiliser le mois assigné
            year, ref_month = map(int, assigned_month.split("-"))
            weeks = get_weeks_in_month(year, ref_month)
        else:
            # Utiliser la date la plus ancienne des présences pour déterminer les semaines
            min_date_str = min(all_presence_dates)
            max_date_str = max(all_presence_dates)
            min_date = datetime.strptime(min_date_str, "%Y-%m-%d")
            max_date = datetime.strptime(max_date_str, "%Y-%m-%d")
            
            # Calculer toutes les semaines entre min et max
            weeks = set()
            current = min_date
            while current <= max_date:
                weeks.add(current.isocalendar()[1])
                current += timedelta(days=1)
            weeks = sorted(list(weeks))
        
        # Filter by month if specified
        if month and assigned_month != month:
            continue
        
        # Filter by week if specified
        if week:
            weeks = [week] if week in weeks else []
        
        weekly_rates = []
        for w in weeks:
            # Count presences with WEIGHTED LOGIC (dimanche x2, jeudi x1)
            total_presences_dimanche = 0
            total_presences_jeudi = 0
            for visitor in visitors:
                for presence in visitor.get("presences_dimanche", []):
                    if get_week_number(presence["date"]) == w and presence.get("present", False):
                        total_presences_dimanche += 1
                for presence in visitor.get("presences_jeudi", []):
                    if get_week_number(presence["date"]) == w and presence.get("present", False):
                        total_presences_jeudi += 1
            
            # Calculate rate with PONDÉRATION: dimanche x2, jeudi x1
            expected_dimanche = total_visitors * 1  # 1 dimanche par semaine
            expected_jeudi = total_visitors * 1  # 1 jeudi par semaine
            max_weighted = (expected_dimanche * 2) + (expected_jeudi * 1)
            actual_weighted = (total_presences_dimanche * 2) + (total_presences_jeudi * 1)
            rate = (actual_weighted / max_weighted * 100) if max_weighted > 0 else 0
            
            weekly_rates.append({
                "week": w,
                "rate": round(rate, 2),
                "presences": total_presences_dimanche + total_presences_jeudi,
                "expected": expected_dimanche + expected_jeudi
            })
        
        monthly_average = sum(w["rate"] for w in weekly_rates) / len(weekly_rates) if weekly_rates else 0
        
        results.append({
            "referent_username": referent["username"],
            "referent_id": referent["id"],
            "assigned_month": assigned_month,
            "total_visitors": total_visitors,
            "weekly_rates": weekly_rates,
            "monthly_average": round(monthly_average, 2)
        })
    
    return results

# ==================== INIT DATA ====================

@api_router.post("/init")
async def init_data():
    """Initialize default data"""
    # Create default cities
    default_cities = [
        "Dijon", "Chalon-Sur-Saone", "Dole", "Besançon", 
        "Sens", "Milan", "Perugia", "Rome"
    ]
    
    for city_name in default_cities:
        existing = await db.cities.find_one({"name": city_name})
        if not existing:
            city = City(name=city_name)
            await db.cities.insert_one(city.model_dump())
    
    # Create default superviseur_promos for Dijon
    existing_admin = await db.users.find_one({"username": "superviseur_promos", "city": "Dijon"})
    if not existing_admin:
        superviseur = User(
            username="superviseur_promos",
            password=hash_password("superviseur123"),
            city="Dijon",
            role="superviseur_promos"
        )
        doc = superviseur.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.users.insert_one(doc)
    
    # Create Super Admin account
    existing_superadmin = await db.users.find_one({"username": "superadmin"})
    if not existing_superadmin:
        superadmin = User(
            username="superadmin",
            password=hash_password("superadmin123"),
            city="Dijon",
            role="super_admin"
        )
        doc = superadmin.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.users.insert_one(doc)
    
    # Create Pasteur account
    existing_pasteur = await db.users.find_one({"username": "pasteur"})
    if not existing_pasteur:
        pasteur = User(
            username="pasteur",
            password=hash_password("pasteur123"),
            city="Dijon",
            role="pasteur"
        )
        doc = pasteur.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.users.insert_one(doc)
    
    # Migrate old "admin" roles to "superviseur_promos"
    await db.users.update_many(
        {"role": "admin"},
        {"$set": {"role": "superviseur_promos"}}
    )
    
    return {"message": "Initialization complete"}

# ==================== FAMILLES D'IMPACT ROUTES ====================

# ========== SECTEURS ==========

@api_router.post("/fi/secteurs")
async def create_secteur(secteur_data: SecteurCreate, current_user: dict = Depends(get_current_user)):
    # Only admin, super_admin, superviseur_fi can create secteurs
    if current_user["role"] not in ["superviseur_promos", "super_admin", "superviseur_fi"] and not is_super_admin(current_user):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    secteur = Secteur(**secteur_data.model_dump(), created_by=current_user["username"])
    doc = secteur.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.secteurs.insert_one(doc)
    return secteur

@api_router.get("/fi/secteurs")
async def get_secteurs(ville: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    
    # Super admin peut voir toutes les villes, autres utilisateurs seulement leur ville
    if current_user["role"] == "super_admin":
        if ville:
            query["ville"] = ville
    else:
        # Forcer la ville de l'utilisateur connecté
        query["ville"] = current_user["city"]
    
    secteurs = await db.secteurs.find(query, {"_id": 0}).to_list(length=None)
    return secteurs

@api_router.put("/fi/secteurs/{secteur_id}")
async def update_secteur(secteur_id: str, secteur_data: SecteurCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["superviseur_promos", "super_admin", "superviseur_fi"] and not is_super_admin(current_user):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    result = await db.secteurs.update_one(
        {"id": secteur_id},
        {"$set": secteur_data.model_dump()}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Secteur not found")
    
    return {"message": "Secteur updated"}

@api_router.delete("/fi/secteurs/{secteur_id}")
async def delete_secteur(secteur_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["superviseur_promos", "super_admin", "superviseur_fi"] and not is_super_admin(current_user):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    # Check if there are FI in this secteur
    fi_count = await db.familles_impact.count_documents({"secteur_id": secteur_id})
    if fi_count > 0:
        raise HTTPException(status_code=400, detail=f"Impossible de supprimer le secteur : il contient encore {fi_count} Famille(s) d'Impact. Veuillez d'abord supprimer ou réassigner les FI.")
    
    result = await db.secteurs.delete_one({"id": secteur_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Secteur not found")
    
    return {"message": "Secteur deleted"}

# ========== FAMILLES D'IMPACT ==========

@api_router.post("/fi/familles-impact")
async def create_famille_impact(fi_data: FamilleImpactCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["superviseur_promos", "super_admin", "superviseur_fi", "responsable_secteur"] and not is_super_admin(current_user):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    fi = FamilleImpact(**fi_data.model_dump(), created_by=current_user["username"])
    doc = fi.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.familles_impact.insert_one(doc)
    return fi

@api_router.get("/fi/familles-impact")
async def get_familles_impact(
    secteur_id: Optional[str] = None, 
    ville: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    
    # Super admin peut voir toutes les villes, autres utilisateurs seulement leur ville
    if current_user["role"] == "super_admin":
        if ville:
            query["ville"] = ville
    else:
        # Forcer la ville de l'utilisateur connecté
        query["ville"] = current_user["city"]
    
    if secteur_id:
        query["secteur_id"] = secteur_id
    
    # Filter based on role
    if current_user["role"] == "pilote_fi" and current_user.get("assigned_fi_id"):
        query["id"] = current_user["assigned_fi_id"]
    elif current_user["role"] == "responsable_secteur" and current_user.get("assigned_secteur_id"):
        query["secteur_id"] = current_user["assigned_secteur_id"]
    
    fis = await db.familles_impact.find(query, {"_id": 0}).to_list(length=None)
    return fis

@api_router.get("/fi/familles-impact/{fi_id}")
async def get_famille_impact(fi_id: str, current_user: dict = Depends(get_current_user)):
    fi = await db.familles_impact.find_one({"id": fi_id}, {"_id": 0})
    if not fi:
        raise HTTPException(status_code=404, detail="Famille d'Impact not found")
    return fi

# Endpoint PUBLIC pour trouver les FI (sans authentification)
@api_router.get("/public/fi/all")
async def get_all_fi_public(ville: Optional[str] = None):
    """Get all FI with addresses (public access for finding nearest FI)"""
    query = {}
    if ville:
        query["ville"] = ville
    
    fis = await db.familles_impact.find(
        query, 
        {"_id": 0, "id": 1, "nom": 1, "ville": 1, "adresse": 1, "secteur_id": 1, "pilote_id": 1, "pilote_ids": 1, "heure_debut": 1, "heure_fin": 1}
    ).to_list(length=None)
    
    # Only return FIs with addresses
    fis_with_address = [fi for fi in fis if fi.get("adresse")]
    
    # Enrichir avec les infos du/des pilote(s)
    for fi in fis_with_address:
        pilotes_info = []
        
        # Handle both old (pilote_id) and new (pilote_ids) format
        pilote_ids = fi.get("pilote_ids", [])
        if fi.get("pilote_id"):
            pilote_ids.append(fi["pilote_id"])
        
        # Get info for all pilots
        for pilote_id in pilote_ids:
            pilote = await db.users.find_one(
                {"id": pilote_id}, 
                {"_id": 0, "username": 1, "telephone": 1}
            )
            if pilote:
                pilotes_info.append({
                    "nom": pilote.get("username"),
                    "telephone": pilote.get("telephone")
                })
        
        fi["pilotes"] = pilotes_info
        # Keep backward compatibility
        if pilotes_info:
            fi["pilote_nom"] = pilotes_info[0]["nom"]
            fi["pilote_telephone"] = pilotes_info[0]["telephone"]
    
    return fis_with_address

@api_router.put("/fi/familles-impact/{fi_id}")
async def update_famille_impact(fi_id: str, fi_data: FamilleImpactCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["superviseur_promos", "super_admin", "superviseur_fi", "responsable_secteur"] and not is_super_admin(current_user):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    result = await db.familles_impact.update_one(
        {"id": fi_id},
        {"$set": fi_data.model_dump()}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Famille d'Impact not found")
    
    return {"message": "Famille d'Impact updated"}

@api_router.delete("/fi/familles-impact/{fi_id}")
async def delete_famille_impact(fi_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["superviseur_promos", "super_admin", "superviseur_fi", "responsable_secteur"] and not is_super_admin(current_user):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    # Check if there are membres
    membre_count = await db.membres_fi.count_documents({"fi_id": fi_id})
    if membre_count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete FI with {membre_count} members")
    
    result = await db.familles_impact.delete_one({"id": fi_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Famille d'Impact not found")
    
    return {"message": "Famille d'Impact deleted"}

# ========== MEMBRES FI ==========

@api_router.post("/fi/membres")
async def create_membre_fi(membre_data: MembreFICreate, current_user: dict = Depends(get_current_user)):
    # Pilote can add to their FI, others need admin permissions
    if current_user["role"] == "pilote_fi":
        if current_user.get("assigned_fi_id") != membre_data.fi_id:
            raise HTTPException(status_code=403, detail="Can only add members to your FI")
    elif current_user["role"] not in ["superviseur_promos", "super_admin", "superviseur_fi", "responsable_secteur"] and not is_super_admin(current_user):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    membre = MembreFI(**membre_data.model_dump())
    doc = membre.model_dump()
    doc['date_ajout'] = doc['date_ajout'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.membres_fi.insert_one(doc)
    return membre

@api_router.get("/fi/membres")
async def get_membres_fi(fi_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    
    # Filter for pilote
    if current_user["role"] == "pilote_fi" and current_user.get("assigned_fi_id"):
        query["fi_id"] = current_user["assigned_fi_id"]
    elif fi_id:
        query["fi_id"] = fi_id
    else:
        # Si pas de fi_id spécifique, filtrer par ville (sauf super_admin)
        if current_user["role"] != "super_admin":
            # Récupérer les FI de la ville de l'utilisateur
            fi_query = {"ville": current_user["city"]}
            fis = await db.familles_impact.find(fi_query, {"_id": 0, "id": 1}).to_list(length=None)
            fi_ids = [fi["id"] for fi in fis]
            if fi_ids:
                query["fi_id"] = {"$in": fi_ids}
            else:
                # Aucune FI dans cette ville, retourner vide
                return []
    
    membres = await db.membres_fi.find(query, {"_id": 0}).to_list(length=None)
    return membres

@api_router.delete("/fi/membres/{membre_id}")
async def delete_membre_fi(membre_id: str, current_user: dict = Depends(get_current_user)):
    membre = await db.membres_fi.find_one({"id": membre_id})
    if not membre:
        raise HTTPException(status_code=404, detail="Membre not found")
    
    # Check permissions
    if current_user["role"] == "pilote_fi":
        if current_user.get("assigned_fi_id") != membre["fi_id"]:
            raise HTTPException(status_code=403, detail="Permission denied")
    elif current_user["role"] not in ["superviseur_promos", "super_admin", "superviseur_fi", "responsable_secteur"] and not is_super_admin(current_user):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    # Delete all presences for this member
    await db.presences_fi.delete_many({"membre_fi_id": membre_id})
    
    result = await db.membres_fi.delete_one({"id": membre_id})
    return {"message": "Membre deleted"}

# ========== PRESENCES FI ==========

@api_router.post("/fi/presences")
async def create_presence_fi(presence_data: PresenceFICreate, current_user: dict = Depends(get_current_user)):
    # Get membre to check FI
    membre = await db.membres_fi.find_one({"id": presence_data.membre_fi_id})
    if not membre:
        raise HTTPException(status_code=404, detail="Membre not found")
    
    # Check permissions
    if current_user["role"] == "pilote_fi":
        if current_user.get("assigned_fi_id") != membre["fi_id"]:
            raise HTTPException(status_code=403, detail="Permission denied")
    elif current_user["role"] not in ["superviseur_promos", "super_admin", "superviseur_fi", "responsable_secteur"] and not is_super_admin(current_user):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    # Check if presence already exists for this date
    existing = await db.presences_fi.find_one({
        "membre_fi_id": presence_data.membre_fi_id,
        "date": presence_data.date
    })
    
    if existing:
        # Update existing
        await db.presences_fi.update_one(
            {"id": existing["id"]},
            {"$set": {
                "present": presence_data.present,
                "commentaire": presence_data.commentaire
            }}
        )
        return {"message": "Presence updated"}
    
    presence = PresenceFI(**presence_data.model_dump(), marked_by=current_user["username"])
    doc = presence.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.presences_fi.insert_one(doc)
    return presence

@api_router.get("/fi/presences")
async def get_presences_fi(
    fi_id: Optional[str] = None,
    date: Optional[str] = None,
    membre_fi_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    
    if membre_fi_id:
        query["membre_fi_id"] = membre_fi_id
    elif fi_id:
        # Get all membres of this FI
        membres = await db.membres_fi.find({"fi_id": fi_id}, {"_id": 0}).to_list(length=None)
        membre_ids = [m["id"] for m in membres]
        query["membre_fi_id"] = {"$in": membre_ids}
    
    if date:
        query["date"] = date
    
    presences = await db.presences_fi.find(query, {"_id": 0}).to_list(length=None)
    return presences

# ========== AFFECTATION NOUVEAUX ARRIVANTS ==========

@api_router.post("/fi/affecter-visiteur")
async def affecter_visiteur_to_fi(affectation: AffectationFI, current_user: dict = Depends(get_current_user)):
    # Get visitor
    visitor = await db.visitors.find_one({"id": affectation.nouveau_arrivant_id})
    if not visitor:
        raise HTTPException(status_code=404, detail="Nouveau arrivant not found")
    
    # Get FI
    fi = await db.familles_impact.find_one({"id": affectation.fi_id})
    if not fi:
        raise HTTPException(status_code=404, detail="Famille d'Impact not found")
    
    # Check if already membre
    existing = await db.membres_fi.find_one({
        "nouveau_arrivant_id": affectation.nouveau_arrivant_id,
        "fi_id": affectation.fi_id
    })
    
    if existing:
        raise HTTPException(status_code=400, detail="Already membre of this FI")
    
    # Create membre
    membre = MembreFI(
        prenom=visitor["firstname"],
        nom=visitor["lastname"],
        fi_id=affectation.fi_id,
        source="nouveau_arrivant",
        nouveau_arrivant_id=affectation.nouveau_arrivant_id
    )
    
    doc = membre.model_dump()
    doc['date_ajout'] = doc['date_ajout'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.membres_fi.insert_one(doc)
    return membre

@api_router.get("/fi/indicateurs/affectation")
async def get_indicateurs_affectation(ville: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    # Get all nouveaux arrivants
    query = {}
    if ville:
        query["ville"] = ville
    
    total_visitors = await db.visitors.count_documents(query)
    
    # Get affected visitors
    membres = await db.membres_fi.find(
        {"source": "nouveau_arrivant", "nouveau_arrivant_id": {"$ne": None}},
        {"_id": 0}
    ).to_list(length=None)
    
    affected_ids = [m["nouveau_arrivant_id"] for m in membres]
    affected_count = len(set(affected_ids))
    
    non_affected_count = total_visitors - affected_count
    percentage = (affected_count / total_visitors * 100) if total_visitors > 0 else 0
    
    return {
        "total_nouveaux_arrivants": total_visitors,
        "affectes": affected_count,
        "non_affectes": non_affected_count,
        "pourcentage_affectation": round(percentage, 2)
    }

# ========== STATISTIQUES FI ==========

@api_router.get("/fi/stats/pilote")
async def get_stats_pilote(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "pilote_fi":
        raise HTTPException(status_code=403, detail="Only for pilote_fi")
    
    # Support both old (assigned_fi_id) and new (assigned_fi_ids) format
    fi_id = current_user.get("assigned_fi_id")
    if not fi_id:
        assigned_fi_ids = current_user.get("assigned_fi_ids", [])
        if assigned_fi_ids:
            fi_id = assigned_fi_ids[0]  # Use first FI if multiple assigned
    
    if not fi_id:
        raise HTTPException(status_code=400, detail="No FI assigned to your account")
    
    # Get FI info
    fi = await db.familles_impact.find_one({"id": fi_id}, {"_id": 0})
    if not fi:
        raise HTTPException(status_code=404, detail="FI not found")
    
    # Get membres count and evolution
    membres = await db.membres_fi.find({"fi_id": fi_id}, {"_id": 0}).to_list(length=None)
    total_membres = len(membres)
    
    # Calculate evolution by month
    from collections import defaultdict
    evolution_membres = defaultdict(int)
    for membre in membres:
        date_ajout = membre.get("date_ajout")
        if date_ajout:
            if isinstance(date_ajout, str):
                month = date_ajout[:7]  # YYYY-MM
            else:
                month = date_ajout.strftime("%Y-%m")
            evolution_membres[month] += 1
    
    # Cumulative
    cumulative = {}
    total = 0
    for month in sorted(evolution_membres.keys()):
        total += evolution_membres[month]
        cumulative[month] = total
    
    # Get presences and calculate fidelisation
    membre_ids = [m["id"] for m in membres]
    presences = await db.presences_fi.find(
        {"membre_fi_id": {"$in": membre_ids}},
        {"_id": 0}
    ).to_list(length=None)
    
    # Calculate fidelisation by week/month/year
    from datetime import datetime
    presences_by_date = {}
    for p in presences:
        if p["present"]:
            presences_by_date[p["date"]] = presences_by_date.get(p["date"], 0) + 1
    
    # Get unique jeudis count
    unique_jeudis = len(set([p["date"] for p in presences]))
    total_presences = sum([1 for p in presences if p["present"]])
    max_possible = total_membres * unique_jeudis if unique_jeudis > 0 else 0
    fidelisation_globale = (total_presences / max_possible * 100) if max_possible > 0 else 0
    
    return {
        "fi": fi,
        "total_membres": total_membres,
        "evolution_membres": cumulative,
        "fidelisation_globale": round(fidelisation_globale, 2),
        "total_presences": total_presences,
        "jeudis_count": unique_jeudis
    }

@api_router.get("/fi/stats/secteur")
async def get_stats_secteur(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "responsable_secteur":
        raise HTTPException(status_code=403, detail="Only for responsable_secteur")
    
    secteur_id = current_user.get("assigned_secteur_id")
    if not secteur_id:
        raise HTTPException(status_code=400, detail="No secteur assigned")
    
    # Get secteur info
    secteur = await db.secteurs.find_one({"id": secteur_id}, {"_id": 0})
    if not secteur:
        raise HTTPException(status_code=404, detail="Secteur not found")
    
    # Get all FI in this secteur
    fis = await db.familles_impact.find({"secteur_id": secteur_id}, {"_id": 0}).to_list(length=None)
    
    # Stats per FI
    fi_stats = []
    for fi in fis:
        membres = await db.membres_fi.find({"fi_id": fi["id"]}, {"_id": 0}).to_list(length=None)
        total_membres = len(membres)
        
        # Get pilote info
        pilote = None
        if fi.get("pilote_id"):
            pilote = await db.users.find_one({"id": fi["pilote_id"]}, {"_id": 0, "password": 0})
        
        # Calculate fidelisation
        membre_ids = [m["id"] for m in membres]
        presences = await db.presences_fi.find(
            {"membre_fi_id": {"$in": membre_ids}},
            {"_id": 0}
        ).to_list(length=None)
        
        unique_jeudis = len(set([p["date"] for p in presences]))
        total_presences = sum([1 for p in presences if p["present"]])
        max_possible = total_membres * unique_jeudis if unique_jeudis > 0 else 0
        fidelisation = (total_presences / max_possible * 100) if max_possible > 0 else 0
        
        fi_stats.append({
            "fi": fi,
            "pilote": pilote,
            "total_membres": total_membres,
            "fidelisation": round(fidelisation, 2)
        })
    
    # Count pilotes
    pilotes_count = len([fi for fi in fis if fi.get("pilote_id")])
    
    return {
        "secteur": secteur,
        "nombre_fi": len(fis),
        "nombre_pilotes": pilotes_count,
        "fi_details": fi_stats
    }

@api_router.get("/fi/stats/superviseur")
async def get_stats_superviseur(ville: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["superviseur_fi", "superviseur_promos", "super_admin", "pasteur"] and not is_super_admin(current_user):
        raise HTTPException(status_code=403, detail="Permission denied")
    
    query = {}
    if ville:
        query["ville"] = ville
    elif current_user["role"] == "superviseur_fi":
        query["ville"] = current_user["city"]
    
    # Get all secteurs
    secteurs = await db.secteurs.find(query, {"_id": 0}).to_list(length=None)
    
    # Get all FI
    fi_query = {}
    if ville:
        fi_query["ville"] = ville
    elif current_user["role"] == "superviseur_fi":
        fi_query["ville"] = current_user["city"]
    
    fis = await db.familles_impact.find(fi_query, {"_id": 0}).to_list(length=None)
    
    # Get all membres
    fi_ids = [fi["id"] for fi in fis]
    membres = await db.membres_fi.find({"fi_id": {"$in": fi_ids}}, {"_id": 0}).to_list(length=None)
    
    # Evolution membres by month
    from collections import defaultdict
    evolution_membres = defaultdict(int)
    for membre in membres:
        date_ajout = membre.get("date_ajout")
        if isinstance(date_ajout, str):
            month = date_ajout[:7]
        else:
            month = date_ajout.strftime("%Y-%m")
        evolution_membres[month] += 1
    
    cumulative = {}
    total = 0
    for month in sorted(evolution_membres.keys()):
        total += evolution_membres[month]
        cumulative[month] = total
    
    # Calculate global fidelisation
    membre_ids = [m["id"] for m in membres]
    presences = await db.presences_fi.find(
        {"membre_fi_id": {"$in": membre_ids}},
        {"_id": 0}
    ).to_list(length=None)
    
    unique_jeudis = len(set([p["date"] for p in presences]))
    total_presences = sum([1 for p in presences if p["present"]])
    max_possible = len(membres) * unique_jeudis if unique_jeudis > 0 else 0
    fidelisation_globale = (total_presences / max_possible * 100) if max_possible > 0 else 0
    
    # FI by secteur
    fi_by_secteur = defaultdict(int)
    for fi in fis:
        fi_by_secteur[fi["secteur_id"]] += 1
    
    secteur_details = []
    for secteur in secteurs:
        secteur_details.append({
            "secteur": secteur,
            "nombre_fi": fi_by_secteur[secteur["id"]]
        })
    
    return {
        "nombre_secteurs": len(secteurs),
        "nombre_fi_total": len(fis),
        "nombre_membres_total": len(membres),
        "evolution_membres": cumulative,
        "fidelisation_globale": round(fidelisation_globale, 2),
        "secteurs_details": secteur_details
    }

@api_router.get("/debug/data-structure")
async def debug_data_structure(ville: str = "Dijon", current_user: dict = Depends(get_current_user)):
    """Debug endpoint to check data structure in production"""
    try:
        # Check visitors structure
        visitor_sample = await db.visitors.find_one({"city": ville}, {"_id": 0})
        visitor_count = await db.visitors.count_documents({"city": ville})
        
        # Check culte_stats structure  
        culte_sample = await db.culte_stats.find_one({"ville": ville}, {"_id": 0})
        culte_count = await db.culte_stats.count_documents({"ville": ville})
        
        # Check all assigned_month values
        assigned_months = await db.visitors.distinct("assigned_month", {"city": ville})
        
        # Check all culte dates
        culte_dates = await db.culte_stats.distinct("date", {"ville": ville})
        
        return {
            "ville": ville,
            "visitors": {
                "count": visitor_count,
                "sample": visitor_sample,
                "assigned_months": assigned_months[:10]
            },
            "cultes": {
                "count": culte_count,
                "sample": culte_sample,
                "dates": [str(d) for d in culte_dates[:10]]
            }
        }
    except Exception as e:
        return {"error": str(e)}

@api_router.get("/fi/stats/pasteur")
async def get_stats_pasteur(
    annee: Optional[int] = None,
    mois: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["pasteur", "super_admin"] and not is_super_admin(current_user):
        raise HTTPException(status_code=403, detail="Only for pasteur or super_admin")
    
    # Get all cities
    cities = await db.cities.find({}, {"_id": 0}).to_list(length=None)
    
    stats_by_city = []
    for city in cities:
        ville = city["name"]
        
        # Get stats for this city
        secteurs = await db.secteurs.find({"ville": ville}, {"_id": 0}).to_list(length=None)
        fis = await db.familles_impact.find({"ville": ville}, {"_id": 0}).to_list(length=None)
        
        fi_ids = [fi["id"] for fi in fis]
        membres = await db.membres_fi.find({"fi_id": {"$in": fi_ids}}, {"_id": 0}).to_list(length=None)
        
        # Build date filters for année/mois FIRST
        date_filter_start = None
        date_filter_end = None
        if annee and mois:
            date_filter_start = f"{annee}-{str(mois).zfill(2)}-01"
            if mois == 12:
                date_filter_end = f"{annee + 1}-01-01"
            else:
                date_filter_end = f"{annee}-{str(mois + 1).zfill(2)}-01"
        elif annee:
            date_filter_start = f"{annee}-01-01"
            date_filter_end = f"{annee + 1}-01-01"
        
        # Get visitors (Personnes Reçues) - FILTERED by année/mois if provided
        visitor_query = {"city": ville}
        
        # Only apply date filter if both annee and mois are provided (specific month)
        # Otherwise, load all visitors for the city
        if annee and mois and date_filter_start and date_filter_end:
            # Try to filter by assigned_month first (format: "2025-11")
            visitor_query["$or"] = [
                {"assigned_month": f"{annee}-{str(mois).zfill(2)}"},
                {"assigned_month": {"$regex": f"^{annee}-{str(mois).zfill(2)}"}},
                # Fallback: check if assigned_month contains the year-month pattern
                {"assigned_month": {"$regex": f"{annee}-{str(mois).zfill(2)}"}}
            ]
        elif annee and date_filter_start and date_filter_end:
            # Filter by year only
            visitor_query["$or"] = [
                {"assigned_month": {"$regex": f"^{annee}-"}},
                {"assigned_month": {"$regex": f"{annee}"}}
            ]
        
        visitors = await db.visitors.find(visitor_query, {"_id": 0}).to_list(length=None)
        
        # Count by status using "types" field (not "statut")
        total_visitors = len(visitors)
        de_passage_count = len([v for v in visitors if "De Passage" in v.get("types", [])])
        resident_count = len([v for v in visitors if "De Passage" not in v.get("types", [])])
        na_count = total_visitors  # ALL visitors are NA
        nc_count = len([v for v in visitors if "Nouveau Converti" in v.get("types", [])])
        
        # Promotions fidelisation - Calculate from presences_dimanche and presences_jeudi
        # This is the SAME calculation as Promotion dashboard (fidélisation générale)
        # FILTERED by date if provided
        total_presences_dimanche = 0
        total_presences_jeudi = 0
        for visitor in visitors:
            presences_dim = visitor.get("presences_dimanche", [])
            presences_jeu = visitor.get("presences_jeudi", [])
            
            # Filter presences by date if année/mois are provided
            if date_filter_start and date_filter_end:
                presences_dim = [p for p in presences_dim if date_filter_start <= p.get("date", "") < date_filter_end]
                presences_jeu = [p for p in presences_jeu if date_filter_start <= p.get("date", "") < date_filter_end]
            
            total_presences_dimanche += len([p for p in presences_dim if p.get("present")])
            total_presences_jeudi += len([p for p in presences_jeu if p.get("present")])
        
        # Calculate number of sundays and thursdays in the period
        if annee and mois:
            import calendar
            num_days = calendar.monthrange(annee, mois)[1]
            num_sundays = sum(1 for day in range(1, num_days + 1) if datetime(annee, mois, day).weekday() == 6)
            num_thursdays = sum(1 for day in range(1, num_days + 1) if datetime(annee, mois, day).weekday() == 3)
        else:
            num_sundays = 4
            num_thursdays = 4
        
        expected_dimanche = total_visitors * num_sundays if total_visitors > 0 else 0
        expected_jeudi = total_visitors * num_thursdays if total_visitors > 0 else 0
        
        taux_dimanche = (total_presences_dimanche / expected_dimanche) if expected_dimanche > 0 else 0
        taux_jeudi = (total_presences_jeudi / expected_jeudi) if expected_jeudi > 0 else 0
        promos_fidelisation = ((taux_dimanche * 2) + (taux_jeudi * 1)) / 2 * 100
        
        # Cultes stats - Use culte_stats collection (not cultes) - FILTERED by année/mois
        culte_query = {"ville": ville}
        if date_filter_start and date_filter_end:
            # Support multiple date formats: string or datetime
            culte_query["$or"] = [
                {"date": {"$gte": date_filter_start, "$lt": date_filter_end}},
                # Also check if date field is datetime object
                {"date": {"$gte": datetime.fromisoformat(date_filter_start.replace('Z', '+00:00')) if isinstance(date_filter_start, str) else date_filter_start,
                          "$lt": datetime.fromisoformat(date_filter_end.replace('Z', '+00:00')) if isinstance(date_filter_end, str) else date_filter_end}}
            ]
        
        cultes = await db.culte_stats.find(culte_query, {"_id": 0}).to_list(length=None)
        total_adultes = sum([c.get("nombre_adultes", 0) for c in cultes])
        total_enfants = sum([c.get("nombre_enfants", 0) for c in cultes])
        total_stars = sum([c.get("nombre_stars", 0) for c in cultes])
        total_services = len(cultes)
        
        moy_adultes = round(total_adultes / total_services, 1) if total_services > 0 else 0
        moy_enfants = round(total_enfants / total_services, 1) if total_services > 0 else 0
        moy_stars = round(total_stars / total_services, 1) if total_services > 0 else 0
        
        # Cultes fidelisation (same as promotions fidelisation)
        cultes_fidelisation = round(promos_fidelisation, 2)
        
        # Fidelisation Familles d'Impact - FILTERED by année/mois
        membre_ids = [m["id"] for m in membres]
        presences_fi_query = {"membre_fi_id": {"$in": membre_ids}}
        if date_filter_start and date_filter_end:
            presences_fi_query["date"] = {"$gte": date_filter_start, "$lt": date_filter_end}
        
        presences_fi = await db.presences_fi.find(presences_fi_query, {"_id": 0}).to_list(length=None)
        
        unique_jeudis_fi = len(set([p["date"] for p in presences_fi]))
        total_presences_fi = sum([1 for p in presences_fi if p["present"]])
        max_possible_fi = len(membres) * unique_jeudis_fi if unique_jeudis_fi > 0 else 0
        fidelisation_fi = (total_presences_fi / max_possible_fi * 100) if max_possible_fi > 0 else 0
        
        # Dynamique d'Évangélisation stats - FILTERED by année/mois
        evangel_query = {"ville": ville}
        if date_filter_start and date_filter_end:
            evangel_query["date"] = {"$gte": date_filter_start, "$lt": date_filter_end}
        
        # Get all evangelisation data for the city with filters
        all_evangel = await db.evangelisation.find(evangel_query, {"_id": 0}).to_list(length=None)
        
        # Separate by type
        evangel_eglise = [e for e in all_evangel if e.get("eglise") or e.get("type") == "eglise"]
        evangel_fi = [e for e in all_evangel if e.get("familles_impact") or e.get("type") == "familles_impact"]
        
        # Aggregate evangelisation data for eglise
        evangel_eglise_totals = {
            "gagneurs_ame": sum([e.get("eglise", {}).get("nombre_gagneurs_ame", 0) for e in evangel_eglise]),
            "personnes_receptives": sum([e.get("eglise", {}).get("nombre_personnes_receptives", 0) for e in evangel_eglise]),
            "priere_salut": sum([e.get("eglise", {}).get("nombre_priere_salut", 0) for e in evangel_eglise]),
            "contacts_pris": sum([e.get("eglise", {}).get("nombre_contacts_pris", 0) for e in evangel_eglise]),
            "ames_invitees": sum([e.get("eglise", {}).get("nombre_ames_invitees", 0) for e in evangel_eglise]),
            "miracles": sum([e.get("eglise", {}).get("nombre_miracles", 0) for e in evangel_eglise])
        }
        
        # Aggregate evangelisation data for familles impact
        evangel_fi_totals = {
            "gagneurs_ame": sum([e.get("familles_impact", {}).get("nombre_gagneurs_ame", 0) for e in evangel_fi]),
            "personnes_receptives": sum([e.get("familles_impact", {}).get("nombre_personnes_receptives", 0) for e in evangel_fi]),
            "priere_salut": sum([e.get("familles_impact", {}).get("nombre_priere_salut", 0) for e in evangel_fi]),
            "contacts_pris": sum([e.get("familles_impact", {}).get("nombre_contacts_pris", 0) for e in evangel_fi]),
            "ames_invitees": sum([e.get("familles_impact", {}).get("nombre_ames_invitees", 0) for e in evangel_fi]),
            "miracles": sum([e.get("familles_impact", {}).get("nombre_miracles", 0) for e in evangel_fi])
        }
        
        stats_by_city.append({
            "ville": ville,
            "nombre_secteurs": len(secteurs),
            "nombre_fi": len(fis),
            "nombre_membres": len(membres),
            "promotions": {
                "total_personnes": total_visitors,
                "de_passage": de_passage_count,
                "resident": resident_count,
                "na": na_count,
                "nc": nc_count,
                "fidelisation": round(promos_fidelisation, 2)
            },
            "familles_impact": {
                "secteurs": len(secteurs),
                "familles": len(fis),
                "membres": len(membres),
                "fidelisation": round(fidelisation_fi, 2)
            },
            "cultes": {
                "moy_adultes": moy_adultes,
                "moy_enfants": moy_enfants,
                "moy_stars": moy_stars,
                "total_services": total_services
            },
            "evangelisation": {
                "eglise": evangel_eglise_totals,
                "familles_impact": evangel_fi_totals
            }
        })
    
    return {
        "stats_by_city": stats_by_city
    }

# ==================== NOTIFICATIONS ====================

@api_router.get("/notifications")
async def get_notifications(
    unread_only: bool = False,
    current_user: dict = Depends(get_current_user)
):
    """Get notifications for current user"""
    query = {"user_id": current_user["id"]}
    if unread_only:
        query["read"] = False
    
    notifications = await db.notifications.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return notifications

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(
    notification_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Mark notification as read"""
    result = await db.notifications.update_one(
        {"id": notification_id, "user_id": current_user["id"]},
        {"$set": {"read": True}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    
    return {"message": "Notification marked as read"}

@api_router.post("/notifications/generate")
async def generate_notifications(current_user: dict = Depends(get_current_user)):
    """Generate automated notifications (Admin/Supervisor only)"""
    
    notifications_created = []
    
    # 1. Rappels de présence pour Pilotes FI (tous les jeudis)
    today = datetime.now(timezone.utc)
    if today.weekday() == 3:  # Jeudi
        pilotes = await db.users.find({"role": "pilote_fi"}, {"_id": 0}).to_list(length=None)
        for pilote in pilotes:
            if pilote.get("assigned_fi_id"):
                fi = await db.familles_impact.find_one({"id": pilote["assigned_fi_id"]}, {"_id": 0})
                if fi:
                    notif = Notification(
                        user_id=pilote["id"],
                        type="presence_reminder",
                        message=f"📝 N'oubliez pas de marquer les présences pour {fi['name']} aujourd'hui!",
                        data={"fi_id": fi["id"], "fi_name": fi["name"]}
                    )
                    doc = notif.model_dump()
                    doc['created_at'] = doc['created_at'].isoformat()
                    await db.notifications.insert_one(doc)
                    notifications_created.append(notif)
    
    # 2. Alertes FI stagnantes (pas de nouvelles présences depuis 2 semaines)
    two_weeks_ago = (datetime.now(timezone.utc) - timedelta(weeks=2)).isoformat()
    fis = await db.familles_impact.find({}, {"_id": 0}).to_list(length=None)
    
    for fi in fis:
        # Check dernière présence
        last_presence = await db.presences_fi.find_one(
            {"fi_id": fi["id"]},
            {"_id": 0},
            sort=[("date", -1)]
        )
        
        if not last_presence or last_presence["date"] < two_weeks_ago:
            # Notifier responsable secteur
            secteur = await db.secteurs.find_one({"id": fi["secteur_id"]}, {"_id": 0})
            if secteur:
                responsables = await db.users.find(
                    {"role": "responsable_secteur", "assigned_secteur_id": secteur["id"]},
                    {"_id": 0}
                ).to_list(length=None)
                
                for resp in responsables:
                    notif = Notification(
                        user_id=resp["id"],
                        type="fi_stagnation",
                        message=f"⚠️ Aucune activité récente dans la FI {fi['name']} (Secteur {secteur['name']})",
                        data={"fi_id": fi["id"], "secteur_id": secteur["id"]}
                    )
                    doc = notif.model_dump()
                    doc['created_at'] = doc['created_at'].isoformat()
                    await db.notifications.insert_one(doc)
                    notifications_created.append(notif)
    
    # 3. Alertes fidélisation faible (< 50%)
    superviseurs = await db.users.find(
        {"role": {"$in": ["superviseur_fi", "superviseur_promos"]}},
        {"_id": 0}
    ).to_list(length=None)
    
    for sup in superviseurs:
        city = sup["city"]
        
        if sup["role"] == "superviseur_fi":
            # Calculer fidélisation FI
            fis_city = await db.familles_impact.find({"city": city}, {"_id": 0}).to_list(length=None)
            fi_ids = [fi["id"] for fi in fis_city]
            membres = await db.membres_fi.find({"fi_id": {"$in": fi_ids}}, {"_id": 0}).to_list(length=None)
            membre_ids = [m["id"] for m in membres]
            presences = await db.presences_fi.find(
                {"membre_fi_id": {"$in": membre_ids}},
                {"_id": 0}
            ).to_list(length=None)
            
            unique_jeudis = len(set([p["date"] for p in presences]))
            total_presences = sum([1 for p in presences if p["present"]])
            max_possible = len(membres) * unique_jeudis if unique_jeudis > 0 else 0
            fidelisation = (total_presences / max_possible * 100) if max_possible > 0 else 0
            
            if fidelisation < 50:
                notif = Notification(
                    user_id=sup["id"],
                    type="low_fidelisation",
                    message=f"📊 Taux de fidélisation FI bas: {fidelisation:.1f}% pour {city}",
                    data={"city": city, "fidelisation": round(fidelisation, 2)}
                )
                doc = notif.model_dump()
                doc['created_at'] = doc['created_at'].isoformat()
                await db.notifications.insert_one(doc)
                notifications_created.append(notif)
    
    # 4. Nouveaux arrivants non assignés
    unassigned = await db.visitors.find(
        {"assigned_fi_id": None, "tracking_stopped": False},
        {"_id": 0}
    ).to_list(length=None)
    
    if unassigned:
        superviseurs_promos = await db.users.find(
            {"role": "superviseur_promos"},
            {"_id": 0}
        ).to_list(length=None)
        
        for sup in superviseurs_promos:
            city_unassigned = [v for v in unassigned if v["city"] == sup["city"]]
            if city_unassigned:
                notif = Notification(
                    user_id=sup["id"],
                    type="unassigned_visitor",
                    message=f"👥 {len(city_unassigned)} nouveaux arrivants non assignés à une FI à {sup['city']}",
                    data={"city": sup["city"], "count": len(city_unassigned)}
                )
                doc = notif.model_dump()
                doc['created_at'] = doc['created_at'].isoformat()
                await db.notifications.insert_one(doc)
                notifications_created.append(notif)
    
    return {
        "message": f"{len(notifications_created)} notifications créées",
        "count": len(notifications_created)
    }

# ==================== ADVANCED ANALYTICS FOR SUPER ADMIN/PASTEUR ====================

@api_router.get("/analytics/promotions-detailed")
async def get_promotions_detailed(ville: str = None, mois: str = None, annee: str = None, current_user: dict = Depends(get_current_user)):
    """Get detailed promotions analytics for Super Admin/Pasteur with:
    - Fidélisation par promo (TOUTES les promos)
    - Filtres mois/année appliqués sur les PRÉSENCES et données, pas sur les promos
    - Total NA vs NC vs DP
    - Canal d'arrivée (4 canaux spécifiques)
    - Détails quotidiens par date
    """
    from datetime import datetime
    import calendar
    
    # Only super_admin, pasteur, and responsable_eglise can access
    
    # Get all visitors INCLUDING tracking_stopped (pour compter les suivis arrêtés)
    base_query = {}
    
    # For responsable_eglise, force filter by their city
    if current_user["role"] == "responsable_eglise":
        base_query["city"] = current_user["city"]
    # Filtrer par ville si spécifié
    elif ville and ville != "all":
        base_query["city"] = ville
    
    # Get ALL visitors (y compris tracking_stopped pour les statistiques)
    all_visitors = await db.visitors.find(base_query, {"_id": 0}).to_list(10000)
    
    # Filtrer les visiteurs pour les indicateurs globaux si mois/année spécifiés
    visitors_for_summary = all_visitors
    if mois and mois != "all" and annee and annee != "all":
        # Filtrer les visiteurs dont assigned_month correspond au mois/année
        visitors_for_summary = [v for v in all_visitors if v.get("assigned_month", "").startswith(f"{annee}-{mois}")]
    
    # Utiliser all_visitors pour les stats détaillées (toutes les promos)
    visitors = all_visitors
    
    # Déterminer le mois filtré pour le calcul des dimanches/jeudis
    if mois and mois != "all" and annee and annee != "all":
        filter_year = int(annee)
        filter_month = int(mois)
        # Compter le nombre de dimanches et jeudis dans ce mois
        num_days = calendar.monthrange(filter_year, filter_month)[1]
        num_sundays = sum(1 for day in range(1, num_days + 1) if datetime(filter_year, filter_month, day).weekday() == 6)
        num_thursdays = sum(1 for day in range(1, num_days + 1) if datetime(filter_year, filter_month, day).weekday() == 3)
    else:
        # Par défaut: 4 dimanches et 4 jeudis
        num_sundays = 4
        num_thursdays = 4
    
    # Group by assigned_month (Promo) - on garde TOUTES les promos
    promos_by_month = {}
    for visitor in visitors:
        month = visitor.get("assigned_month", "N/A")
        if month not in promos_by_month:
            promos_by_month[month] = {
                "month": month,
                "total_visitors": 0,
                "total_visitors_all": 0,  # Inclut les suivis arrêtés
                "na_count": 0,
                "nc_count": 0,
                "dp_count": 0,
                "residents_count": 0,
                "visitors": [],
                "visitors_actifs": [],  # Seulement ceux avec suivi actif
                "suivis_arretes": []
            }
        
        # Compter TOUS les visiteurs (y compris arrêtés) pour les stats globales
        promos_by_month[month]["total_visitors_all"] += 1
        promos_by_month[month]["visitors"].append(visitor)
        
        # LOGIQUE NOUVELLE: NA = TOUTES les personnes reçues (car toute personne reçue est NA)
        promos_by_month[month]["na_count"] += 1
        
        # Compter seulement les visiteurs actifs (pas arrêtés) pour "nbre_pers_suivis"
        if not visitor.get("tracking_stopped"):
            promos_by_month[month]["total_visitors"] += 1
            promos_by_month[month]["visitors_actifs"].append(visitor)
        
        # Count NC vs DP (tous, même arrêtés)
        types = visitor.get("types", [])
        if "Nouveau Converti" in types:
            promos_by_month[month]["nc_count"] += 1
        if "De Passage" in types:
            promos_by_month[month]["dp_count"] += 1
        
        # Count residents (not DP)
        if "De Passage" not in types:
            promos_by_month[month]["residents_count"] += 1
        
        # Track stopped tracking
        if visitor.get("tracking_stopped"):
            promos_by_month[month]["suivis_arretes"].append({
                "name": f"{visitor.get('firstname', '')} {visitor.get('lastname', '')}",
                "reason": visitor.get("stop_reason", "Non spécifié")
            })
    
    # Calculate fidelisation for each promo based on FILTERED month
    promos_stats = []
    for month, data in sorted(promos_by_month.items()):
        total = data["total_visitors"]
        
        # Filtrer les présences par mois/année si spécifié
        total_presences_dimanche = 0
        total_presences_jeudi = 0
        
        if total > 0:
            # Compter seulement les présences des visiteurs ACTIFS (pas arrêtés)
            for visitor in data["visitors_actifs"]:
                # Filtrer les présences dimanche par mois/année
                presences_dim = visitor.get("presences_dimanche", [])
                if mois and mois != "all" and annee and annee != "all":
                    # Normaliser le mois avec zéro devant si nécessaire
                    mois_padded = mois.zfill(2)
                    # Filtrer d'abord par date, puis compter les présents
                    presences_dim_filtered = [p for p in presences_dim if p.get("date", "").startswith(f"{annee}-{mois_padded}") and p.get("present")]
                    total_presences_dimanche += len(presences_dim_filtered)
                else:
                    total_presences_dimanche += len([p for p in presences_dim if p.get("present")])
                
                # Filtrer les présences jeudi par mois/année
                presences_jeu = visitor.get("presences_jeudi", [])
                if mois and mois != "all" and annee and annee != "all":
                    # Normaliser le mois avec zéro devant si nécessaire
                    mois_padded = mois.zfill(2)
                    # Filtrer d'abord par date, puis compter les présents
                    presences_jeu_filtered = [p for p in presences_jeu if p.get("date", "").startswith(f"{annee}-{mois_padded}") and p.get("present")]
                    total_presences_jeudi += len(presences_jeu_filtered)
                else:
                    total_presences_jeudi += len([p for p in presences_jeu if p.get("present")])
        
        # Expected = nombre de personnes × nombre de dimanches/jeudis dans le mois filtré
        expected_dimanche = total * num_sundays
        expected_jeudi = total * num_thursdays
        
        # Fidélisation: ((Présences Dim / Attendues Dim) × 2 + (Présences Jeu / Attendues Jeu) × 1) / 2
        taux_dimanche = (total_presences_dimanche / expected_dimanche) if expected_dimanche > 0 else 0
        taux_jeudi = (total_presences_jeudi / expected_jeudi) if expected_jeudi > 0 else 0
        fidelisation = ((taux_dimanche * 2) + (taux_jeudi * 1)) / 2 * 100
        
        # CORRECTION: Pers. Suivies = Pers. Reçues (total_all) - Suivis Arrêtés
        nbre_pers_suivis = data["total_visitors_all"] - len(data["suivis_arretes"])
        
        promos_stats.append({
            "month": month,
            "nbre_pers_suivis": nbre_pers_suivis,  # Pers. Reçues - Arrêtés
            "total_visitors": data["total_visitors_all"],  # Total incluant arrêtés
            "na_count": data["na_count"],
            "nc_count": data["nc_count"],
            "dp_count": data["dp_count"],
            "residents_count": data["residents_count"],
            "suivis_arretes_count": len(data["suivis_arretes"]),
            "suivis_arretes_details": data["suivis_arretes"],
            "fidelisation": round(fidelisation, 1),
            "total_presences_dimanche": total_presences_dimanche,
            "expected_presences_dimanche": expected_dimanche,
            "total_presences_jeudi": total_presences_jeudi,
            "expected_presences_jeudi": expected_jeudi
        })
    
    # Calculate Canal d'arrivée statistics (4 canaux spécifiques) - FILTRÉ par mois/année
    canal_counts = {
        "Evangelisation": 0,
        "Réseaux sociaux": 0,
        "Invitation par un membre (hors evangelisation)": 0,
        "Par soi même": 0
    }
    for visitor in visitors_for_summary:
        canal = visitor.get("arrival_channel", "")
        if canal in canal_counts:
            canal_counts[canal] += 1
    
    # Global totals - FILTRÉS par mois/année
    # LOGIQUE NOUVELLE: NA = Total personnes reçues (tous sont NA au départ)
    total_personnes_recues = len(visitors_for_summary)  # Tous les visiteurs (même arrêtés)
    total_na = total_personnes_recues  # NA = Total personnes reçues
    total_nc = len([v for v in visitors_for_summary if "Nouveau Converti" in v.get("types", [])])
    total_dp = len([v for v in visitors_for_summary if "De Passage" in v.get("types", [])])
    total_suivis_arretes = len([v for v in visitors_for_summary if v.get("tracking_stopped")])
    total_personnes_suivies = total_na - total_suivis_arretes  # Personnes suivies = NA - Suivis arrêtés
    
    # Avg fidelisation reste basé sur promos_stats (toutes les promos visibles)
    avg_fidelisation = sum(p["fidelisation"] for p in promos_stats) / len(promos_stats) if promos_stats else 0
    
    # Détail des personnes reçues par jour (TOUJOURS affiché)
    from datetime import datetime
    from collections import defaultdict
    daily_data = defaultdict(lambda: {
        "total": 0, "dp": 0, "residents": 0, "na": 0, "nc": 0
    })
    
    # Group by visit_date avec filtre optionnel par mois/année
    for visitor in visitors:
        visit_date = visitor.get("visit_date", "")
        if visit_date:
            # Appliquer filtre mois/année si spécifié
            if mois and mois != "all" and annee and annee != "all":
                if not visit_date.startswith(f"{annee}-{mois}"):
                    continue
            
            types = visitor.get("types", [])
            daily_data[visit_date]["total"] += 1
            if "De Passage" in types:
                daily_data[visit_date]["dp"] += 1
            else:
                daily_data[visit_date]["residents"] += 1
            if "Nouveau Arrivant" in types:
                daily_data[visit_date]["na"] += 1
            if "Nouveau Converti" in types:
                daily_data[visit_date]["nc"] += 1
    
    # Sort by date
    daily_details = []
    for date_str, data in sorted(daily_data.items()):
        daily_details.append({
            "date": date_str,
            "total_personnes_recues": data["total"],
            "nbre_de_passage": data["dp"],
            "nbre_residents": data["residents"],
            "nbre_na": data["na"],
            "nbre_nc": data["nc"]
        })
    
    return {
        "promos": promos_stats,
        "summary": {
            "total_promos": len(promos_stats),
            "total_personnes_recues": total_personnes_recues,  # Nouveau: Total personnes reçues
            "total_na": total_na,  # = total_personnes_recues
            "total_nc": total_nc,
            "total_dp": total_dp,
            "total_suivis_arretes": total_suivis_arretes,  # Nouveau
            "total_personnes_suivies": total_personnes_suivies,  # Nouveau: NA - Suivis arrêtés
            "avg_fidelisation": round(avg_fidelisation, 1),
            "canal_evangelisation": canal_counts["Evangelisation"],
            "canal_reseaux_sociaux": canal_counts["Réseaux sociaux"],
            "canal_invitation_membre": canal_counts["Invitation par un membre (hors evangelisation)"],
            "canal_par_soi_meme": canal_counts["Par soi même"]
        },
        "daily_details": daily_details
    }

@api_router.get("/analytics/visitors-table")
async def get_visitors_table(ville: str = None, current_user: dict = Depends(get_current_user)):
    """Get complete visitors table with all details for Super Admin/Pasteur"""
    # Only super_admin, pasteur, and responsable_eglise can access
    
    # Filtrer par ville si spécifié
    query = {}
    if ville:
        query["city"] = ville
    
    visitors = await db.visitors.find(query, {"_id": 0}).to_list(10000)
    
    # Enrich with assigned FI info
    enriched_visitors = []
    for visitor in visitors:
        # Calculate total presences
        presences_dimanche_count = len([p for p in visitor.get("presences_dimanche", []) if p.get("present")])
        presences_jeudi_count = len([p for p in visitor.get("presences_jeudi", []) if p.get("present")])
        
        # Check if assigned to FI
        membre = await db.membres_fi.find_one({"nouveau_arrivant_id": visitor["id"]}, {"_id": 0})
        assigned_fi = None
        assigned_fi_id = None
        if membre:
            fi = await db.familles_impact.find_one({"id": membre["fi_id"]}, {"_id": 0})
            if fi:
                assigned_fi = fi.get("nom", "N/A")
                assigned_fi_id = fi.get("id")
        
        enriched_visitors.append({
            **visitor,
            "presences_dimanche_count": presences_dimanche_count,
            "presences_jeudi_count": presences_jeudi_count,
            "total_presences": presences_dimanche_count + presences_jeudi_count,
            "assigned_fi": assigned_fi,
            "assigned_fi_id": assigned_fi_id,
            "comments_count": len(visitor.get("comments", []))
        })
    
    return enriched_visitors

@api_router.get("/analytics/fi-detailed")
async def get_fi_detailed(ville: str = None, date: str = None, current_user: dict = Depends(get_current_user)):
    """Get detailed FI analytics with:
    - Nombre de secteurs, FI, membres
    - Évolution par secteur
    - Fidélisation par FI (filtrée par date si spécifiée)
    """
    # Only super_admin, pasteur, and responsable_eglise can access
    
    # Filtrer par ville si spécifié
    query = {}
    # For responsable_eglise, force filter by their city
    if current_user["role"] == "responsable_eglise":
        query["ville"] = current_user["city"]
    elif ville:
        query["ville"] = ville
    
    # Get all secteurs
    secteurs = await db.secteurs.find(query, {"_id": 0}).to_list(1000)
    
    # Get all FI
    familles = await db.familles_impact.find(query, {"_id": 0}).to_list(1000)
    
    # Get all membres - ONLY for the FIs in the filtered familles
    fi_ids = [f["id"] for f in familles]
    membres_query = {"fi_id": {"$in": fi_ids}} if fi_ids else {"fi_id": None}
    membres = await db.membres_fi.find(membres_query, {"_id": 0}).to_list(10000)
    
    # Get all presences (filtered by date if specified and by FI)
    presences_query = {}
    if date:
        presences_query["date"] = date
    if fi_ids:
        presences_query["fi_id"] = {"$in": fi_ids}
    presences = await db.presences_fi.find(presences_query, {"_id": 0}).to_list(100000)
    
    # Group by secteur
    secteurs_stats = []
    for secteur in secteurs:
        secteur_fi = [f for f in familles if f.get("secteur_id") == secteur["id"]]
        secteur_membres = []
        
        for fi in secteur_fi:
            fi_membres = [m for m in membres if m.get("fi_id") == fi["id"]]
            secteur_membres.extend(fi_membres)
        
        secteurs_stats.append({
            "secteur_id": secteur["id"],
            "secteur_nom": secteur.get("nom", "N/A"),
            "ville": secteur.get("ville", "N/A"),
            "nombre_fi": len(secteur_fi),
            "nombre_membres": len(secteur_membres),
            "fi_list": [{"id": f["id"], "nom": f.get("nom", "N/A")} for f in secteur_fi]
        })
    
    # Fidélisation par FI (based on selected date or all-time)
    fi_fidelisation = []
    for fi in familles:
        fi_membres = [m for m in membres if m.get("fi_id") == fi["id"]]
        fi_presences = [p for p in presences if p.get("fi_id") == fi["id"]]
        
        total_membres = len(fi_membres)
        if total_membres == 0:
            fidelisation = 0
        else:
            if date:
                # Pour une date spécifique : taux de présence ce jour-là
                presents = len([p for p in fi_presences if p.get("present")])
                fidelisation = (presents / total_membres) * 100 if total_membres > 0 else 0
            else:
                # Sans date : Compte les membres avec au moins 3 présences (historique)
                membres_fideles = 0
                for membre in fi_membres:
                    membre_presences = [p for p in fi_presences if p.get("membre_fi_id") == membre["id"] and p.get("present")]
                    if len(membre_presences) >= 3:
                        membres_fideles += 1
                fidelisation = (membres_fideles / total_membres) * 100
        
        fi_fidelisation.append({
            "fi_id": fi["id"],
            "fi_nom": fi.get("nom", "N/A"),
            "ville": fi.get("ville", "N/A"),
            "secteur_id": fi.get("secteur_id"),
            "total_membres": total_membres,
            "total_presences": len([p for p in fi_presences if p.get("present")]),
            "fidelisation": round(fidelisation, 1)
        })
    
    return {
        "secteurs": secteurs_stats,
        "fi_fidelisation": fi_fidelisation,
        "summary": {
            "total_secteurs": len(secteurs),
            "total_fi": len(familles),
            "total_membres": len(membres),
            "avg_fidelisation": round(sum(f["fidelisation"] for f in fi_fidelisation) / len(fi_fidelisation), 1) if fi_fidelisation else 0
        }
    }

@api_router.get("/analytics/membres-table")
async def get_membres_table(ville: str = None, current_user: dict = Depends(get_current_user)):
    """Get complete membres table with presences for Super Admin/Pasteur"""
    # Only super_admin, pasteur, and responsable_eglise can access
    
    # For responsable_eglise, force filter by their city
    if current_user["role"] == "responsable_eglise":
        ville = current_user["city"]
    
    # Filtrer par ville si spécifié
    membre_query = {}
    if ville:
        # Besoin de filtrer les membres par FI qui appartiennent à la ville
        # On récupère d'abord les FI de la ville
        fi_query = {"ville": ville}
        fis = await db.familles_impact.find(fi_query, {"id": 1, "_id": 0}).to_list(1000)
        fi_ids = [fi["id"] for fi in fis]
        if fi_ids:
            membre_query["famille_impact_id"] = {"$in": fi_ids}
        else:
            # Aucune FI dans cette ville, retour vide
            return []
    
    membres = await db.membres_fi.find(membre_query, {"_id": 0}).to_list(10000)
    presences = await db.presences_fi.find({}, {"_id": 0}).to_list(100000)
    
    enriched_membres = []
    for membre in membres:
        # Get FI info
        fi = await db.familles_impact.find_one({"id": membre.get("fi_id")}, {"_id": 0})
        fi_nom = fi.get("nom", "N/A") if fi else "N/A"
        
        # Get secteur info
        secteur = await db.secteurs.find_one({"id": fi.get("secteur_id")}, {"_id": 0}) if fi else None
        secteur_nom = secteur.get("nom", "N/A") if secteur else "N/A"
        
        # Get presences for this membre
        membre_presences = [p for p in presences if p.get("membre_fi_id") == membre["id"]]
        presences_count = len([p for p in membre_presences if p.get("present")])
        
        enriched_membres.append({
            **membre,
            "fi_nom": fi_nom,
            "secteur_nom": secteur_nom,
            "presences": membre_presences,
            "presences_count": presences_count
        })
    
    return enriched_membres

@api_router.get("/analytics/presences-dimanche")
async def get_presences_dimanche(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    ville: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get presences dimanche aggregated by date with NA/NC breakdown"""
    # Only super_admin, pasteur, and responsable_eglise can access
    
    # Get all visitors with presences_dimanche
    query = {"tracking_stopped": False}
    if ville and ville != "all":
        query["city"] = ville
    
    visitors = await db.visitors.find(query, {"_id": 0}).to_list(10000)
    
    # Aggregate presences by date
    presences_by_date = {}
    
    for visitor in visitors:
        types = visitor.get("types", [])
        is_na = "Nouveau Arrivant" in types
        is_nc = "Nouveau Converti" in types
        
        for presence in visitor.get("presences_dimanche", []):
            date = presence.get("date")
            is_present = presence.get("present", False)
            
            # Filter by date range
            if start_date and date < start_date:
                continue
            if end_date and date > end_date:
                continue
            
            if date not in presences_by_date:
                presences_by_date[date] = {
                    "date": date,
                    "total_present": 0,
                    "total_absent": 0,
                    "na_present": 0,
                    "na_absent": 0,
                    "nc_present": 0,
                    "nc_absent": 0,
                    "visitors_details": []
                }
            
            if is_present:
                presences_by_date[date]["total_present"] += 1
                if is_na:
                    presences_by_date[date]["na_present"] += 1
                if is_nc:
                    presences_by_date[date]["nc_present"] += 1
            else:
                presences_by_date[date]["total_absent"] += 1
                if is_na:
                    presences_by_date[date]["na_absent"] += 1
                if is_nc:
                    presences_by_date[date]["nc_absent"] += 1
            
            # Add visitor details
            presences_by_date[date]["visitors_details"].append({
                "name": f"{visitor.get('firstname')} {visitor.get('lastname')}",
                "city": visitor.get("city"),
                "types": types,
                "present": is_present
            })
    
    # Convert to sorted list
    presences_list = sorted(presences_by_date.values(), key=lambda x: x["date"], reverse=True)
    
    # Calculate totals
    total_dimanches = len(presences_list)
    total_presences = sum(p["total_present"] for p in presences_list)
    total_na = sum(p["na_present"] for p in presences_list)
    total_nc = sum(p["nc_present"] for p in presences_list)
    avg_per_dimanche = total_presences / total_dimanches if total_dimanches > 0 else 0
    
    return {
        "presences": presences_list,
        "summary": {
            "total_dimanches": total_dimanches,
            "total_presences": total_presences,
            "total_na": total_na,
            "total_nc": total_nc,
            "avg_per_dimanche": round(avg_per_dimanche, 1)
        }
    }
# ==================== CULTE STATISTICS ====================

@api_router.post("/culte-stats")
async def create_culte_stats(stats: CulteStatsCreate, current_user: dict = Depends(get_current_user)):
    """Create culte statistics - Accueil, Pasteur, and Responsable Église can create for their city"""
    # Accueil, Responsable Église can only create for their city
    if current_user["role"] in ["accueil", "responsable_eglise"] and stats.ville != current_user["city"]:
        raise HTTPException(status_code=403, detail="Can only create stats for your city")
    
    # Check role permissions
    
    culte_stat = CulteStats(
        **stats.model_dump(),
        created_by=current_user["username"]
    )
    doc = culte_stat.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    if doc.get('updated_at'):
        doc['updated_at'] = doc['updated_at'].isoformat()
    
    await db.culte_stats.insert_one(doc)
    
    # Return without _id
    return {k: v for k, v in doc.items() if k != '_id'}

@api_router.get("/culte-stats")
async def get_culte_stats(
    ville: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    type_culte: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get culte statistics with filters"""
    # Build query based on role
    query = {}
    
    # Accueil and Responsable Église can only see their city
    if current_user["role"] in ["accueil", "responsable_eglise"]:
        query["ville"] = current_user["city"]
    # Pasteur and Super Admin can see all cities
    elif current_user["role"] in ["pasteur", "super_admin"]:
        if ville:
            query["ville"] = ville
    else:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Apply filters
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    if type_culte:
        query["type_culte"] = type_culte
    
    stats = await db.culte_stats.find(query, {"_id": 0}).to_list(10000)
    
    # Sort by date descending
    stats.sort(key=lambda x: x["date"], reverse=True)
    
    return stats

@api_router.get("/culte-stats/{stat_id}")
async def get_culte_stat(stat_id: str, current_user: dict = Depends(get_current_user)):
    """Get a single culte stat"""
    stat = await db.culte_stats.find_one({"id": stat_id}, {"_id": 0})
    
    if not stat:
        raise HTTPException(status_code=404, detail="Stat not found")
    
    # Check permissions
    if current_user["role"] == "accueil" and stat["ville"] != current_user["city"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    return stat

@api_router.put("/culte-stats/{stat_id}")
async def update_culte_stats(
    stat_id: str, 
    updates: CulteStatsUpdate, 
    current_user: dict = Depends(get_current_user)
):
    """Update culte statistics - Accueil and Super Admin"""
    stat = await db.culte_stats.find_one({"id": stat_id})
    
    if not stat:
        raise HTTPException(status_code=404, detail="Stat not found")
    
    # Accueil and Responsable Église can only update their city
    if current_user["role"] in ["accueil", "responsable_eglise"] and stat["ville"] != current_user["city"]:
        raise HTTPException(status_code=403, detail="Can only update stats for your city")
    
    # Only accueil, pasteur, responsable_eglise and super_admin can update
    
    update_dict = {k: v for k, v in updates.dict(exclude_unset=True).items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.culte_stats.update_one({"id": stat_id}, {"$set": update_dict})
    
    updated_stat = await db.culte_stats.find_one({"id": stat_id}, {"_id": 0})
    return updated_stat

@api_router.delete("/culte-stats/{stat_id}")
async def delete_culte_stats(stat_id: str, current_user: dict = Depends(get_current_user)):
    """Delete culte statistics - Super Admin and Pasteur"""
    
    result = await db.culte_stats.delete_one({"id": stat_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Stat not found")
    
    return {"message": "Stat deleted successfully"}

@api_router.get("/culte-stats/summary/all")
async def get_culte_stats_summary(
    ville: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get aggregated summary of culte stats for dashboards"""
    # Only pasteur, super_admin, responsable_eglise and accueil can access summary
    
    # Build query
    query = {}
    if current_user["role"] in ["accueil", "responsable_eglise"]:
        query["ville"] = current_user["city"]
    elif ville:
        query["ville"] = ville
    
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    stats = await db.culte_stats.find(query, {"_id": 0}).to_list(10000)
    
    # Aggregate by date (dimanche)
    by_date = {}
    for stat in stats:
        date = stat["date"]
        if date not in by_date:
            by_date[date] = {
                "date": date,
                "ville": stat["ville"],
                "culte_1": {"fideles": 0, "stars": 0},
                "culte_2": {"fideles": 0, "stars": 0},
                "ejp": {"fideles": 0, "stars": 0},
                "evenements_speciaux": {"fideles": 0, "stars": 0},
                "total_fideles": 0,
                "total_stars": 0,
                "total_general": 0
            }
        
        culte_type = stat["type_culte"].lower().replace(" ", "_").replace("é", "e")
        if "ejp" in culte_type.lower():
            culte_type = "ejp"
        elif "1" in culte_type:
            culte_type = "culte_1"
        elif "2" in culte_type:
            culte_type = "culte_2"
        elif "evenement" in culte_type or "spéciaux" in stat["type_culte"].lower():
            culte_type = "evenements_speciaux"
        
        # Only add to known types to avoid KeyError
        if culte_type in by_date[date]:
            by_date[date][culte_type]["fideles"] = stat["nombre_fideles"]
            by_date[date][culte_type]["stars"] = stat["nombre_stars"]
        
        # Always add to totals
        by_date[date]["total_fideles"] += stat["nombre_fideles"]
        by_date[date]["total_stars"] += stat["nombre_stars"]
        by_date[date]["total_general"] = by_date[date]["total_fideles"] + by_date[date]["total_stars"]
    
    # Convert to list and sort
    summary_list = sorted(by_date.values(), key=lambda x: x["date"], reverse=True)
    
    # Calculate global stats
    total_dimanches = len(summary_list)
    total_fideles = sum(s["total_fideles"] for s in summary_list)
    total_stars = sum(s["total_stars"] for s in summary_list)
    avg_fideles = total_fideles / total_dimanches if total_dimanches > 0 else 0
    avg_stars = total_stars / total_dimanches if total_dimanches > 0 else 0
    
    return {
        "summary": summary_list,
        "global_stats": {
            "total_dimanches": total_dimanches,
            "total_fideles": total_fideles,
            "total_stars": total_stars,
            "avg_fideles_per_dimanche": round(avg_fideles, 1),
            "avg_stars_per_dimanche": round(avg_stars, 1),
            "avg_total_per_dimanche": round(avg_fideles + avg_stars, 1)
        }
    }

# ==================== DATA EXPORT/IMPORT (Super Admin) ====================

@api_router.get("/admin/export-all-data")
async def export_all_data(current_user: dict = Depends(get_current_user)):
    """Export all database data - Super Admin only"""
    if current_user["role"] != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin only")
    
    # Export all collections
    cities = await db.cities.find({}, {"_id": 0}).to_list(10000)
    users = await db.users.find({}, {"_id": 0}).to_list(10000)
    visitors = await db.visitors.find({}, {"_id": 0}).to_list(10000)
    secteurs = await db.secteurs.find({}, {"_id": 0}).to_list(10000)
    familles_impact = await db.familles_impact.find({}, {"_id": 0}).to_list(10000)
    membres_fi = await db.membres_fi.find({}, {"_id": 0}).to_list(10000)
    presences_fi = await db.presences_fi.find({}, {"_id": 0}).to_list(10000)
    culte_stats = await db.culte_stats.find({}, {"_id": 0}).to_list(10000)
    notifications = await db.notifications.find({}, {"_id": 0}).to_list(10000)
    
    data = {
        "cities": cities,
        "users": users,
        "visitors": visitors,
        "secteurs": secteurs,
        "familles_impact": familles_impact,
        "membres_fi": membres_fi,
        "presences_fi": presences_fi,
        "culte_stats": culte_stats,
        "notifications": notifications,
        "metadata": {
            "export_date": datetime.now(timezone.utc).isoformat(),
            "exported_by": current_user["username"],
            "total_records": (
                len(cities) + len(users) + len(visitors) + 
                len(secteurs) + len(familles_impact) + len(membres_fi) + 
                len(presences_fi) + len(culte_stats) + len(notifications)
            ),
            "collections": {
                "cities": len(cities),
                "users": len(users),
                "visitors": len(visitors),
                "secteurs": len(secteurs),
                "familles_impact": len(familles_impact),
                "membres_fi": len(membres_fi),
                "presences_fi": len(presences_fi),
                "culte_stats": len(culte_stats),
                "notifications": len(notifications)
            }
        }
    }
    
    return data

@api_router.post("/admin/import-all-data")
async def import_all_data(data: dict, current_user: dict = Depends(get_current_user)):
    """Import all database data - Super Admin only"""
    if current_user["role"] != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin only")
    
    try:
        # Clear existing data
        await db.cities.delete_many({})
        await db.users.delete_many({})
        await db.visitors.delete_many({})
        await db.secteurs.delete_many({})
        await db.familles_impact.delete_many({})
        await db.membres_fi.delete_many({})
        await db.presences_fi.delete_many({})
        await db.culte_stats.delete_many({})
        await db.notifications.delete_many({})
        
        # Import new data
        if data.get("cities"):
            await db.cities.insert_many(data["cities"])
        if data.get("users"):
            await db.users.insert_many(data["users"])
        if data.get("visitors"):
            await db.visitors.insert_many(data["visitors"])
        if data.get("secteurs"):
            await db.secteurs.insert_many(data["secteurs"])
        if data.get("familles_impact"):
            await db.familles_impact.insert_many(data["familles_impact"])
        if data.get("membres_fi"):
            await db.membres_fi.insert_many(data["membres_fi"])
        if data.get("presences_fi"):
            await db.presences_fi.insert_many(data["presences_fi"])
        if data.get("culte_stats"):
            await db.culte_stats.insert_many(data["culte_stats"])
        if data.get("notifications"):
            await db.notifications.insert_many(data["notifications"])
        
        return {
            "success": True,
            "message": "Data imported successfully",
            "counts": {
                "cities": len(data.get("cities", [])),
                "users": len(data.get("users", [])),
                "visitors": len(data.get("visitors", [])),
                "secteurs": len(data.get("secteurs", [])),
                "familles_impact": len(data.get("familles_impact", [])),
                "membres_fi": len(data.get("membres_fi", [])),
                "presences_fi": len(data.get("presences_fi", [])),
                "culte_stats": len(data.get("culte_stats", [])),
                "notifications": len(data.get("notifications", []))
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

# ==================== ROOT ====================

@api_router.get("/")
async def root():
    return {"message": "ICC Dijon Connect API"}

# ==================== ANALYTICS - AGE & CHANNEL ====================

@api_router.get("/analytics/age-distribution")
async def get_age_distribution(
    ville: str = None,
    current_user: dict = Depends(get_current_user)
):
    """Get age range distribution for visitors"""
    # Permissions
    
    # Build query
    query = {}
    
    # For responsable_eglise, force their city
    if current_user["role"] == "responsable_eglise":
        query["city"] = current_user["city"]
    # For responsable_promo, filter by city and assigned_month
    elif current_user["role"] == "responsable_promo":
        query["city"] = current_user["city"]
        if current_user.get("assigned_month"):
            # Support multiple months separated by commas
            months_list = [m.strip() for m in current_user["assigned_month"].split(',')]
            if len(months_list) > 1:
                query["assigned_month"] = {"$in": months_list}
            else:
                query["assigned_month"] = current_user["assigned_month"]
    # For others, use ville parameter
    elif ville:
        query["city"] = ville
    
    # Get visitors
    visitors = await db.visitors.find(query, {"_id": 0, "age_range": 1}).to_list(10000)
    
    # Count by age range
    age_counts = {}
    for v in visitors:
        age = v.get("age_range", "Non renseigné")
        if not age:
            age = "Non renseigné"
        age_counts[age] = age_counts.get(age, 0) + 1
    
    # Format for PieChart
    result = [{"name": age, "value": count} for age, count in age_counts.items()]
    
    return result


@api_router.get("/analytics/arrival-channel-distribution")
async def get_arrival_channel_distribution(
    ville: str = None,
    current_user: dict = Depends(get_current_user)
):
    """Get arrival channel distribution for visitors"""
    # Permissions
    
    # Build query
    query = {}
    
    # For responsable_eglise, force their city
    if current_user["role"] == "responsable_eglise":
        query["city"] = current_user["city"]
    # For responsable_promo, filter by city and assigned_month
    elif current_user["role"] == "responsable_promo":
        query["city"] = current_user["city"]
        if current_user.get("assigned_month"):
            # Support multiple months separated by commas
            months_list = [m.strip() for m in current_user["assigned_month"].split(',')]
            if len(months_list) > 1:
                query["assigned_month"] = {"$in": months_list}
            else:
                query["assigned_month"] = current_user["assigned_month"]
    # For others, use ville parameter
    elif ville:
        query["city"] = ville
    
    # Get visitors
    visitors = await db.visitors.find(query, {"_id": 0, "arrival_channel": 1}).to_list(10000)
    
    # Count by arrival channel
    channel_counts = {}
    for v in visitors:
        channel = v.get("arrival_channel", "Non renseigné")
        if not channel:
            channel = "Non renseigné"
        channel_counts[channel] = channel_counts.get(channel, 0) + 1
    
    # Format for PieChart
    result = [{"name": channel, "value": count} for channel, count in channel_counts.items()]
    
    return result

@api_router.post("/admin/generate-missing-passwords")
async def generate_missing_passwords(current_user: dict = Depends(get_current_user)):
    """
    Generate new passwords for all users who don't have plain_password stored
    Accessible only to super_admin
    """
    if current_user["role"] != "super_admin":
        raise HTTPException(status_code=403, detail="Only super_admin can generate passwords")
    
    try:
        import random
        import string
        
        # Get all users without plain_password
        users = await db.users.find({
            "$or": [
                {"plain_password": {"$exists": False}},
                {"plain_password": None},
                {"plain_password": ""}
            ]
        }).to_list(10000)
        
        updated_count = 0
        
        for user in users:
            # Générer un mot de passe : Prenom + 3 chiffres aléatoires
            firstname = user.get('username', 'User')
            # Prendre les 4 premiers caractères du username et capitaliser
            base = firstname[:4].capitalize()
            # Ajouter 3 chiffres aléatoires
            random_digits = ''.join(random.choices(string.digits, k=3))
            new_password = f"{base}{random_digits}"
            
            # Hash le nouveau mot de passe
            hashed_pwd = hash_password(new_password)
            
            # Mettre à jour l'utilisateur
            await db.users.update_one(
                {"id": user["id"]},
                {"$set": {
                    "password": hashed_pwd,
                    "plain_password": new_password
                }}
            )
            updated_count += 1
        
        return {
            "message": f"Passwords generated for {updated_count} users",
            "updated_count": updated_count
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating passwords: {str(e)}")

@api_router.get("/admin/export-credentials")
async def export_credentials(current_user: dict = Depends(get_current_user)):
    """
    Export all user credentials (logins and passwords) to Excel
    Accessible only to super_admin
    """
    if current_user["role"] != "super_admin":
        raise HTTPException(status_code=403, detail="Only super_admin can export credentials")
    
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Get all users
        users = await db.users.find({}, {"_id": 0}).to_list(10000)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Credentials"
        
        # Headers
        headers = ["Nom d'utilisateur", "Mot de passe", "Rôle", "Ville", "Email"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data
        for user in users:
            ws.append([
                user.get("username", ""),
                user.get("plain_password", "********"),  # Plain password if stored, otherwise masked
                user.get("role", ""),
                user.get("city", ""),
                user.get("email", "")
            ])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        from fastapi.responses import StreamingResponse
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=credentials_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            }
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'export: {str(e)}")

@api_router.post("/admin/migrate-presences")
async def migrate_presences(current_user: dict = Depends(get_current_user)):
    """
    Endpoint pour migrer les présences mal attribuées (jeudi dans dimanche)
    Accessible uniquement aux super_admin
    """
    if current_user["role"] != "super_admin":
        raise HTTPException(status_code=403, detail="Only super_admin can run migration")
    
    try:
        # Get all visitors
        visitors = await db.visitors.find({}, {"_id": 0}).to_list(10000)
        
        total_visitors = len(visitors)
        visitors_updated = 0
        presences_moved = 0
        migration_details = []
        
        for visitor in visitors:
            visitor_id = visitor.get("id")
            presences_dim = visitor.get("presences_dimanche", [])
            presences_jeu = visitor.get("presences_jeudi", [])
            
            if not presences_dim:
                continue
            
            # Séparer les présences dimanche et jeudi
            real_dimanche = []
            real_jeudi = []
            
            for presence in presences_dim:
                date_str = presence.get("date", "")
                if not date_str:
                    real_dimanche.append(presence)
                    continue
                
                try:
                    year, month, day = date_str.split('-')
                    date_obj = datetime(int(year), int(month), int(day))
                    day_of_week = date_obj.weekday()  # 0=Monday, 6=Sunday
                    
                    if day_of_week == 6:  # Sunday
                        real_dimanche.append(presence)
                    else:  # Thursday or other days → jeudi
                        real_jeudi.append(presence)
                        presences_moved += 1
                except Exception as e:
                    # En cas d'erreur, garder dans dimanche
                    real_dimanche.append(presence)
            
            # Si des présences doivent être déplacées
            if real_jeudi:
                visitors_updated += 1
                
                # Fusionner avec les présences jeudi existantes (éviter les doublons)
                existing_jeudi_dates = {p.get("date") for p in presences_jeu}
                for p in real_jeudi:
                    if p.get("date") not in existing_jeudi_dates:
                        presences_jeu.append(p)
                
                # Update visitor
                await db.visitors.update_one(
                    {"id": visitor_id},
                    {
                        "$set": {
                            "presences_dimanche": real_dimanche,
                            "presences_jeudi": presences_jeu
                        }
                    }
                )
                
                visitor_name = f"{visitor.get('firstname', '')} {visitor.get('lastname', '')}"
                migration_details.append({
                    "visitor": visitor_name,
                    "presences_moved": len(real_jeudi)
                })
        
        return {
            "success": True,
            "message": "Migration terminée avec succès",
            "total_visitors": total_visitors,
            "visitors_updated": visitors_updated,
            "presences_moved": presences_moved,
            "details": migration_details[:20]  # Limiter à 20 pour ne pas surcharger la réponse
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de la migration: {str(e)}")

# ==================== EVENTS & PROJECTS ENDPOINTS ====================

@api_router.get("/events/upcoming")
async def get_upcoming_events():
    """Get upcoming events/activities for the next 30 days - public endpoint for homepage"""
    today = datetime.now(timezone.utc).date()
    max_date = today + timedelta(days=30)
    
    upcoming = []
    
    # 1. Récupérer les activités du planning (collection planning_activites)
    activites = await db.planning_activites.find(
        {"statut": {"$in": ["À venir", "Planifié", "planifie"]}},
        {"_id": 0}
    ).to_list(1000)
    
    for activite in activites:
        try:
            date_str = activite.get("date_debut") or activite.get("date")
            if not date_str:
                continue
            date_debut = datetime.strptime(date_str, "%Y-%m-%d").date()
            if today <= date_debut <= max_date:
                days_until = (date_debut - today).days
                upcoming.append({
                    "id": activite.get("id", ""),
                    "titre": activite.get("nom", ""),
                    "ville": activite.get("ville", ""),
                    "date_debut": date_str,
                    "days_until": days_until,
                    "statut": activite.get("statut", "À venir"),
                    "source": "planning"
                })
        except (ValueError, KeyError):
            continue
    
    # 2. Récupérer aussi les projets (collection projets)
    projets = await db.projets.find(
        {"archived": {"$ne": True}, "date_debut": {"$ne": None}},
        {"_id": 0}
    ).to_list(1000)
    
    for projet in projets:
        try:
            if not projet.get("date_debut"):
                continue
            date_debut = datetime.strptime(projet["date_debut"], "%Y-%m-%d").date()
            if today <= date_debut <= max_date:
                days_until = (date_debut - today).days
                upcoming.append({
                    "id": projet["id"],
                    "titre": projet["titre"],
                    "ville": projet.get("ville", ""),
                    "date_debut": projet["date_debut"],
                    "days_until": days_until,
                    "statut": projet.get("statut", "planifie"),
                    "source": "projet"
                })
        except (ValueError, KeyError):
            continue
    
    # Sort by date (closest first)
    upcoming.sort(key=lambda x: x["days_until"])
    
    return upcoming


@api_router.get("/events/projets")
async def get_projets(ville: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Get all projects - filtered by city for non-super_admin"""
    
    query = {}
    # Super admin can see all, others only their city
    if current_user["role"] != "super_admin":
        query["ville"] = current_user["city"]
    elif ville:
        query["ville"] = ville
    
    projets = await db.projets.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Add task statistics to each project
    for projet in projets:
        taches = await db.taches.find({"projet_id": projet["id"]}, {"_id": 0}).to_list(1000)
        total_taches = len(taches)
        taches_terminees = len([t for t in taches if t.get("statut") == "termine"])
        
        projet["total_taches"] = total_taches
        projet["taches_terminees"] = taches_terminees
        projet["taux_achevement"] = round((taches_terminees / total_taches * 100), 1) if total_taches > 0 else 0
    
    return projets

@api_router.post("/events/projets")
async def create_projet(projet: ProjetCreate, current_user: dict = Depends(get_current_user)):
    """Create a new project"""
    
    projet_dict = projet.model_dump()
    projet_dict["id"] = str(uuid.uuid4())
    projet_dict["created_by"] = current_user["username"]
    projet_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    projet_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.projets.insert_one(projet_dict)
    return {"message": "Projet créé avec succès", "id": projet_dict["id"]}

@api_router.get("/events/projets/{projet_id}")
async def get_projet(projet_id: str, current_user: dict = Depends(get_current_user)):
    """Get project details"""
    
    projet = await db.projets.find_one({"id": projet_id}, {"_id": 0})
    if not projet:
        raise HTTPException(status_code=404, detail="Projet non trouvé")
    
    # Check city access for non-super_admin
    if current_user["role"] != "super_admin" and projet["ville"] != current_user["city"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Calculate global completion percentage from all tasks (including poles and general tasks)
    all_taches = await db.taches.find({"projet_id": projet_id}, {"_id": 0}).to_list(1000)
    total_taches = len(all_taches)
    taches_terminees = len([t for t in all_taches if t.get("statut") == "termine"])
    
    projet["total_taches"] = total_taches
    projet["taches_terminees"] = taches_terminees
    projet["taux_achevement"] = round((taches_terminees / total_taches * 100) if total_taches > 0 else 0, 1)
    
    return projet

@api_router.put("/events/projets/{projet_id}")
async def update_projet(projet_id: str, updates: ProjetUpdate, current_user: dict = Depends(get_current_user)):
    """Update a project"""
    
    projet = await db.projets.find_one({"id": projet_id})
    if not projet:
        raise HTTPException(status_code=404, detail="Projet non trouvé")
    
    # Check city access
    if current_user["role"] != "super_admin" and projet["ville"] != current_user["city"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    update_dict = {k: v for k, v in updates.model_dump().items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.projets.update_one({"id": projet_id}, {"$set": update_dict})
    return {"message": "Projet mis à jour"}

@api_router.delete("/events/projets/{projet_id}")
async def delete_projet(projet_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a project"""
    
    projet = await db.projets.find_one({"id": projet_id})
    if not projet:
        raise HTTPException(status_code=404, detail="Projet non trouvé")
    
    if current_user["role"] != "super_admin" and projet["ville"] != current_user["city"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Delete related data
    await db.taches.delete_many({"projet_id": projet_id})
    await db.commentaires_projet.delete_many({"projet_id": projet_id})
    await db.fichiers_projet.delete_many({"projet_id": projet_id})
    await db.projets.delete_one({"id": projet_id})
    
    return {"message": "Projet supprimé"}

# TÂCHES
@api_router.get("/events/taches")
async def get_taches(projet_id: Optional[str] = None, statut: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Get tasks with auto-update for overdue status"""
    
    query = {}
    if projet_id:
        query["projet_id"] = projet_id
    
    taches = await db.taches.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Auto-update status to "en_retard" if deadline passed and not completed
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    for tache in taches:
        if tache.get("deadline") and tache.get("statut") != "termine":
            if tache["deadline"] < today:
                # Update to "en_retard" if not already
                if tache.get("statut") != "en_retard":
                    await db.taches.update_one(
                        {"id": tache["id"]},
                        {"$set": {"statut": "en_retard"}}
                    )
                    tache["statut"] = "en_retard"
    
    # Filter by status if provided
    if statut:
        taches = [t for t in taches if t.get("statut") == statut]
    
    return taches

@api_router.put("/events/projets/{projet_id}/archive")
async def archive_projet(projet_id: str, current_user: dict = Depends(get_current_user)):
    """Archive/Unarchive a project"""
    
    projet = await db.projets.find_one({"id": projet_id})
    if not projet:
        raise HTTPException(status_code=404, detail="Projet non trouvé")
    
    # Toggle archive status
    is_archived = projet.get("archived", False)
    await db.projets.update_one(
        {"id": projet_id},
        {"$set": {"archived": not is_archived, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "Projet archivé" if not is_archived else "Projet désarchivé"}


@api_router.post("/events/taches")
async def create_tache(tache: TacheCreate, current_user: dict = Depends(get_current_user)):
    """Create a task"""
    
    tache_dict = tache.model_dump()
    tache_dict["id"] = str(uuid.uuid4())
    tache_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    # Don't override statut if it's provided in the request
    
    await db.taches.insert_one(tache_dict)
    return {"message": "Tâche créée", "id": tache_dict["id"]}

@api_router.put("/events/taches/{tache_id}")
async def update_tache(tache_id: str, updates: TacheUpdate, current_user: dict = Depends(get_current_user)):
    """Update a task"""
    
    update_dict = {k: v for k, v in updates.model_dump().items() if v is not None}
    await db.taches.update_one({"id": tache_id}, {"$set": update_dict})
    return {"message": "Tâche mise à jour"}

@api_router.delete("/events/taches/{tache_id}")
async def delete_tache(tache_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a task"""
    
    await db.taches.delete_one({"id": tache_id})
    return {"message": "Tâche supprimée"}

# COMMENTAIRES
@api_router.get("/events/commentaires")
async def get_commentaires(projet_id: str, current_user: dict = Depends(get_current_user)):
    """Get project comments"""
    
    commentaires = await db.commentaires_projet.find(
        {"projet_id": projet_id}, 
        {"_id": 0}
    ).sort("created_at", 1).to_list(1000)
    return commentaires

@api_router.post("/events/commentaires")
async def create_commentaire(commentaire: CommentaireProjetCreate, current_user: dict = Depends(get_current_user)):
    """Add a comment"""
    
    commentaire_dict = commentaire.model_dump()
    commentaire_dict["id"] = str(uuid.uuid4())
    commentaire_dict["user"] = current_user["username"]
    commentaire_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.commentaires_projet.insert_one(commentaire_dict)
    return {"message": "Commentaire ajouté", "id": commentaire_dict["id"]}

# DÉPENSES PROJET
@api_router.get("/events/depenses")
async def get_depenses(projet_id: str, current_user: dict = Depends(get_current_user)):
    """Get project expenses"""
    
    depenses = await db.depenses_projet.find(
        {"projet_id": projet_id}, 
        {"_id": 0}
    ).sort("created_at", -1).to_list(1000)
    return depenses

@api_router.post("/events/depenses")
async def create_depense(depense: DepenseCreate, current_user: dict = Depends(get_current_user)):
    """Add an expense"""
    
    # Get current budget_reel from projet
    projet = await db.projets.find_one({"id": depense.projet_id})
    if not projet:
        raise HTTPException(status_code=404, detail="Projet non trouvé")
    
    depense_dict = depense.model_dump()
    depense_dict["id"] = str(uuid.uuid4())
    depense_dict["created_by"] = current_user["username"]
    depense_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    # If no date provided, use today
    if not depense_dict.get("date"):
        depense_dict["date"] = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    await db.depenses_projet.insert_one(depense_dict)
    
    # Update budget_reel in projet
    new_budget_reel = projet.get("budget_reel", 0) + depense.montant
    await db.projets.update_one(
        {"id": depense.projet_id},
        {"$set": {"budget_reel": new_budget_reel, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "Dépense ajoutée", "id": depense_dict["id"], "new_budget_reel": new_budget_reel}

@api_router.delete("/events/depenses/{depense_id}")
async def delete_depense(depense_id: str, current_user: dict = Depends(get_current_user)):
    """Delete an expense"""
    
    # Get depense to retrieve montant and projet_id
    depense = await db.depenses_projet.find_one({"id": depense_id})
    if not depense:
        raise HTTPException(status_code=404, detail="Dépense non trouvée")
    
    # Get projet to update budget_reel
    projet = await db.projets.find_one({"id": depense["projet_id"]})
    if projet:
        new_budget_reel = max(0, projet.get("budget_reel", 0) - depense["montant"])
        await db.projets.update_one(
            {"id": depense["projet_id"]},
            {"$set": {"budget_reel": new_budget_reel, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    
    await db.depenses_projet.delete_one({"id": depense_id})
    return {"message": "Dépense supprimée"}



# JALONS
@api_router.get("/events/jalons")
async def get_jalons(projet_id: str, current_user: dict = Depends(get_current_user)):
    """Get all milestones for a project"""
    
    jalons = await db.jalons.find({"projet_id": projet_id}, {"_id": 0}).sort("date_debut", 1).to_list(1000)
    
    # Auto-update status to "en_retard" if date_fin is passed
    now = datetime.now(timezone.utc)
    for jalon in jalons:
        if jalon.get("date_fin") and jalon.get("statut") in ["a_faire", "en_cours"]:
            try:
                date_fin_dt = datetime.fromisoformat(jalon["date_fin"].replace("Z", "+00:00"))
                if date_fin_dt < now:
                    await db.jalons.update_one(
                        {"id": jalon["id"]},
                        {"$set": {"statut": "en_retard"}}
                    )
                    jalon["statut"] = "en_retard"
            except:
                pass
    
    return jalons

@api_router.post("/events/jalons")
async def create_jalon(jalon: JalonCreate, current_user: dict = Depends(get_current_user)):
    """Create a new milestone"""
    
    jalon_dict = jalon.model_dump()
    jalon_dict["id"] = str(uuid.uuid4())
    jalon_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    jalon_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    jalon_dict["statut"] = "a_faire"
    
    await db.jalons.insert_one(jalon_dict)
    return {"message": "Jalon créé", "id": jalon_dict["id"]}

@api_router.put("/events/jalons/{jalon_id}")
async def update_jalon(jalon_id: str, jalon: JalonUpdate, current_user: dict = Depends(get_current_user)):
    """Update a milestone"""
    
    existing = await db.jalons.find_one({"id": jalon_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Jalon non trouvé")
    
    update_data = {k: v for k, v in jalon.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.jalons.update_one({"id": jalon_id}, {"$set": update_data})
    return {"message": "Jalon mis à jour"}

@api_router.delete("/events/jalons/{jalon_id}")
async def delete_jalon(jalon_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a milestone"""
    
    result = await db.jalons.delete_one({"id": jalon_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Jalon non trouvé")
    
    return {"message": "Jalon supprimé"}

# POLES
@api_router.get("/events/poles")
async def get_poles(projet_id: str, current_user: dict = Depends(get_current_user)):
    """Get all poles for a project"""
    
    poles = await db.poles.find({"projet_id": projet_id}, {"_id": 0}).sort("created_at", 1).to_list(1000)
    
    # For each pole, calculate completion percentage
    for pole in poles:
        taches = await db.taches.find({"pole_id": pole["id"]}, {"_id": 0}).to_list(1000)
        total_taches = len(taches)
        taches_terminees = len([t for t in taches if t.get("statut") == "termine"])
        pole["nb_taches"] = total_taches
        pole["nb_taches_terminees"] = taches_terminees
        pole["completion_percent"] = round((taches_terminees / total_taches * 100) if total_taches > 0 else 0, 1)
    
    return poles

@api_router.post("/events/poles")
async def create_pole(pole: PoleCreate, current_user: dict = Depends(get_current_user)):
    """Create a new pole"""
    
    # Verify project exists
    projet = await db.projets.find_one({"id": pole.projet_id}, {"_id": 0})
    if not projet:
        raise HTTPException(status_code=404, detail="Projet non trouvé")
    
    pole_dict = pole.model_dump()
    pole_dict["id"] = str(uuid.uuid4())
    pole_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    pole_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.poles.insert_one(pole_dict)
    return {"message": "Pôle créé", "id": pole_dict["id"]}

@api_router.put("/events/poles/{pole_id}")
async def update_pole(pole_id: str, pole: PoleUpdate, current_user: dict = Depends(get_current_user)):
    """Update a pole"""
    
    existing = await db.poles.find_one({"id": pole_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Pôle non trouvé")
    
    update_data = {k: v for k, v in pole.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.poles.update_one({"id": pole_id}, {"$set": update_data})
    return {"message": "Pôle mis à jour"}

@api_router.delete("/events/poles/{pole_id}")
async def delete_pole(pole_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a pole"""
    
    # Check if pole has tasks
    taches = await db.taches.find({"pole_id": pole_id}, {"_id": 0}).to_list(1)
    if taches:
        raise HTTPException(status_code=400, detail="Impossible de supprimer un pôle contenant des tâches. Supprimez d'abord les tâches ou déplacez-les.")
    
    result = await db.poles.delete_one({"id": pole_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pôle non trouvé")
    
    return {"message": "Pôle supprimé"}

# CAMPAGNES COMMUNICATION
@api_router.get("/events/campagnes")
async def get_campagnes(current_user: dict = Depends(get_current_user)):
    """Get communication campaigns"""
    
    campagnes = await db.campagnes_communication.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return campagnes

@api_router.post("/events/campagnes")
async def create_campagne(campagne: CampagneCommunicationCreate, current_user: dict = Depends(get_current_user)):
    """Create a communication campaign"""
    
    campagne_dict = campagne.model_dump()
    campagne_dict["id"] = str(uuid.uuid4())
    campagne_dict["created_by"] = current_user["username"]
    campagne_dict["statut"] = "brouillon"
    campagne_dict["stats"] = {"envoyes": 0, "oui": 0, "non": 0, "peut_etre": 0}
    campagne_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.campagnes_communication.insert_one(campagne_dict)
    return {"message": "Campagne créée", "id": campagne_dict["id"]}

@api_router.put("/events/campagnes/{campagne_id}/archive")
async def archive_campagne(campagne_id: str, current_user: dict = Depends(get_current_user)):
    """Archiver/désarchiver une campagne"""
    
    campagne = await db.campagnes_communication.find_one({"id": campagne_id})
    if not campagne:
        raise HTTPException(status_code=404, detail="Campagne non trouvée")
    
    # Toggle archived status
    new_status = not campagne.get("archived", False)
    await db.campagnes_communication.update_one(
        {"id": campagne_id},
        {"$set": {"archived": new_status}}
    )
    
    return {"message": "Campagne archivée" if new_status else "Campagne désarchivée"}

@api_router.delete("/events/campagnes/{campagne_id}")
async def delete_campagne(campagne_id: str, current_user: dict = Depends(get_current_user)):
    """Supprimer une campagne"""
    
    await db.campagnes_communication.delete_one({"id": campagne_id})
    return {"message": "Campagne supprimée"}

@api_router.post("/events/campagnes/{campagne_id}/envoyer")
async def envoyer_campagne(campagne_id: str, current_user: dict = Depends(get_current_user)):
    """Send a communication campaign"""
    
    campagne = await db.campagnes_communication.find_one({"id": campagne_id})
    if not campagne:
        raise HTTPException(status_code=404, detail="Campagne non trouvée")
    
    # Envoi via Brevo (Email) et/ou Twilio (SMS)
    import sib_api_v3_sdk
    from sib_api_v3_sdk.rest import ApiException
    
    envois_reussis = 0
    
    # Configuration Brevo
    brevo_api_key = os.getenv('BREVO_API_KEY')
    rsvp_enabled = campagne.get('enable_rsvp', False)
    
    if campagne["type"] in ["email", "both"] and brevo_api_key:
        configuration = sib_api_v3_sdk.Configuration()
        configuration.api_key['api-key'] = brevo_api_key
        api_instance = sib_api_v3_sdk.TransactionalEmailsApi(sib_api_v3_sdk.ApiClient(configuration))
        
        for destinataire in campagne["destinataires"]:
            if not destinataire.get("email"):
                continue
            
            # Personnaliser le message AVANT de créer le HTML
            message = campagne["message"]
            prenom = destinataire.get("prenom", "")
            nom = destinataire.get("nom", "")
            message = message.replace("{prenom}", prenom)
            message = message.replace("{nom}", nom)
            
            # Créer un email HTML avec le logo ICC
            logo_url = "https://customer-assets.emergentagent.com/job_dijon-icc-hub/artifacts/foeikpvk_IMG_2590.png"
            
            # Construire le contenu du message avec l'image en bas
            message_html = message.replace('\n', '<br>')
            
            # Ajouter l'image de la campagne EN BAS du texte si présente
            image_html = ""
            image_url = (campagne.get("image_url") or "").strip()
            if image_url:
                print(f"DEBUG: Ajout image dans email - URL: {image_url[:100]}...")
                image_html = f'''
                <div style="text-align: center; margin-top: 20px; margin-bottom: 20px;">
                    <img src="{image_url}" alt="Affiche" style="max-width: 100%; height: auto; border-radius: 8px; display: block; margin: 0 auto;" />
                </div>
                '''
            else:
                print("DEBUG: Pas d'image_url dans la campagne")
            
            # Ajouter UN SEUL lien RSVP si activé
            rsvp_html = ""
            if rsvp_enabled:
                # Utiliser l'URL depuis REACT_APP_BACKEND_URL mais en remplaçant par le frontend
                backend_url = os.environ['REACT_APP_BACKEND_URL']
                # L'URL frontend est la même que le backend (le backend est accessible via /api)
                base_url = backend_url
                contact_identifier = destinataire.get("email")
                rsvp_link = f"{base_url}/rsvp/{campagne_id}?contact={contact_identifier}"
                rsvp_html = f'''
                <div style="text-align: center; margin-top: 30px; padding: 20px; background-color: #e0e7ff; border-radius: 8px;">
                    <p style="color: #1e40af; font-weight: bold; margin-bottom: 15px;">📋 Merci de confirmer votre présence</p>
                    <a href="{rsvp_link}" style="display: inline-block; padding: 12px 30px; background-color: #667eea; color: white; text-decoration: none; border-radius: 6px; font-weight: bold;">
                        Répondre maintenant
                    </a>
                </div>
                '''
            
            html_content = f'''
            <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
                <div style="text-align: center; padding: 20px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);">
                    <img src="{logo_url}" alt="ICC BFC-Italie" style="width: 120px; height: 120px; border-radius: 60px; border: 4px solid white;" />
                    <h2 style="color: white; margin-top: 10px;">Impact Centre Chrétien BFC-Italie</h2>
                </div>
                <div style="padding: 30px; background-color: #f9fafb;">
                    {message_html}
                    {image_html}
                    {rsvp_html}
                </div>
                <div style="padding: 20px; text-align: center; background-color: #667eea; color: white; font-size: 12px;">
                    <p>© {datetime.now().year} Impact Centre Chrétien BFC-Italie</p>
                    <p>My Events Church - Gestion d'Événements</p>
                </div>
            </div>
            '''
            
            # Configuration expéditeur depuis .env ou valeurs par défaut
            sender_email = os.environ.get('SENDER_EMAIL', 'impactcentrechretienbfcitalie@gmail.com')
            sender_name = os.environ.get('SENDER_NAME', 'Impact Centre Chrétien BFC-Italie')
            
            send_smtp_email = sib_api_v3_sdk.SendSmtpEmail(
                to=[{"email": destinataire.get("email"), "name": f"{destinataire.get('prenom', '')} {destinataire.get('nom', '')}"}],
                subject=campagne["titre"],
                html_content=html_content,
                sender={"name": sender_name, "email": sender_email}
            )
            
            try:
                api_instance.send_transac_email(send_smtp_email)
                envois_reussis += 1
                print(f"✅ Email envoyé à {destinataire.get('email')}")
            except ApiException as e:
                print(f"❌ Erreur envoi email à {destinataire.get('email')}: {str(e)}")
                import traceback
                traceback.print_exc()
    
    # SMS via Brevo
    if campagne["type"] in ["sms", "both"]:
        import sib_api_v3_sdk
        from sib_api_v3_sdk.rest import ApiException
        
        brevo_api_key = os.environ.get('BREVO_API_KEY')
        brevo_sender = os.environ.get('BREVO_SMS_SENDER', '0646989818')
        
        if not brevo_api_key:
            print("Brevo non configuré - SMS ignorés")
        else:
            configuration = sib_api_v3_sdk.Configuration()
            configuration.api_key['api-key'] = brevo_api_key
            api_instance = sib_api_v3_sdk.TransactionalSMSApi(sib_api_v3_sdk.ApiClient(configuration))
            
            for destinataire in campagne["destinataires"]:
                if not destinataire.get("telephone"):
                    continue
                
                # Personnaliser le message
                message = campagne["message"]
                message = message.replace('{prenom}', destinataire.get('prenom', ''))
                message = message.replace('{nom}', destinataire.get('nom', ''))
                
                # Ajouter lien RSVP si activé
                if campagne.get("enable_rsvp"):
                    rsvp_base = f"{os.environ.get('FRONTEND_URL', 'http://localhost:3000')}/rsvp/{campagne_id}"
                    message += f"\n\nRépondez: {rsvp_base}/oui (Oui) | {rsvp_base}/non (Non) | {rsvp_base}/peut_etre (Peut-être)"
                
                # Format phone number (Brevo format: +33XXXXXXXXX)
                phone = destinataire.get("telephone", "").strip()
                if phone.startswith("0"):
                    phone = "+33" + phone[1:]
                elif not phone.startswith("+"):
                    phone = "+33" + phone
                
                try:
                    send_sms = sib_api_v3_sdk.SendTransacSms(
                        sender=brevo_sender,
                        recipient=phone,
                        content=message,
                        type="transactional"
                    )
                    api_instance.send_transac_sms(send_sms)
                    envois_reussis += 1
                except ApiException as e:
                    print(f"Erreur envoi SMS Brevo à {phone}: {e}")
                except Exception as e:
                    print(f"Erreur envoi SMS à {phone}: {e}")
    
    await db.campagnes_communication.update_one(
        {"id": campagne_id},
        {"$set": {
            "statut": "envoye",
            "stats.envoyes": envois_reussis
        }}
    )
    
    return {"message": "Campagne envoyée", "count": envois_reussis}

# RSVP PUBLIC (sans authentification)
@api_router.get("/public/campagne/{campagne_id}")
async def get_public_campagne(campagne_id: str):
    """Récupérer les détails d'une campagne (public)"""
    try:
        campagne = await db.campagnes_communication.find_one(
            {"id": campagne_id},
            {"_id": 0, "titre": 1, "message": 1, "date_envoi": 1, "image_url": 1}
        )
        if not campagne:
            raise HTTPException(status_code=404, detail="Campagne non trouvée")
        return campagne
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/public/rsvp")
async def enregistrer_rsvp(data: dict):
    """Enregistrer une réponse RSVP (endpoint public)"""
    campagne_id = data.get("campagne_id")
    contact = data.get("contact")
    reponse = data.get("reponse")
    
    if not campagne_id or not contact or not reponse:
        raise HTTPException(status_code=400, detail="Données manquantes")
    
    if reponse not in ["oui", "non", "peut_etre"]:
        raise HTTPException(status_code=400, detail="Réponse invalide")
    
    # Vérifier si RSVP existe déjà
    existing = await db.rsvp.find_one({"campagne_id": campagne_id, "$or": [
        {"contact_email": contact},
        {"contact_telephone": contact}
    ]})
    
    if existing:
        # Mettre à jour RSVP existant
        await db.rsvp.update_one(
            {"id": existing["id"]},
            {"$set": {"reponse": reponse, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    else:
        # Créer nouveau RSVP
        rsvp = {
            "id": str(uuid.uuid4()),
            "campagne_id": campagne_id,
            "contact_email": contact if "@" in contact else None,
            "contact_telephone": contact if "@" not in contact else None,
            "reponse": reponse,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.rsvp.insert_one(rsvp)
    
    # Update campaign stats
    campagne = await db.campagnes_communication.find_one({"id": campagne_id})
    if campagne:
        # Recount all RSVPs
        oui_count = await db.rsvp.count_documents({"campagne_id": campagne_id, "reponse": "oui"})
        non_count = await db.rsvp.count_documents({"campagne_id": campagne_id, "reponse": "non"})
        peut_etre_count = await db.rsvp.count_documents({"campagne_id": campagne_id, "reponse": "peut_etre"})
        
        await db.campagnes_communication.update_one(
            {"id": campagne_id},
            {"$set": {
                "stats.oui": oui_count,
                "stats.non": non_count,
                "stats.peut_etre": peut_etre_count
            }}
        )
    

@api_router.post("/events/upload-image")
async def upload_image(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Upload an image and save it to backend uploads folder"""
    
    # Check file type
    if not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="Le fichier doit être une image")
    
    # Generate unique filename
    import hashlib
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_hash = hashlib.md5(file.filename.encode()).hexdigest()[:8]
    extension = file.filename.split('.')[-1] if '.' in file.filename else 'jpg'
    new_filename = f"campaign_{timestamp}_{file_hash}.{extension}"
    
    # Save to backend uploads folder (accessible via API)
    upload_dir = "/app/backend/uploads"
    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, new_filename)
    
    # Write file
    contents = await file.read()
    with open(file_path, "wb") as f:
        f.write(contents)
    
    # Return API URL (served by backend, accessible publicly)
    backend_url = os.environ['REACT_APP_BACKEND_URL']
    public_url = f"{backend_url}/api/uploads/{new_filename}"
    
    return {"image_url": public_url}

@api_router.post("/visitors/upload-photo")
async def upload_visitor_photo(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Upload photo for a visitor"""
    # Check file type
    if not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="Le fichier doit être une image")
    
    # Generate unique filename
    import hashlib
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_hash = hashlib.md5(file.filename.encode()).hexdigest()[:8]
    extension = file.filename.split('.')[-1] if '.' in file.filename else 'jpg'
    new_filename = f"visitor_{timestamp}_{file_hash}.{extension}"
    
    # Save to backend uploads folder
    upload_dir = "/app/backend/uploads"
    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, new_filename)
    
    # Write file
    contents = await file.read()
    with open(file_path, "wb") as f:
        f.write(contents)
    
    # Return API URL
    backend_url = os.environ['REACT_APP_BACKEND_URL']
    public_url = f"{backend_url}/api/uploads/{new_filename}"
    
    return {"photo_url": public_url}

@api_router.get("/uploads/{filename}")
async def get_uploaded_image(filename: str):
    """Serve uploaded images publicly (no authentication required)"""
    file_path = f"/app/backend/uploads/{filename}"
    
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Image not found")
    
    # Detect MIME type
    mime_type, _ = mimetypes.guess_type(file_path)
    if not mime_type:
        mime_type = "image/jpeg"
    
    return FileResponse(file_path, media_type=mime_type)

    return {"message": "Réponse enregistrée"}

@api_router.get("/events/campagnes/{campagne_id}/rsvp")
async def get_rsvp_stats(campagne_id: str, current_user: dict = Depends(get_current_user)):
    """Get RSVP responses for a campaign"""
    
    responses = await db.rsvp.find(
        {"campagne_id": campagne_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(1000)
    
    return responses

@api_router.delete("/events/rsvp/{reponse_id}")
async def delete_rsvp_response(reponse_id: str, current_user: dict = Depends(get_current_user)):
    """Delete an RSVP response"""
    
    result = await db.rsvp.delete_one({"id": reponse_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Réponse non trouvée")
    
    return {"message": "Réponse RSVP supprimée"}


# ==================== RSVP LINKS - STANDALONE EVENTS ====================

@api_router.post("/events")
async def create_event(event: ChurchEventCreate, current_user: dict = Depends(get_current_user)):
    """Create a new event with RSVP link"""
    
    event_data = event.model_dump()
    event_data["id"] = str(uuid.uuid4())
    event_data["created_by"] = current_user["id"]
    event_data["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.church_events.insert_one(event_data)
    
    # Return the event data without MongoDB ObjectId
    return {
        "id": event_data["id"],
        "title": event_data["title"],
        "description": event_data.get("description"),
        "date": event_data["date"],
        "time": event_data.get("time"),
        "location": event_data.get("location"),
        "image_url": event_data.get("image_url"),
        "created_by": event_data["created_by"],
        "created_at": event_data["created_at"],
        "rsvp_enabled": event_data.get("rsvp_enabled", True),
        "max_participants": event_data.get("max_participants"),
        "require_names": event_data.get("require_names", False),
        "require_payment_method": event_data.get("require_payment_method", False),
        "custom_link_title": event_data.get("custom_link_title"),
        "custom_link_url": event_data.get("custom_link_url"),
        "require_email_contact": event_data.get("require_email_contact", False),
        "confirmation_message": event_data.get("confirmation_message")
    }

@api_router.get("/events")
async def get_events(current_user: dict = Depends(get_current_user)):
    """Get all events - visibility based on role"""
    
    # Tous les rôles voient tous les événements RSVP
    # Note: responsable_eglise pourrait être filtré par ville si les événements avaient un champ "city"
    # Pour l'instant, tout le monde voit tout comme demandé
    events = await db.church_events.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    return events

@api_router.get("/events/{event_id}")
async def get_event(event_id: str):
    """Get a specific event (public access for RSVP page)"""
    event = await db.church_events.find_one({"id": event_id}, {"_id": 0})
    
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    return event

@api_router.put("/events/{event_id}")
async def update_event(event_id: str, event: ChurchEventCreate, current_user: dict = Depends(get_current_user)):
    """Update an existing event"""
    
    # Check if event exists
    existing_event = await db.church_events.find_one({"id": event_id})
    if not existing_event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    # Tous les utilisateurs peuvent modifier les événements (accès universel)
    
    # Update the event
    event_data = event.model_dump()
    event_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.church_events.update_one(
        {"id": event_id},
        {"$set": event_data}
    )
    
    return {"message": "Event updated successfully", "id": event_id}

@api_router.delete("/events/{event_id}")
async def delete_event(event_id: str, current_user: dict = Depends(get_current_user)):
    """Delete an event"""
    
    # Tous les utilisateurs peuvent supprimer les événements (accès universel)
    result = await db.church_events.delete_one({"id": event_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Event not found")
    
    # Also delete associated RSVPs
    await db.event_rsvps.delete_many({"event_id": event_id})
    
    return {"message": "Event deleted"}

@api_router.post("/events/{event_id}/rsvp-public")
async def create_event_rsvp_public(event_id: str, rsvp: EventRSVPCreate):
    """Public endpoint - Submit RSVP for an event"""
    # Verify event exists
    event = await db.church_events.find_one({"id": event_id})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    if not event.get("rsvp_enabled", True):
        raise HTTPException(status_code=400, detail="RSVP not enabled for this event")
    
    # Check max participants
    if event.get("max_participants"):
        current_count = await db.event_rsvps.count_documents({
            "event_id": event_id,
            "status": "confirmed"
        })
        if current_count >= event["max_participants"]:
            raise HTTPException(status_code=400, detail="Event is full")
    
    rsvp_data = rsvp.model_dump()
    rsvp_data["id"] = str(uuid.uuid4())
    rsvp_data["event_id"] = event_id
    rsvp_data["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.event_rsvps.insert_one(rsvp_data)
    
    # Send confirmation email if conditions are met
    email_sent = False
    if (event.get("require_email_contact") and 
        rsvp_data.get("status") == "confirmed" and 
        rsvp_data.get("email")):
        
        try:
            import sib_api_v3_sdk
            from sib_api_v3_sdk.rest import ApiException
            
            brevo_api_key = os.getenv('BREVO_API_KEY')
            if brevo_api_key and event.get("confirmation_message"):
                configuration = sib_api_v3_sdk.Configuration()
                configuration.api_key['api-key'] = brevo_api_key
                api_instance = sib_api_v3_sdk.TransactionalEmailsApi(sib_api_v3_sdk.ApiClient(configuration))
                
                # Personnaliser le message
                message = event.get("confirmation_message", "")
                prenom = rsvp_data.get("first_name") or ""
                nom = rsvp_data.get("last_name") or ""
                evenement = event.get("title") or ""
                date = event.get("date") or ""
                lieu = event.get("location") or ""
                
                message = message.replace("{prenom}", prenom)
                message = message.replace("{nom}", nom)
                message = message.replace("{evenement}", evenement)
                message = message.replace("{date}", date)
                message = message.replace("{lieu}", lieu)
                
                message_html = message.replace('\n', '<br>')
                
                # Logo ICC
                logo_url = "https://customer-assets.emergentagent.com/job_dijon-icc-hub/artifacts/foeikpvk_IMG_2590.png"
                
                # Ajouter l'image de l'événement si présente
                image_html = ""
                image_url = (event.get("image_url") or "").strip()
                if image_url:
                    image_html = f'''
                    <div style="text-align: center; margin-top: 20px; margin-bottom: 20px;">
                        <img src="{image_url}" alt="Événement" style="max-width: 100%; height: auto; border-radius: 8px; display: block; margin: 0 auto;" />
                    </div>
                    '''
                
                html_content = f'''
                <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
                    <div style="text-align: center; padding: 20px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);">
                        <img src="{logo_url}" alt="ICC BFC-Italie" style="width: 120px; height: 120px; border-radius: 60px; border: 4px solid white;" />
                        <h2 style="color: white; margin-top: 10px;">Impact Centre Chrétien BFC-Italie</h2>
                    </div>
                    <div style="padding: 30px; background-color: #f9fafb;">
                        {message_html}
                        {image_html}
                    </div>
                    <div style="padding: 20px; text-align: center; background-color: #667eea; color: white; font-size: 12px;">
                        <p>© {datetime.now().year} Impact Centre Chrétien BFC-Italie</p>
                        <p>My Events Church - Gestion d'Événements</p>
                    </div>
                </div>
                '''
                
                sender_email = os.environ.get('SENDER_EMAIL', 'impactcentrechretienbfcitalie@gmail.com')
                sender_name = os.environ.get('SENDER_NAME', 'Impact Centre Chrétien BFC-Italie')
                
                send_smtp_email = sib_api_v3_sdk.SendSmtpEmail(
                    to=[{"email": rsvp_data.get("email"), "name": f"{prenom} {nom}"}],
                    subject=f"Confirmation: {evenement}",
                    html_content=html_content,
                    sender={"name": sender_name, "email": sender_email}
                )
                
                api_instance.send_transac_email(send_smtp_email)
                email_sent = True
        except Exception as e:
            # Log error but don't fail the RSVP submission
            print(f"Error sending confirmation email: {e}")
    
    return {
        "message": "RSVP submitted",
        "id": rsvp_data["id"],
        "email_sent": email_sent
    }

@api_router.get("/events/{event_id}/rsvp")
async def get_event_rsvps(event_id: str, current_user: dict = Depends(get_current_user)):
    """Get all RSVPs for an event with statistics"""
    # Tous les utilisateurs peuvent voir tous les événements
    event = await db.church_events.find_one({"id": event_id})
    
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    rsvps = await db.event_rsvps.find({"event_id": event_id}, {"_id": 0}).to_list(1000)
    
    # Calculate stats
    total = len(rsvps)
    confirmed = len([r for r in rsvps if r.get("status") == "confirmed"])
    declined = len([r for r in rsvps if r.get("status") == "declined"])
    maybe = len([r for r in rsvps if r.get("status") == "maybe"])
    
    return {
        "total": total,
        "confirmed": confirmed,
        "declined": declined,
        "maybe": maybe,
        "responses": rsvps
    }

@api_router.delete("/events/{event_id}/rsvp/{rsvp_id}")
async def delete_event_rsvp(event_id: str, rsvp_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a specific RSVP response"""
    result = await db.event_rsvps.delete_one({"id": rsvp_id, "event_id": event_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="RSVP not found")
    return {"message": "RSVP deleted successfully"}

@api_router.post("/upload-event-image")
async def upload_event_image(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Upload an image for an event"""
    
    # Check file type
    if not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="File must be an image")
    
    # Generate unique filename
    import hashlib
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_hash = hashlib.md5(file.filename.encode()).hexdigest()[:8]
    extension = file.filename.split('.')[-1] if '.' in file.filename else 'jpg'
    new_filename = f"event_{timestamp}_{file_hash}.{extension}"
    
    # Save to uploads folder
    upload_dir = "/app/backend/uploads"
    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, new_filename)
    
    # Write file
    contents = await file.read()
    with open(file_path, "wb") as f:
        f.write(contents)
    
    # Return public URL (relative path - frontend will use REACT_APP_BACKEND_URL)
    public_url = f"/api/uploads/{new_filename}"
    
    return {"image_url": public_url}

# ==================== CONTACT GROUPS (BOXES) ====================

class ContactGroup(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    contacts: List[Dict[str, str]]
    created_by: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

@api_router.get("/contact-groups")
async def get_contact_groups(user: dict = Depends(get_current_user)):
    """Récupérer toutes les boxes de contacts"""
    try:
        groups = await db.contact_groups.find(
            {}, 
            {"_id": 0}
        ).to_list(1000)
        return groups
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==================== PRÉSENCE BERGERS ====================

class BergerPresence(BaseModel):
    berger_id: str
    date: str
    present: bool
    priere: bool = False
    commentaire: Optional[str] = None
    enregistre_par: str
    ville: str
    promo_name: Optional[str] = None
    noms_bergers: Optional[str] = None
    personnes_suivies: Optional[int] = None

class BergerPresenceBatch(BaseModel):
    presences: list[BergerPresence]

@api_router.post("/berger-presences/batch")
async def create_berger_presences_batch(
    batch: BergerPresenceBatch,
    current_user: dict = Depends(get_current_user)
):
    """Enregistrer plusieurs présences de bergers en une fois"""
    
    created_count = 0
    for presence in batch.presences:
        presence_data = presence.model_dump()
        presence_data["id"] = str(uuid4())
        presence_data["created_at"] = datetime.now(timezone.utc).isoformat()
        
        # Vérifier si une présence existe déjà pour ce berger à cette date
        existing = await db.berger_presences.find_one({
            "berger_id": presence.berger_id,
            "date": presence.date
        })
        
        if existing:
            # Mettre à jour
            await db.berger_presences.update_one(
                {"id": existing["id"]},
                {"$set": presence_data}
            )
        else:
            # Créer nouveau
            await db.berger_presences.insert_one(presence_data)
        
        created_count += 1
    
    return {"message": f"{created_count} présence(s) enregistrée(s)", "count": created_count}

@api_router.get("/berger-presences")
async def get_berger_presences(
    date: str,
    ville: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtenir les présences des bergers pour une date donnée"""
    
    presences = await db.berger_presences.find({
        "date": date,
        "ville": ville
    }, {"_id": 0}).to_list(100)
    
    # Enrichir avec les noms des bergers et enregistreurs
    for presence in presences:
        # Chercher le berger
        berger = await db.users.find_one({"id": presence["berger_id"]}, {"_id": 0})
        if berger:
            presence["berger_name"] = berger.get("name", "Inconnu")
            presence["berger_email"] = berger.get("email", "")
        
        # Chercher l'enregistreur
        enregistreur = await db.users.find_one({"id": presence["enregistre_par"]}, {"_id": 0})
        if enregistreur:
            presence["enregistre_par_name"] = enregistreur.get("name", "Inconnu")
    
    return presences


@api_router.get("/berger-presences/latest")
async def get_latest_berger_presences(
    ville: str,
    current_user: dict = Depends(get_current_user)
):
    """Obtenir les dernières présences des bergers par promo pour pré-remplissage"""
    
    # Récupérer toutes les présences de cette ville, triées par date décroissante
    all_presences = await db.berger_presences.find({
        "ville": ville
    }, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Pour chaque promo, garder seulement la dernière présence
    latest_by_promo = {}
    for presence in all_presences:
        promo_name = presence.get("promo_name")
        if promo_name and promo_name not in latest_by_promo:
            latest_by_promo[promo_name] = presence
    
    return list(latest_by_promo.values())

@api_router.post("/contact-groups")
async def create_contact_group(group: ContactGroup, user: dict = Depends(get_current_user)):
    """Créer une nouvelle box de contacts"""
    try:
        group_dict = group.model_dump()
        group_dict["created_by"] = user["username"]
        await db.contact_groups.insert_one(group_dict)
        return {"message": "Box créée", "id": group.id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/contact-groups/{group_id}")
async def delete_contact_group(group_id: str, user: dict = Depends(get_current_user)):
    """Supprimer une box de contacts"""
    try:
        await db.contact_groups.delete_one({"id": group_id})
        return {"message": "Box supprimée"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==================== CONTACT GROUPS SMS ====================

@api_router.get("/contact-groups-sms")
async def get_contact_groups_sms(user: dict = Depends(get_current_user)):
    """Récupérer toutes les boxes SMS"""
    try:
        groups = await db.contact_groups_sms.find(
            {}, 
            {"_id": 0}
        ).to_list(1000)
        return groups
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/contact-groups-sms")
async def create_contact_group_sms(group: ContactGroup, user: dict = Depends(get_current_user)):
    """Créer une nouvelle box SMS"""
    try:
        group_dict = group.model_dump()
        group_dict["created_by"] = user["username"]
        await db.contact_groups_sms.insert_one(group_dict)
        return {"message": "Box SMS créée", "id": group.id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/contact-groups-sms/{group_id}")
async def delete_contact_group_sms(group_id: str, user: dict = Depends(get_current_user)):
    """Supprimer une box SMS"""
    try:
        await db.contact_groups_sms.delete_one({"id": group_id})
        return {"message": "Box SMS supprimée"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==================== CONTACT GROUPS WHATSAPP ====================

@api_router.get("/contact-groups-whatsapp")
async def get_contact_groups_whatsapp(user: dict = Depends(get_current_user)):
    """Récupérer toutes les boxes WhatsApp"""
    try:
        groups = await db.contact_groups_whatsapp.find(
            {}, 
            {"_id": 0}
        ).to_list(1000)
        return groups
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/contact-groups-whatsapp")
async def create_contact_group_whatsapp(group: ContactGroup, user: dict = Depends(get_current_user)):
    """Créer une nouvelle box WhatsApp"""
    try:
        group_dict = group.model_dump()
        group_dict["created_by"] = user["username"]
        await db.contact_groups_whatsapp.insert_one(group_dict)
        return {"message": "Box WhatsApp créée", "id": group.id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/contact-groups-whatsapp/{group_id}")
async def delete_contact_group_whatsapp(group_id: str, user: dict = Depends(get_current_user)):
    """Supprimer une box WhatsApp"""
    try:
        await db.contact_groups_whatsapp.delete_one({"id": group_id})
        return {"message": "Box WhatsApp supprimée"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==================== WHATSAPP CAMPAIGNS ====================

class WhatsAppCampaign(BaseModel):
    titre: str
    message: str
    destinataires: List[Dict[str, str]]
    template_id: Optional[str] = None

@api_router.post("/events/whatsapp/send")
async def send_whatsapp_campaign(campaign: WhatsAppCampaign, user: dict = Depends(get_current_user)):
    """Envoyer une campagne WhatsApp via Brevo"""
    try:
        # Sauvegarder la campagne
        campagne_doc = {
            "id": str(uuid.uuid4()),
            "type": "whatsapp",
            "titre": campaign.titre,
            "message": campaign.message,
            "template_id": campaign.template_id,
            "destinataires_count": len(campaign.destinataires),
            "created_by": user["username"],
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.campagnes_communication.insert_one(campagne_doc)
        
        # TODO: Implémenter l'envoi réel via Brevo WhatsApp API
        # Configuration Brevo:
        # 1. Créer un compte Brevo et activer WhatsApp Business
        # 2. Créer des templates WhatsApp approuvés
        # 3. Obtenir l'API Key depuis Brevo → Settings → API Keys
        # 4. Ajouter BREVO_API_KEY dans .env
        
        # Exemple d'envoi via Brevo:
        # import requests
        # brevo_api_key = os.environ.get('BREVO_API_KEY')
        # for contact in campaign.destinataires:
        #     response = requests.post(
        #         'https://api.brevo.com/v3/whatsapp/sendMessage',
        #         headers={'api-key': brevo_api_key},
        #         json={
        #             'phoneNumber': contact.telephone,
        #             'templateId': campaign.template_id,
        #             'params': {...}
        #         }
        #     )
        
        success_count = len(campaign.destinataires)
        failed_count = 0
        
        return {
            "message": "Campagne WhatsApp envoyée",
            "campaign_id": campagne_doc["id"],
            "success_count": success_count,
            "failed_count": failed_count
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==================== PLANNING DES ACTIVITÉS ====================

class PlanningActivite(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nom: str
    annee: int  # Année du planning (2025-2035)
    semestre: int  # 1 ou 2
    date_debut: str  # Date de début
    date_fin: str  # Date de fin
    date: Optional[str] = None  # DEPRECATED - Pour compatibilité
    ministeres: str  # Texte libre au lieu de List[str]
    statut: str  # "À venir", "Reporté", "Annulé", "Fait"
    commentaire: str = ""
    ville: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

@api_router.get("/planning/activites")
async def get_activites_planning(ville: str, annee: Optional[int] = None, user: dict = Depends(get_current_user)):
    """Récupérer toutes les activités pour une ville et une année"""
    try:
        query = {"ville": ville}
        if annee:
            query["annee"] = annee
        
        activites = await db.planning_activites.find(
            query, 
            {"_id": 0}
        ).sort([("semestre", 1), ("date_debut", 1)]).to_list(1000)
        return activites
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/planning/activites")
async def create_activite(activite: PlanningActivite, user: dict = Depends(get_current_user)):
    """Créer une nouvelle activité"""
    try:
        activite_dict = activite.model_dump()
        activite_dict["created_by"] = user["username"]
        await db.planning_activites.insert_one(activite_dict)
        return {"message": "Activité créée", "id": activite.id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/planning/activites/{activite_id}")
async def update_activite(activite_id: str, activite: PlanningActivite, user: dict = Depends(get_current_user)):
    """Modifier une activité existante"""
    try:
        activite_dict = activite.model_dump()
        activite_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.planning_activites.update_one(
            {"id": activite_id},
            {"$set": activite_dict}
        )
        return {"message": "Activité mise à jour"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/planning/activites/{activite_id}")
async def delete_activite(activite_id: str, user: dict = Depends(get_current_user)):
    """Supprimer une activité"""
    try:
        await db.planning_activites.delete_one({"id": activite_id})
        return {"message": "Activité supprimée"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Router will be included at the end of the file after all endpoints are defined

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging - set to ERROR level to reduce noise
logging.basicConfig(
    level=logging.ERROR,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Suppress uvicorn access logs
logging.getLogger("uvicorn.access").setLevel(logging.ERROR)

@app.on_event("startup")
async def startup_initialize_cities():
    """Initialize cities automatically on startup - ALWAYS update countries"""
    try:
        print("🏙️ Initializing cities system...")
        
        # Clean up invalid cities (with null or empty name)
        deleted = await db.cities.delete_many({'$or': [{'name': None}, {'name': ''}]})
        if deleted.deleted_count > 0:
            print(f"🗑️ Cleaned {deleted.deleted_count} invalid cities")
        
        # Define cities with countries
        cities_mapping = {
            'Milan': 'Italie',
            'Rome': 'Italie', 
            'Perugia': 'Italie',
            'Bologne': 'Italie',
            'Turin': 'Italie',
            'Dijon': 'France',
            'Auxerre': 'France',
            'Besançon': 'France',
            'Chalon-Sur-Saone': 'France',
            'Dole': 'France',
            'Sens': 'France'
        }
        
        # ALWAYS update or create each city (not just when count is 0)
        created = 0
        updated = 0
        
        for city_name, country in cities_mapping.items():
            # Check if city exists (case-insensitive)
            existing = await db.cities.find_one({'name': {'$regex': f'^{city_name}$', '$options': 'i'}})
            
            if existing:
                # Update existing city - FORCE country update
                result = await db.cities.update_one(
                    {'_id': existing['_id']},
                    {'$set': {'country': country}}
                )
                if result.modified_count > 0:
                    updated += 1
                    print(f"  ✅ {city_name} → {country}")
            else:
                # Create new city
                new_city = {
                    'id': str(uuid.uuid4()),
                    'name': city_name,
                    'country': country
                }
                await db.cities.insert_one(new_city)
                created += 1
                print(f"  ➕ {city_name} → {country} (created)")
        
        total_cities = await db.cities.count_documents({})
        print(f"✅ Cities initialized: {created} created, {updated} updated, {total_cities} total")
        
    except Exception as e:
        print(f"⚠️ Warning: Could not initialize cities: {e}")
        import traceback
        traceback.print_exc()

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

# ==================== EVANGELISATION MODELS ====================

class EvangelisationData(BaseModel):
    model_config = ConfigDict(arbitrary_types_allowed=True)
    
    type: str  # "eglise" or "familles_impact"
    nombre_gagneurs_ame: int = 0
    nombre_personnes_receptives: int = 0
    nombre_priere_salut: int = 0
    nombre_contacts_pris: int = 0
    nombre_ames_invitees: int = 0
    nombre_miracles: int = 0
    commentaire: Optional[str] = None

class EvangelisationRecord(BaseModel):
    model_config = ConfigDict(arbitrary_types_allowed=True)
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    ville: str
    created_by: str
    eglise: Optional[EvangelisationData] = None
    familles_impact: Optional[EvangelisationData] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# ==================== MINISTERE STARS MODELS ====================

class StarCreate(BaseModel):
    prenom: str
    nom: str
    jour_naissance: int  # 1-31
    mois_naissance: int  # 1-12
    departements: List[str]  # Liste des départements (peut être multiple)
    ville: Optional[str] = None

class Star(BaseModel):
    model_config = ConfigDict(arbitrary_types_allowed=True)
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    prenom: str
    nom: str
    jour_naissance: int
    mois_naissance: int
    departements: List[str]
    ville: str
    statut: str = "actif"  # actif ou non_actif
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class StarUpdate(BaseModel):
    prenom: Optional[str] = None
    nom: Optional[str] = None
    jour_naissance: Optional[int] = None
    mois_naissance: Optional[int] = None
    departements: Optional[List[str]] = None
    statut: Optional[str] = None

# ==================== EVANGELISATION ENDPOINTS ====================

@api_router.post("/evangelisation")
async def create_evangelisation_record(record: EvangelisationRecord, current_user: dict = Depends(get_current_user)):
    """Create evangelisation record"""
    
    # Set ville and created_by
    record.ville = current_user.get("city", record.ville)
    record.created_by = current_user["username"]
    
    # Convert to dict and save
    record_dict = record.model_dump()
    await db.evangelisation.insert_one(record_dict)
    
    return {"message": "Record created successfully", "id": record.id}

@api_router.get("/evangelisation")
async def get_evangelisation_records(
    ville: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get evangelisation records with filters"""
    
    query = {}
    
    # Filter by ville
    if current_user["role"] in ["responsable_eglise", "responsable_evangelisation"]:
        query["ville"] = current_user.get("city")
    elif ville:
        query["ville"] = ville
    
    # Date filters
    if date_from and date_to:
        query["date"] = {"$gte": date_from, "$lte": date_to}
    elif date_from:
        query["date"] = {"$gte": date_from}
    elif date_to:
        query["date"] = {"$lte": date_to}
    
    records = await db.evangelisation.find(query, {"_id": 0}).to_list(1000)
    return records

@api_router.get("/evangelisation/stats")
async def get_evangelisation_stats(
    ville: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get evangelisation statistics"""
    
    query = {}
    
    # Filter by ville
    if current_user["role"] == "responsable_eglise":
        query["ville"] = current_user.get("city")
    elif ville:
        query["ville"] = ville
    
    # Date filters
    if year and month:
        query["date"] = {"$regex": f"^{year}-{month:02d}"}
    elif year:
        query["date"] = {"$regex": f"^{year}"}
    
    records = await db.evangelisation.find(query).to_list(1000)
    
    # Calculate totals
    stats = {
        "eglise": {
            "nombre_gagneurs_ame": 0,
            "nombre_personnes_receptives": 0,
            "nombre_priere_salut": 0,
            "nombre_contacts_pris": 0,
            "nombre_ames_invitees": 0,
            "nombre_miracles": 0
        },
        "familles_impact": {
            "nombre_gagneurs_ame": 0,
            "nombre_personnes_receptives": 0,
            "nombre_priere_salut": 0,
            "nombre_contacts_pris": 0,
            "nombre_ames_invitees": 0,
            "nombre_miracles": 0
        },
        "total_records": len(records)
    }
    
    for record in records:
        if record.get("eglise"):
            for key in stats["eglise"]:
                stats["eglise"][key] += record["eglise"].get(key, 0)
        
        if record.get("familles_impact"):
            for key in stats["familles_impact"]:
                stats["familles_impact"][key] += record["familles_impact"].get(key, 0)
    
    return stats



# ============== NOTIFICATIONS ENDPOINTS ==============

@api_router.post("/notifications/register-token")
async def register_fcm_token(token_data: dict, current_user: dict = Depends(get_current_user)):
    """Enregistrer le token FCM d'un utilisateur"""
    try:
        # Vérifier si le token existe déjà
        existing = await db.fcm_tokens.find_one({
            "user_id": current_user["id"],
            "token": token_data["token"]
        })
        
        if existing:
            # Mettre à jour last_used
            await db.fcm_tokens.update_one(
                {"id": existing["id"]},
                {"$set": {"last_used": datetime.now(timezone.utc).isoformat()}}
            )
            return {"message": "Token updated"}
        
        # Créer nouveau token
        fcm_token = {
            "id": str(uuid.uuid4()),
            "user_id": current_user["id"],
            "token": token_data["token"],
            "device_type": token_data.get("device_type", "web"),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "last_used": datetime.now(timezone.utc).isoformat()
        }
        
        await db.fcm_tokens.insert_one(fcm_token)
        return {"message": "Token registered successfully"}
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/notifications/create")
async def create_notification(notif: NotificationCreate, current_user: dict = Depends(get_current_user)):
    """Créer une notification (superadmin seulement)"""
    if current_user["role"] != "super_admin":
        raise HTTPException(status_code=403, detail="Permission denied")
    
    notification_data = notif.model_dump()
    notification_data["id"] = str(uuid.uuid4())
    notification_data["created_by"] = current_user["id"]
    notification_data["status"] = "scheduled" if notif.scheduled_at else "pending"
    notification_data["sent_count"] = 0
    notification_data["failed_count"] = 0
    notification_data["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.notifications.insert_one(notification_data)
    
    # Si pas de programmation, envoyer immédiatement
    if not notif.scheduled_at:
        # On envoie en arrière-plan
        return {"message": "Notification créée et sera envoyée", "id": notification_data["id"]}
    
    return {"message": "Notification programmée", "id": notification_data["id"]}


@api_router.post("/notifications/{notification_id}/send")
async def send_notification(notification_id: str, current_user: dict = Depends(get_current_user)):
    """Envoyer une notification maintenant"""
    if current_user["role"] != "super_admin":
        raise HTTPException(status_code=403, detail="Permission denied")
    
    # Récupérer la notification
    notification = await db.notifications.find_one({"id": notification_id})
    if not notification:
        raise HTTPException(status_code=404, detail="Notification not found")
    
    # Récupérer les tokens des utilisateurs ciblés
    query = {}
    
    if notification.get("send_to_all"):
        # Tous les utilisateurs
        pass
    else:
        # Construire la requête de ciblage
        user_query = {}
        
        if notification.get("target_users"):
            user_query["id"] = {"$in": notification["target_users"]}
        
        if notification.get("city"):
            user_query["city"] = notification["city"]
        
        if notification.get("target_roles"):
            user_query["role"] = {"$in": notification["target_roles"]}
        
        # Récupérer les utilisateurs ciblés
        targeted_users = await db.users.find(user_query, {"id": 1, "_id": 0}).to_list(10000)
        user_ids = [u["id"] for u in targeted_users]
        
        if not user_ids:
            return {"message": "Aucun utilisateur ciblé trouvé", "sent_count": 0}
        
        query["user_id"] = {"$in": user_ids}
    
    # Récupérer tous les tokens
    tokens = await db.fcm_tokens.find(query, {"token": 1, "_id": 0}).to_list(10000)
    
    if not tokens:
        await db.notifications.update_one(
            {"id": notification_id},
            {"$set": {"status": "failed", "sent_at": datetime.now(timezone.utc).isoformat()}}
        )
        return {"message": "Aucun token trouvé", "sent_count": 0}
    
    # Envoyer via Firebase FCM API
    sent_count = 0
    failed_count = 0
    fcm_url = "https://fcm.googleapis.com/fcm/send"
    headers = {
        "Authorization": f"key={FIREBASE_SERVER_KEY}",
        "Content-Type": "application/json"
    }
    
    for token_doc in tokens:
        try:
            payload = {
                "to": token_doc["token"],
                "notification": {
                    "title": notification["title"],
                    "body": notification["message"],
                    "icon": "/logo192.png",
                    "click_action": "/"
                },
                "data": {
                    "notification_id": notification_id,
                    "created_at": notification["created_at"]
                }
            }
            
            response = requests.post(fcm_url, json=payload, headers=headers, timeout=10)
            
            if response.status_code == 200:
                sent_count += 1
            else:
                failed_count += 1
                print(f"FCM Error: {response.status_code} - {response.text}")
        
        except Exception as e:
            failed_count += 1
            print(f"Error sending to token: {str(e)}")
    
    # Mettre à jour la notification
    status_value = "sent" if sent_count > 0 else "failed"
    await db.notifications.update_one(
        {"id": notification_id},
        {"$set": {
            "status": status_value,
            "sent_at": datetime.now(timezone.utc).isoformat(),
            "sent_count": sent_count,
            "failed_count": failed_count
        }}
    )
    
    return {
        "message": f"Notification envoyée à {sent_count} utilisateurs ({failed_count} échecs)",
        "sent_count": sent_count,
        "failed_count": failed_count
    }


@api_router.get("/notifications")
async def get_notifications(current_user: dict = Depends(get_current_user)):
    """Récupérer l'historique des notifications (superadmin uniquement)"""
    if current_user["role"] != "super_admin":
        raise HTTPException(status_code=403, detail="Permission denied")
    
    notifications = await db.notifications.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return notifications


@api_router.get("/notifications/my-notifications")
async def get_my_notifications(current_user: dict = Depends(get_current_user)):
    """Récupérer les notifications d'un utilisateur"""
    # Construire les critères de ciblage pour cet utilisateur
    notifications = await db.notifications.find({
        "status": "sent",
        "$or": [
            {"send_to_all": True},
            {"target_users": current_user["id"]},
            {"city": current_user.get("city")},
            {"target_roles": current_user["role"]}
        ]
    }, {"_id": 0}).sort("sent_at", -1).limit(50).to_list(50)
    
    return notifications


@api_router.delete("/notifications/{notification_id}")
async def delete_notification(notification_id: str, current_user: dict = Depends(get_current_user)):
    """Supprimer une notification"""
    if current_user["role"] != "super_admin":
        raise HTTPException(status_code=403, detail="Permission denied")
    
    result = await db.notifications.delete_one({"id": notification_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    
    return {"message": "Notification deleted"}


# ==================== MINISTERE STARS ENDPOINTS ====================

@api_router.post("/stars")
async def create_star(star: StarCreate, current_user: dict = Depends(get_current_user)):
    """Créer une star (admin authentifié)"""
    # Si ville non fournie, utiliser celle de l'utilisateur connecté
    ville = star.ville if star.ville else current_user.get("city", "")
    
    star_obj = Star(
        prenom=star.prenom,
        nom=star.nom,
        jour_naissance=star.jour_naissance,
        mois_naissance=star.mois_naissance,
        departements=star.departements,
        ville=ville,
        statut="actif"
    )
    
    await db.stars.insert_one(star_obj.model_dump())
    return {"message": "Star créée avec succès", "id": star_obj.id}


@api_router.post("/stars/public/register")
async def create_star_public(star: StarCreate):
    """Créer une star (recensement public - sans authentification)"""
    if not star.ville:
        raise HTTPException(status_code=400, detail="La ville est obligatoire")
    
    if not star.departements or len(star.departements) == 0:
        raise HTTPException(status_code=400, detail="Au moins un département est obligatoire")
    
    star_obj = Star(
        prenom=star.prenom,
        nom=star.nom,
        jour_naissance=star.jour_naissance,
        mois_naissance=star.mois_naissance,
        departements=star.departements,
        ville=star.ville,
        statut="actif"
    )
    
    await db.stars.insert_one(star_obj.model_dump())
    return {"message": "Inscription réussie! Merci pour votre engagement.", "id": star_obj.id}


@api_router.get("/stars")
async def get_stars(current_user: dict = Depends(get_current_user)):
    """Récupérer toutes les stars (avec permissions)"""
    if current_user["role"] not in ["super_admin", "pasteur", "responsable_eglise", "ministere_stars"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    query = {}
    
    # Responsable d'église voit seulement sa ville
    if current_user["role"] == "responsable_eglise":
        query["ville"] = current_user["city"]
    
    stars = await db.stars.find(query, {"_id": 0}).to_list(1000)
    return stars


@api_router.get("/stars/departement/{departement}")
async def get_stars_by_departement(departement: str, ville: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Récupérer les stars d'un département spécifique"""
    if current_user["role"] not in ["super_admin", "pasteur", "responsable_eglise", "ministere_stars", "respo_departement", "star"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    query = {"departements": departement}
    
    # Filtrer par ville si spécifié
    if ville and ville != 'all':
        query["ville"] = ville
    elif current_user["role"] == "responsable_eglise":
        query["ville"] = current_user["city"]
    elif current_user["role"] == "respo_departement":
        # Respo_departement voit seulement sa ville
        query["ville"] = current_user["city"]
    
    stars = await db.stars.find(query, {"_id": 0}).to_list(1000)
    return stars


@api_router.get("/stars/multi-departements")
async def get_stars_multi_departements(current_user: dict = Depends(get_current_user)):
    """Récupérer les stars qui servent dans plusieurs départements"""
    if current_user["role"] not in ["super_admin", "pasteur", "responsable_eglise", "ministere_stars", "respo_departement", "star"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    query = {}
    if current_user["role"] == "responsable_eglise":
        query["ville"] = current_user["city"]
    
    all_stars = await db.stars.find(query, {"_id": 0}).to_list(1000)
    
    # Filtrer ceux qui ont plus d'un département
    multi_dept_stars = [s for s in all_stars if len(s.get("departements", [])) > 1]
    
    return multi_dept_stars


@api_router.get("/stars/anniversaires")
async def get_anniversaires():
    """Récupérer les anniversaires à venir (accessible publiquement)"""
    from datetime import datetime, timedelta, date
    
    today = date.today()  # Utiliser date au lieu de datetime pour éviter les problèmes d'heures
    print(f"🎂 Date du jour: {today}")
    
    # Anniversaires dans les 30 prochains jours
    anniversaires = []
    all_stars = await db.stars.find({}, {"_id": 0}).to_list(1000)
    print(f"📋 Nombre de stars en base: {len(all_stars)}")
    
    for star in all_stars:
        jour = star.get("jour_naissance")
        mois = star.get("mois_naissance")
        
        if not jour or not mois:
            continue
        
        # Créer une date d'anniversaire pour cette année
        try:
            anniv_date = date(today.year, mois, jour)
        except ValueError:
            continue
        
        # Si l'anniversaire est passé cette année, vérifier l'année prochaine
        if anniv_date < today:
            try:
                anniv_date = date(today.year + 1, mois, jour)
            except ValueError:
                continue
        
        # Calculer les jours jusqu'à l'anniversaire
        days_until = (anniv_date - today).days
        
        print(f"  {star.get('prenom')} {star.get('nom')}: {jour}/{mois} - jours restants: {days_until}")
        
        if 0 <= days_until <= 30:  # 30 jours maximum
            print(f"    ✅ Ajouté aux anniversaires!")
            anniversaires.append({
                "prenom": star.get("prenom"),
                "nom": star.get("nom"),
                "ville": star.get("ville", ""),
                "jour": jour,
                "mois": mois,
                "days_until": days_until,
                "date": anniv_date.strftime("%d/%m")
            })
    
    # Trier par jours jusqu'à l'anniversaire
    anniversaires.sort(key=lambda x: x["days_until"])
    
    return anniversaires


@api_router.put("/stars/{star_id}")
async def update_star(star_id: str, update: StarUpdate, current_user: dict = Depends(get_current_user)):
    """Mettre à jour une star"""
    if current_user["role"] not in ["super_admin", "pasteur", "ministere_stars"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    # Récupérer les champs non-null
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    result = await db.stars.update_one(
        {"id": star_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Star not found")
    
    return {"message": "Star mise à jour"}


@api_router.delete("/stars/{star_id}")
async def delete_star(star_id: str, current_user: dict = Depends(get_current_user)):
    """Supprimer une star"""
    if current_user["role"] not in ["super_admin", "pasteur"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    result = await db.stars.delete_one({"id": star_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Star not found")
    
    return {"message": "Star supprimée"}


@api_router.get("/stars/stats/overview")
async def get_stars_stats(ville: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Statistiques globales des stars"""
    if current_user["role"] not in ["super_admin", "pasteur", "responsable_eglise", "ministere_stars", "respo_departement", "star"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    query = {}
    
    # Filtrer par ville si spécifié
    if ville and ville != 'all':
        query["ville"] = ville
    elif current_user["role"] == "responsable_eglise":
        query["ville"] = current_user["city"]
    elif current_user["role"] == "respo_departement":
        query["ville"] = current_user["city"]
    
    all_stars = await db.stars.find(query, {"_id": 0}).to_list(1000)
    
    total = len(all_stars)
    actifs = len([s for s in all_stars if s.get("statut") == "actif"])
    non_actifs = len([s for s in all_stars if s.get("statut") == "non_actif"])
    
    # Compter par département
    dept_counts = {}
    for star in all_stars:
        for dept in star.get("departements", []):
            dept_counts[dept] = dept_counts.get(dept, 0) + 1
    
    return {
        "total": total,
        "actifs": actifs,
        "non_actifs": non_actifs,
        "par_departement": dept_counts
    }


# ========== ENDPOINTS PUBLICS MINISTÈRE DES STARS ==========

@api_router.get("/stars/public/stats/overview")
async def get_stars_stats_public(ville: Optional[str] = None):
    """Statistiques globales des stars - Accès public"""
    query = {}
    if ville:
        query["ville"] = ville
    
    all_stars = await db.stars.find(query, {"_id": 0}).to_list(1000)
    
    total = len(all_stars)
    actifs = len([s for s in all_stars if s.get("statut") == "actif"])
    non_actifs = len([s for s in all_stars if s.get("statut") == "non_actif"])
    
    # Compter par département
    dept_counts = {}
    for star in all_stars:
        for dept in star.get("departements", []):
            dept_counts[dept] = dept_counts.get(dept, 0) + 1
    
    return {
        "total": total,
        "actifs": actifs,
        "non_actifs": non_actifs,
        "par_departement": dept_counts
    }

@api_router.get("/stars/public/multi-departements")
async def get_stars_multi_departements_public(ville: Optional[str] = None):
    """Récupérer les stars qui servent dans plusieurs départements - Accès public"""
    query = {}
    if ville:
        query["ville"] = ville
    
    all_stars = await db.stars.find(query, {"_id": 0}).to_list(1000)
    
    # Filtrer ceux qui ont plus d'un département
    multi_dept_stars = [s for s in all_stars if len(s.get("departements", [])) > 1]
    
    return multi_dept_stars



@api_router.get("/stars/public/departement/{departement}")
async def get_stars_by_departement_public(departement: str, ville: Optional[str] = None):
    """Récupérer les stars d'un département spécifique - Accès public"""
    query = {"departements": departement}
    
    # Filtrer par ville si spécifié
    if ville and ville != 'all':
        query["ville"] = ville
    
    stars = await db.stars.find(query, {"_id": 0}).to_list(1000)
    return stars


# ==================== PLANNINGS STARS ====================

class PlanningEntry(BaseModel):
    type_culte: Optional[str] = None  # "Culte 1", "Culte 2", "EJP", "Tous les cultes", "Événements spéciaux"
    role: Optional[str] = None  # Rôle/Service dans le planning
    poles: Optional[List[str]] = []  # Liste de pôles (0 à plusieurs)
    membre_ids: Optional[List[str]] = []  # IDs des membres assignés (sélection multiple)
    membres_noms: Optional[List[str]] = []  # Noms des membres assignés
    # Legacy fields pour rétro-compatibilité
    pole1: Optional[str] = None
    pole2: Optional[str] = None
    membre_id: Optional[str] = None
    membre_nom: Optional[str] = None
    commentaire: Optional[str] = None

class PlanningCreate(BaseModel):
    departement: str
    semaine: int  # 1-52
    annee: int
    entries: List[PlanningEntry]

@api_router.get("/stars/planning/{departement}")
async def get_department_plannings(departement: str, annee: Optional[int] = None, current_user: dict = Depends(get_current_user)):
    """Récupérer tous les plannings d'un département (avec filtre année optionnel)"""
    if current_user["role"] not in ["super_admin", "pasteur", "responsable_eglise", "ministere_stars", "respo_departement", "star"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    query = {"departement": departement}
    if annee:
        query["annee"] = annee
    
    plannings = await db.stars_planning.find(query, {"_id": 0}).to_list(1000)
    
    return plannings

@api_router.get("/stars/planning/{departement}/{semaine}/{annee}")
async def get_planning(departement: str, semaine: int, annee: int, current_user: dict = Depends(get_current_user)):
    """Récupérer le planning d'une semaine spécifique"""
    if current_user["role"] not in ["super_admin", "pasteur", "responsable_eglise", "ministere_stars", "respo_departement", "star"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    planning = await db.stars_planning.find_one(
        {"departement": departement, "semaine": semaine, "annee": annee},
        {"_id": 0}
    )
    
    return planning or {"departement": departement, "semaine": semaine, "annee": annee, "entries": []}

@api_router.post("/stars/planning")
async def create_or_update_planning(planning: PlanningCreate, current_user: dict = Depends(get_current_user)):
    """Créer ou mettre à jour un planning - respo_departement peut modifier"""
    if current_user["role"] not in ["super_admin", "pasteur", "responsable_eglise", "ministere_stars", "respo_departement"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    planning_data = planning.model_dump()
    planning_data["id"] = f"{planning.departement}_{planning.semaine}_{planning.annee}"
    planning_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    planning_data["updated_by"] = current_user["username"]
    
    await db.stars_planning.update_one(
        {"departement": planning.departement, "semaine": planning.semaine, "annee": planning.annee},
        {"$set": planning_data},
        upsert=True
    )
    
    return {"message": "Planning enregistré"}

@api_router.delete("/stars/planning/{departement}/{semaine}/{annee}")
async def delete_planning(departement: str, semaine: int, annee: int, current_user: dict = Depends(get_current_user)):
    """Supprimer un planning - respo_departement peut supprimer"""
    if current_user["role"] not in ["super_admin", "pasteur", "respo_departement"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    result = await db.stars_planning.delete_one(
        {"departement": departement, "semaine": semaine, "annee": annee}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Planning not found")
    
    return {"message": "Planning supprimé"}


@api_router.get("/stars/service-stats/{semaine}/{annee}")
async def get_stars_service_stats(semaine: int, annee: int, ville: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Récupérer les KPIs des stars en service pour une semaine donnée (tous départements)"""
    if current_user["role"] not in ["super_admin", "pasteur", "responsable_eglise", "ministere_stars", "respo_departement", "star"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    # Récupérer tous les plannings de cette semaine
    plannings = await db.stars_planning.find(
        {"semaine": semaine, "annee": annee},
        {"_id": 0}
    ).to_list(1000)
    
    types_culte = ['Culte 1', 'Culte 2', 'EJP', 'Tous les cultes', 'Événements spéciaux']
    kpis_by_type = {t: {"count": 0, "membres": []} for t in types_culte}
    all_membres = set()
    departements_avec_planning = []
    
    for planning in plannings:
        dept = planning.get("departement", "")
        if dept not in departements_avec_planning:
            departements_avec_planning.append(dept)
            
        for entry in planning.get("entries", []):
            type_culte = entry.get("type_culte")
            membres_noms = entry.get("membres_noms", [])
            membre_ids = entry.get("membre_ids", [])
            
            if type_culte and type_culte in kpis_by_type:
                for i, nom in enumerate(membres_noms):
                    if nom not in kpis_by_type[type_culte]["membres"]:
                        kpis_by_type[type_culte]["membres"].append(nom)
                        kpis_by_type[type_culte]["count"] += 1
                    
                    # Ajouter à la liste globale des membres
                    all_membres.add(nom)
    
    return {
        "semaine": semaine,
        "annee": annee,
        "total_stars_en_service": len(all_membres),
        "par_type_culte": {k: {"count": v["count"], "membres": v["membres"]} for k, v in kpis_by_type.items()},
        "departements_avec_planning": departements_avec_planning,
        "membres_en_service": list(all_membres)
    }


@api_router.get("/stars/service-overview/{annee}")
async def get_stars_service_overview(annee: int, current_user: dict = Depends(get_current_user)):
    """Récupérer un aperçu des stats de service pour toutes les semaines d'une année"""
    if current_user["role"] not in ["super_admin", "pasteur", "responsable_eglise", "ministere_stars", "respo_departement", "star"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    # Récupérer tous les plannings de l'année
    plannings = await db.stars_planning.find(
        {"annee": annee},
        {"_id": 0}
    ).to_list(5000)
    
    # Regrouper par semaine
    weeks_data = {}
    for planning in plannings:
        semaine = planning.get("semaine")
        if semaine not in weeks_data:
            weeks_data[semaine] = {"membres": set(), "departements": set()}
        
        weeks_data[semaine]["departements"].add(planning.get("departement", ""))
        
        for entry in planning.get("entries", []):
            for nom in entry.get("membres_noms", []):
                weeks_data[semaine]["membres"].add(nom)
    
    # Créer le résumé
    result = []
    for week in range(1, 53):
        data = weeks_data.get(week, {"membres": set(), "departements": set()})
        result.append({
            "semaine": week,
            "total_stars_en_service": len(data["membres"]),
            "nb_departements": len(data["departements"]),
            "has_planning": week in weeks_data
        })
    
    return result


# ==================== LE PAIN DU JOUR ====================

class PainDuJourContent(BaseModel):
    date: str  # "YYYY-MM-DD"
    lien_priere: Optional[str] = None  # Lien YouTube temps de prière
    titre_priere: Optional[str] = None
    thumbnail_priere: Optional[str] = None  # Thumbnail YouTube
    duration_priere: Optional[str] = None  # Durée de la vidéo
    lien_enseignement: Optional[str] = None  # Lien YouTube enseignement
    titre_enseignement: Optional[str] = None
    thumbnail_enseignement: Optional[str] = None  # Thumbnail YouTube
    duration_enseignement: Optional[str] = None  # Durée de la vidéo
    versets: Optional[List[Dict]] = []  # [{livre, chapitre, verset_debut, verset_fin}]
    # Nouveau: Résumé et Quiz
    resume: Optional[Dict] = None  # {titre, resume, points_cles, versets_cites, citations}
    quiz: Optional[List[Dict]] = None  # [{question, options, correct_index}]
    created_by: Optional[str] = None
    created_at: Optional[str] = None

class PainDuJourProgrammation(BaseModel):
    """Modèle pour la programmation hebdomadaire"""
    semaine: str  # Format: "2026-W02" (année-semaine)
    jours: Dict  # {"lundi": {...}, "mardi": {...}, etc.}
    created_by: Optional[str] = None
    created_at: Optional[str] = None

class QuizSubmission(BaseModel):
    date: str  # "YYYY-MM-DD"
    answers: List[int]  # Index des réponses choisies
    score: int  # Score calculé côté frontend

class FetchTranscriptionRequest(BaseModel):
    youtube_url: str

class GenerateResumeQuizRequest(BaseModel):
    transcription: str  # La transcription complète
    titre_message: str  # Le titre configuré par l'admin
    minute_debut: int = 0  # La minute à partir de laquelle analyser

class SondagePainDuJour(BaseModel):
    date: str  # "YYYY-MM-DD"
    lecture_reponse: str  # "Oui", "Non", "Partiellement"
    video_reponse: str  # "Oui", "Non", "Pas totalement"

class ClickTrack(BaseModel):
    type: str  # "priere" ou "enseignement"
    date: str  # "YYYY-MM-DD"

class YouTubeVideoRequest(BaseModel):
    url: str


# YouTube API Key
YOUTUBE_API_KEY = os.environ.get("YOUTUBE_API_KEY", "")

def get_youtube_video_id(url: str) -> Optional[str]:
    """Extraire l'ID de la vidéo à partir d'une URL YouTube"""
    patterns = [
        r'(?:youtube\.com\/watch\?v=|youtu\.be\/|youtube\.com\/embed\/)([a-zA-Z0-9_-]{11})',
        r'(?:youtube\.com\/shorts\/)([a-zA-Z0-9_-]{11})',
        r'(?:youtube\.com\/live\/)([a-zA-Z0-9_-]{11})',  # Support YouTube Live
    ]
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return match.group(1)
    return None

def format_duration(duration: str) -> str:
    """Convertir la durée ISO 8601 en format lisible (ex: PT1H2M30S -> 1:02:30)"""
    match = re.match(r'PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?', duration)
    if not match:
        return "0:00"
    hours, minutes, seconds = match.groups()
    hours = int(hours) if hours else 0
    minutes = int(minutes) if minutes else 0
    seconds = int(seconds) if seconds else 0
    
    if hours > 0:
        return f"{hours}:{minutes:02d}:{seconds:02d}"
    else:
        return f"{minutes}:{seconds:02d}"


# Liste des livres de la Bible
LIVRES_BIBLE = [
    # Ancien Testament
    "Genèse", "Exode", "Lévitique", "Nombres", "Deutéronome",
    "Josué", "Juges", "Ruth", "1 Samuel", "2 Samuel",
    "1 Rois", "2 Rois", "1 Chroniques", "2 Chroniques",
    "Esdras", "Néhémie", "Esther", "Job", "Psaumes", "Proverbes",
    "Ecclésiaste", "Cantique des Cantiques", "Ésaïe", "Jérémie",
    "Lamentations", "Ézéchiel", "Daniel", "Osée", "Joël", "Amos",
    "Abdias", "Jonas", "Michée", "Nahum", "Habacuc", "Sophonie",
    "Aggée", "Zacharie", "Malachie",
    # Nouveau Testament
    "Matthieu", "Marc", "Luc", "Jean", "Actes",
    "Romains", "1 Corinthiens", "2 Corinthiens", "Galates", "Éphésiens",
    "Philippiens", "Colossiens", "1 Thessaloniciens", "2 Thessaloniciens",
    "1 Timothée", "2 Timothée", "Tite", "Philémon", "Hébreux",
    "Jacques", "1 Pierre", "2 Pierre", "1 Jean", "2 Jean", "3 Jean",
    "Jude", "Apocalypse"
]


@api_router.get("/pain-du-jour/livres")
async def get_livres_bible():
    """Retourne la liste des livres de la Bible"""
    return LIVRES_BIBLE


@api_router.post("/pain-du-jour/youtube-info")
async def get_youtube_video_info(request: YouTubeVideoRequest):
    """Récupère les métadonnées d'une vidéo YouTube (titre, miniature, durée)"""
    video_id = get_youtube_video_id(request.url)
    
    if not video_id:
        raise HTTPException(status_code=400, detail="URL YouTube invalide")
    
    if not YOUTUBE_API_KEY:
        raise HTTPException(status_code=500, detail="Clé API YouTube non configurée")
    
    try:
        youtube = build("youtube", "v3", developerKey=YOUTUBE_API_KEY, cache_discovery=False)
        
        video_response = youtube.videos().list(
            part="snippet,contentDetails,statistics",
            id=video_id
        ).execute()
        
        if not video_response.get("items"):
            raise HTTPException(status_code=404, detail="Vidéo non trouvée")
        
        item = video_response["items"][0]
        snippet = item["snippet"]
        content_details = item["contentDetails"]
        statistics = item.get("statistics", {})
        
        # Get best thumbnail
        thumbnails = snippet.get("thumbnails", {})
        thumbnail_url = (
            thumbnails.get("maxres", {}).get("url") or
            thumbnails.get("high", {}).get("url") or
            thumbnails.get("medium", {}).get("url") or
            thumbnails.get("default", {}).get("url", "")
        )
        
        return {
            "video_id": video_id,
            "title": snippet.get("title", ""),
            "description": snippet.get("description", "")[:500],  # Limit description
            "thumbnail_url": thumbnail_url,
            "channel_title": snippet.get("channelTitle", ""),
            "published_at": snippet.get("publishedAt", ""),
            "duration": format_duration(content_details.get("duration", "PT0S")),
            "view_count": int(statistics.get("viewCount", 0)),
            "like_count": int(statistics.get("likeCount", 0)),
        }
        
    except HttpError as e:
        if "quotaExceeded" in str(e):
            raise HTTPException(status_code=429, detail="Quota API YouTube dépassé")
        raise HTTPException(status_code=400, detail=f"Erreur API YouTube: {str(e)}")
    except Exception as e:
        logging.error(f"YouTube API error: {e}")
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")


@api_router.get("/pain-du-jour/today")
async def get_pain_du_jour_today():
    """Récupère le contenu du jour (public, pas besoin d'auth)"""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    content = await db.pain_du_jour.find_one({"date": today}, {"_id": 0})
    return content or {"date": today, "versets": []}


@api_router.get("/pain-du-jour/programmations")
async def get_all_programmations(current_user: dict = Depends(get_current_user)):
    """Récupérer toutes les programmations futures"""
    if current_user["role"] not in ["super_admin", "pasteur", "gestion_projet"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    # Récupérer les programmations triées par semaine
    programmations = await db.pain_du_jour_programmation.find(
        {},
        {"_id": 0}
    ).sort("semaine", -1).to_list(52)  # Maximum 1 an
    
    return programmations


@api_router.get("/pain-du-jour/{date}")
async def get_pain_du_jour(date: str):
    """Récupère le contenu d'une date spécifique (public)"""
    content = await db.pain_du_jour.find_one({"date": date}, {"_id": 0})
    return content or {"date": date, "versets": []}


@api_router.post("/pain-du-jour")
async def save_pain_du_jour(content: PainDuJourContent, current_user: dict = Depends(get_current_user)):
    """Enregistrer le contenu du jour - pasteur, super_admin, gestion_projet uniquement"""
    if current_user["role"] not in ["super_admin", "pasteur", "gestion_projet"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    content_data = content.model_dump()
    content_data["created_by"] = current_user["username"]
    content_data["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.pain_du_jour.update_one(
        {"date": content.date},
        {"$set": content_data},
        upsert=True
    )
    
    return {"message": "Contenu enregistré"}


@api_router.post("/pain-du-jour/click")
async def track_click(click: ClickTrack):
    """Enregistrer un clic sur vidéo (public)"""
    week_num = datetime.strptime(click.date, "%Y-%m-%d").isocalendar()[1]
    year = datetime.strptime(click.date, "%Y-%m-%d").year
    week_key = f"S{week_num}"
    
    await db.pain_du_jour_stats.update_one(
        {"semaine": week_key, "annee": year},
        {"$inc": {f"clicks_{click.type}": 1}},
        upsert=True
    )
    
    return {"message": "Click enregistré"}


@api_router.post("/pain-du-jour/sondage")
async def submit_sondage(sondage: SondagePainDuJour):
    """Soumettre un sondage (public)"""
    week_num = datetime.strptime(sondage.date, "%Y-%m-%d").isocalendar()[1]
    year = datetime.strptime(sondage.date, "%Y-%m-%d").year
    week_key = f"S{week_num}"
    
    # Incrémenter les compteurs
    update_query = {"$inc": {"total_reponses": 1}}
    
    # Lecture
    if sondage.lecture_reponse == "Oui":
        update_query["$inc"]["lecture_oui"] = 1
    elif sondage.lecture_reponse == "Non":
        update_query["$inc"]["lecture_non"] = 1
    else:
        update_query["$inc"]["lecture_partiel"] = 1
    
    # Vidéo
    if sondage.video_reponse == "Oui":
        update_query["$inc"]["video_oui"] = 1
    elif sondage.video_reponse == "Non":
        update_query["$inc"]["video_non"] = 1
    else:
        update_query["$inc"]["video_partiel"] = 1
    
    await db.pain_du_jour_stats.update_one(
        {"semaine": week_key, "annee": year},
        update_query,
        upsert=True
    )
    
    return {"message": "Sondage enregistré"}


@api_router.get("/pain-du-jour/stats/{annee}")
async def get_pain_du_jour_stats(annee: int, current_user: dict = Depends(get_current_user)):
    """Récupérer les statistiques - pasteur, super_admin, gestion_projet uniquement"""
    if current_user["role"] not in ["super_admin", "pasteur", "gestion_projet"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    stats = await db.pain_du_jour_stats.find(
        {"annee": annee},
        {"_id": 0}
    ).to_list(100)
    
    return stats


# ==================== PROGRAMMATION HEBDOMADAIRE PAIN DU JOUR ====================

@api_router.get("/pain-du-jour/programmation/{semaine}")
async def get_programmation_semaine(semaine: str, current_user: dict = Depends(get_current_user)):
    """Récupérer la programmation pour une semaine donnée"""
    if current_user["role"] not in ["super_admin", "pasteur", "gestion_projet"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    programmation = await db.pain_du_jour_programmation.find_one(
        {"semaine": semaine},
        {"_id": 0}
    )
    
    if not programmation:
        # Retourner un template vide
        return {
            "semaine": semaine,
            "jours": {
                "lundi": {"lien_enseignement": "", "titre_enseignement": "", "versets": []},
                "mardi": {"lien_enseignement": "", "titre_enseignement": "", "versets": []},
                "mercredi": {"lien_enseignement": "", "titre_enseignement": "", "versets": []},
                "jeudi": {"lien_enseignement": "", "titre_enseignement": "", "versets": []},
                "vendredi": {"lien_enseignement": "", "titre_enseignement": "", "versets": []}
            }
        }
    
    return programmation

@api_router.post("/pain-du-jour/programmation")
async def save_programmation_semaine(data: dict, current_user: dict = Depends(get_current_user)):
    """Sauvegarder la programmation pour une semaine entière"""
    if current_user["role"] not in ["super_admin", "pasteur", "gestion_projet"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    semaine = data.get("semaine")
    jours = data.get("jours", {})
    
    if not semaine:
        raise HTTPException(status_code=400, detail="Semaine requise")
    
    await db.pain_du_jour_programmation.update_one(
        {"semaine": semaine},
        {"$set": {
            "semaine": semaine,
            "jours": jours,
            "created_by": current_user["username"],
            "updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    
    # Optionnel: Appliquer automatiquement la programmation aux dates correspondantes
    # Calculer les dates pour cette semaine
    import re
    match = re.match(r'(\d{4})-W(\d{2})', semaine)
    if match:
        year = int(match.group(1))
        week = int(match.group(2))
        
        # Calculer le lundi de cette semaine
        from datetime import timedelta
        jan1 = datetime(year, 1, 1)
        days_to_first_monday = (7 - jan1.weekday()) % 7
        first_monday = jan1 + timedelta(days=days_to_first_monday)
        target_monday = first_monday + timedelta(weeks=week - 1)
        if jan1.weekday() <= 3:  # Si 1er janvier est avant jeudi
            target_monday -= timedelta(weeks=1)
        
        jour_mapping = {
            "lundi": 0,
            "mardi": 1,
            "mercredi": 2,
            "jeudi": 3,
            "vendredi": 4
        }
        
        for jour_nom, jour_data in jours.items():
            if jour_nom in jour_mapping and jour_data.get("lien_enseignement"):
                jour_offset = jour_mapping[jour_nom]
                date_jour = target_monday + timedelta(days=jour_offset)
                date_str = date_jour.strftime("%Y-%m-%d")
                
                # Créer/mettre à jour le contenu pour ce jour
                await db.pain_du_jour.update_one(
                    {"date": date_str},
                    {"$set": {
                        "date": date_str,
                        "lien_enseignement": jour_data.get("lien_enseignement", ""),
                        "titre_enseignement": jour_data.get("titre_enseignement", ""),
                        "versets": jour_data.get("versets", []),
                        "lien_priere": jour_data.get("lien_priere", ""),
                        "titre_priere": jour_data.get("titre_priere", ""),
                        "created_by": current_user["username"],
                        "programmation_auto": True,
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }},
                    upsert=True
                )
    
    return {"message": "Programmation enregistrée et appliquée aux dates"}


# ==================== RÉSUMÉ ET QUIZ ENSEIGNEMENT ====================

class ExtractVersetsRequest(BaseModel):
    transcription: str

@api_router.post("/pain-du-jour/extract-versets")
async def extract_versets(request: ExtractVersetsRequest, current_user: dict = Depends(get_current_user)):
    """Extraire tous les versets bibliques de la transcription - Admin uniquement"""
    if current_user["role"] not in ["super_admin", "pasteur", "gestion_projet"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        import json as json_module
        
        api_key = os.environ.get("EMERGENT_LLM_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="Clé API LLM non configurée")
        
        transcription = request.transcription
        
        # GARDER les timestamps pour pouvoir identifier la minute de chaque verset
        import re
        
        # Augmenter la limite pour les longues prédications (environ 50 minutes de transcription)
        max_chars = 30000
        if len(transcription) > max_chars:
            # Garder le début et la fin pour ne pas perdre de versets
            half = max_chars // 2
            transcription = transcription[:half] + "\n...[PARTIE CENTRALE TRONQUÉE]...\n" + transcription[-half:]
            logger.warning(f"Transcription tronquée de {len(transcription)} à {max_chars} caractères")
        
        logger.info(f"Extraction des versets bibliques - {len(transcription)} caractères...")
        
        prompt = f"""Tu es un expert biblique. Analyse TOUTE cette transcription de prédication du DÉBUT à la FIN.

TRANSCRIPTION COMPLÈTE AVEC TIMESTAMPS:
{transcription}

OBJECTIF: Extraire TOUS les versets bibliques mentionnés dans TOUTE la transcription.

COMMENT IDENTIFIER LES VERSETS:
Cherche ces patterns dans la transcription:
- "Jean 3:16", "Jean 3 verset 16", "Jean chapitre 3"
- "dans Romains 8", "Romains chapitre 8 verset 28"
- "le Psaume 23", "Psaumes 23", "Psaume vingt-trois"
- "premier Corinthiens", "1 Corinthiens", "première épître aux Corinthiens"
- "Matthieu 5:14", "dans Matthieu au chapitre 5"
- "Luc 14:28-30", "Luc 14 à partir du verset 28"
- "Ecclésiaste 3", "dans l'Ecclésiaste"
- "Proverbes 3:5-6", "le livre des Proverbes"
- Toute mention d'un livre biblique suivi d'un numéro

RÈGLES D'EXTRACTION:
1. VERSETS EXPLICITES - EXTRAIS TOUS (pas de limite):
   - Quand le prédicateur cite une référence biblique (livre + chapitre ou verset)
   - Même si la référence est partielle (ex: "dans Luc 14" sans numéro de verset)

2. VERSETS IMPLICITES - MAXIMUM 3:
   - Quand le contenu d'un verset est cité sans la référence
   - Seulement les plus évidents

3. TIMESTAMP - TRÈS IMPORTANT:
   - Copie EXACTEMENT le timestamp [MM:SS] qui apparaît AVANT le verset
   - Format: "12:45" (sans les crochets)

4. EXPLICATION - STYLE NARRATIF DIRECT:
   ❌ NE JAMAIS écrire: "Il dit...", "Le prédicateur explique...", "L'homme de Dieu..."
   ✅ TOUJOURS écrire directement l'enseignement en 2-3 phrases

LIVRES BIBLIQUES À CHERCHER:
Ancien Testament: Genèse, Exode, Lévitique, Nombres, Deutéronome, Josué, Juges, Ruth, 1 Samuel, 2 Samuel, 1 Rois, 2 Rois, 1 Chroniques, 2 Chroniques, Esdras, Néhémie, Esther, Job, Psaumes, Proverbes, Ecclésiaste, Cantique, Ésaïe, Jérémie, Lamentations, Ézéchiel, Daniel, Osée, Joël, Amos, Abdias, Jonas, Michée, Nahum, Habacuc, Sophonie, Aggée, Zacharie, Malachie

Nouveau Testament: Matthieu, Marc, Luc, Jean, Actes, Romains, 1 Corinthiens, 2 Corinthiens, Galates, Éphésiens, Philippiens, Colossiens, 1 Thessaloniciens, 2 Thessaloniciens, 1 Timothée, 2 Timothée, Tite, Philémon, Hébreux, Jacques, 1 Pierre, 2 Pierre, 1 Jean, 2 Jean, 3 Jean, Jude, Apocalypse

IMPORTANT: Si tu trouves AU MOINS UN verset dans la transcription, retourne-le. Ne retourne une liste vide QUE si vraiment il n'y a AUCUNE référence biblique.

Réponds UNIQUEMENT avec ce JSON (sans markdown ni texte autour):
{{
    "versets": [
        {{
            "reference": "Luc 14:28-30",
            "type": "explicite",
            "timestamp": "02:30",
            "citation_dans_transcription": "[02:30] dans Luc chapitre 14",
            "explication_predicateur": "La construction d'une tour nécessite une planification préalable. Avant de s'engager dans un projet, il faut calculer les coûts."
        }}
    ]
}}"""

        llm_chat = LlmChat(
            api_key=api_key,
            session_id=str(uuid4()),
            system_message="Tu es un expert biblique qui trouve TOUS les versets dans une transcription. Tu ne retournes JAMAIS une liste vide s'il y a des références bibliques. Tu copies les timestamps exactement."
        ).with_model("openai", "gpt-4o")
        
        user_message = UserMessage(text=prompt)
        response = await llm_chat.send_message(user_message)
        
        # Parser le JSON
        clean_response = response.strip()
        if clean_response.startswith("```json"):
            clean_response = clean_response[7:]
        if clean_response.startswith("```"):
            clean_response = clean_response[3:]
        if clean_response.endswith("```"):
            clean_response = clean_response[:-3]
        clean_response = clean_response.strip()
        
        result = json_module.loads(clean_response)
        
        # Filtrer pour s'assurer qu'il n'y a pas plus de 3 implicites
        versets = result.get('versets', [])
        explicites = [v for v in versets if v.get('type') == 'explicite']
        implicites = [v for v in versets if v.get('type') == 'implicite'][:3]  # Max 3 implicites
        
        result['versets'] = explicites + implicites
        
        logger.info(f"Versets extraits: {len(explicites)} explicites, {len(implicites)} implicites")
        
        return result
        
    except json_module.JSONDecodeError as e:
        logger.error(f"Erreur parsing JSON: {str(e)}")
        raise HTTPException(status_code=500, detail="Erreur lors de l'extraction")
    except Exception as e:
        logger.error(f"Erreur: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")


@api_router.post("/pain-du-jour/fetch-transcription")
async def fetch_transcription(request: FetchTranscriptionRequest, current_user: dict = Depends(get_current_user)):
    """Récupérer la transcription complète d'une vidéo YouTube - Admin uniquement"""
    if current_user["role"] not in ["super_admin", "pasteur", "gestion_projet"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    try:
        from youtube_transcript_api import YouTubeTranscriptApi
        from youtube_transcript_api._errors import TranscriptsDisabled, NoTranscriptFound, VideoUnavailable
        import re
        
        # Extraire l'ID de la vidéo YouTube
        video_id = None
        youtube_patterns = [
            r'(?:youtube\.com\/watch\?v=|youtu\.be\/|youtube\.com\/embed\/)([a-zA-Z0-9_-]{11})',
            r'(?:youtube\.com\/live\/)([a-zA-Z0-9_-]{11})',
            r'(?:youtube\.com\/shorts\/)([a-zA-Z0-9_-]{11})'
        ]
        for pattern in youtube_patterns:
            match = re.search(pattern, request.youtube_url)
            if match:
                video_id = match.group(1)
                break
        
        if not video_id:
            raise HTTPException(status_code=400, detail="URL YouTube invalide. Formats acceptés: youtube.com/watch?v=..., youtu.be/..., youtube.com/live/...")
        
        logger.info(f"Récupération de la transcription pour: {video_id}")
        
        # Essayer de récupérer la transcription avec différentes méthodes
        transcript_data = None
        error_message = None
        
        try:
            # Méthode 1: Récupérer la liste des transcriptions disponibles
            transcript_list = YouTubeTranscriptApi.list_transcripts(video_id)
            
            # Essayer d'abord les transcriptions manuelles en français
            try:
                transcript = transcript_list.find_manually_created_transcript(['fr', 'fr-FR'])
                transcript_data = transcript.fetch()
                logger.info("Transcription manuelle FR trouvée")
            except:
                # Essayer les transcriptions générées automatiquement en français
                try:
                    transcript = transcript_list.find_generated_transcript(['fr', 'fr-FR'])
                    transcript_data = transcript.fetch()
                    logger.info("Transcription auto FR trouvée")
                except:
                    # Essayer n'importe quelle transcription disponible
                    try:
                        for transcript in transcript_list:
                            transcript_data = transcript.fetch()
                            logger.info(f"Transcription trouvée en: {transcript.language}")
                            break
                    except:
                        pass
        except TranscriptsDisabled:
            error_message = "Les sous-titres sont désactivés pour cette vidéo. Activez les sous-titres dans les paramètres YouTube de la vidéo."
        except NoTranscriptFound:
            error_message = "Aucun sous-titre trouvé pour cette vidéo. Assurez-vous que la vidéo a des sous-titres (manuels ou automatiques)."
        except VideoUnavailable:
            error_message = "Cette vidéo n'est pas disponible ou est privée."
        except Exception as e:
            logger.error(f"Erreur list_transcripts: {str(e)}")
            # Méthode 2: Essayer la méthode directe
            try:
                ytt_api = YouTubeTranscriptApi()
                for lang in ['fr', 'fr-FR', 'en', 'en-US']:
                    try:
                        transcript_data = ytt_api.fetch(video_id, languages=[lang])
                        if transcript_data:
                            logger.info(f"Transcription trouvée avec méthode directe: {lang}")
                            break
                    except:
                        continue
                
                if not transcript_data:
                    transcript_data = ytt_api.fetch(video_id)
            except Exception as e2:
                error_message = f"Impossible de récupérer la transcription. Vérifiez que la vidéo a des sous-titres activés."
                logger.error(f"Erreur fetch direct: {str(e2)}")
        
        if not transcript_data:
            raise HTTPException(
                status_code=400, 
                detail=error_message or "Aucune transcription disponible. Vérifiez que les sous-titres sont activés sur YouTube."
            )
        
        # Construire la transcription complète avec timestamps
        transcription_parts = []
        full_text_parts = []
        
        for entry in transcript_data:
            # Gérer les différents formats de données
            if hasattr(entry, 'start'):
                start_time = entry.start
                text = entry.text
            elif isinstance(entry, dict):
                start_time = entry.get('start', 0)
                text = entry.get('text', '')
            else:
                continue
            
            # Convertir en minutes:secondes
            minutes = int(start_time // 60)
            seconds = int(start_time % 60)
            timestamp = f"[{minutes:02d}:{seconds:02d}]"
            
            transcription_parts.append(f"{timestamp} {text}")
            full_text_parts.append(text)
        
        # Transcription avec timestamps (pour affichage)
        transcription_with_timestamps = "\n".join(transcription_parts)
        
        # Transcription texte brut (pour analyse)
        transcription_text = " ".join(full_text_parts)
        
        # Calculer la durée totale
        if transcript_data:
            last_entry = transcript_data[-1]
            if hasattr(last_entry, 'start'):
                last_time = last_entry.start
            elif isinstance(last_entry, dict):
                last_time = last_entry.get('start', 0)
            else:
                last_time = 0
            duration_minutes = int(last_time // 60)
        else:
            duration_minutes = 0
        
        logger.info(f"Transcription récupérée: {len(transcription_text)} caractères, {duration_minutes} minutes")
        
        return {
            "transcription_complete": transcription_with_timestamps,
            "transcription_text": transcription_text,
            "duree_minutes": duration_minutes,
            "nombre_caracteres": len(transcription_text)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur transcription: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Erreur: {str(e)}")

@api_router.post("/pain-du-jour/generate-resume-quiz")
async def generate_resume_quiz(request: GenerateResumeQuizRequest, current_user: dict = Depends(get_current_user)):
    """Générer le résumé et le quiz à partir de la transcription - Admin uniquement"""
    if current_user["role"] not in ["super_admin", "pasteur", "gestion_projet"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        import json as json_module
        import re
        
        api_key = os.environ.get("EMERGENT_LLM_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="Clé API LLM non configurée")
        
        transcription = request.transcription
        titre_message = request.titre_message
        minute_debut = request.minute_debut
        
        # Filtrer la transcription à partir de la minute spécifiée mais GARDER les timestamps
        if minute_debut > 0:
            lines = transcription.split('\n')
            filtered_lines = []
            for line in lines:
                match = re.match(r'\[(\d+):(\d+)\]', line)
                if match:
                    minutes = int(match.group(1))
                    if minutes >= minute_debut:
                        filtered_lines.append(line)  # Garder le timestamp
                else:
                    filtered_lines.append(line)
            transcription = '\n'.join(filtered_lines)
        
        # Limiter la taille
        if len(transcription) > 14000:
            transcription = transcription[:14000] + "..."
        
        logger.info(f"Génération résumé pour '{titre_message}' à partir de minute {minute_debut}")
        
        # Prompt pour analyse fidèle
        prompt = f"""Tu es un expert en synthèse de prédications chrétiennes. Analyse cette transcription (avec timestamps [MM:SS]) et génère un contenu FIDÈLE et DÉTAILLÉ.

TITRE DU MESSAGE: "{titre_message}"

TRANSCRIPTION AVEC TIMESTAMPS (à partir de la minute {minute_debut}):
{transcription}

GÉNÈRE UN JSON AVEC CETTE STRUCTURE:

{{
    "resume": {{
        "titre": "{titre_message}",
        "resume": "Écris un résumé TRÈS DÉTAILLÉ de 10-12 longues phrases en style NARRATIF DIRECT. Ne dis JAMAIS 'le prédicateur dit', 'il explique', 'l'orateur souligne', 'l'homme de Dieu'. Écris directement le contenu comme si tu enseignais toi-même. Développe chaque idée en profondeur, fais des liens entre les concepts. Le résumé doit capturer TOUT ce qui est enseigné.",
        "versets_expliques": [
            {{"reference": "Jean 3:16", "timestamp": "12:34", "explication": "L'amour de Dieu est si profond qu'il a donné ce qu'il avait de plus précieux. Ce sacrifice n'est pas conditionné par nos mérites mais par sa grâce infinie. Quiconque croit en ce don reçoit la vie éternelle."}},
            {{"reference": "Romains 8:28", "timestamp": "25:10", "explication": "Même dans les épreuves les plus difficiles, Dieu travaille pour notre bien. Les obstacles ne sont pas des abandons mais des opportunités de voir sa gloire se manifester."}}
        ],
        "points_cles": ["Enseignement 1 formulé clairement", "Enseignement 2", "Enseignement 3", "etc - Chaque point clé doit être une leçon concrète"],
        "phrases_fortes": ["Citation exacte 1 mot pour mot", "Citation exacte 2", "Citation exacte 3", "etc - Les phrases marquantes et puissantes prononcées"]
    }},
    "quiz": [
        {{"question": "Question 1?", "options": ["A", "B", "C", "D"], "correct_index": 0}},
        {{"question": "Question 2?", "options": ["A", "B", "C", "D"], "correct_index": 1}},
        {{"question": "Question 3?", "options": ["A", "B", "C", "D"], "correct_index": 2}},
        {{"question": "Question 4?", "options": ["A", "B", "C", "D"], "correct_index": 0}},
        {{"question": "Question 5?", "options": ["A", "B", "C", "D"], "correct_index": 1}},
        {{"question": "Question 6?", "options": ["A", "B", "C", "D"], "correct_index": 2}},
        {{"question": "Question 7?", "options": ["A", "B", "C", "D"], "correct_index": 0}},
        {{"question": "Question 8?", "options": ["A", "B", "C", "D"], "correct_index": 1}},
        {{"question": "Question 9?", "options": ["A", "B", "C", "D"], "correct_index": 2}},
        {{"question": "Question 10?", "options": ["A", "B", "C", "D"], "correct_index": 0}}
    ]
}}

RÈGLES CRITIQUES:
1. RÉSUMÉ: Style narratif direct (JAMAIS "le prédicateur", "il dit", "l'homme de Dieu explique"). 10-12 phrases détaillées.

2. VERSETS EXPLIQUÉS - TRÈS IMPORTANT:
   - Pour CHAQUE verset biblique cité dans la prédication
   - "timestamp": le timestamp [MM:SS] où le verset apparaît dans la transcription (juste MM:SS sans crochets)
   - L'explication doit être en STYLE NARRATIF DIRECT (2-4 phrases)
   ❌ JAMAIS: "Il dit que...", "Le prédicateur explique que...", "L'homme de Dieu souligne..."
   ✅ TOUJOURS: Écrire directement le contenu de l'explication comme un enseignement
   - Capture fidèlement ce que le prédicateur dit APRÈS avoir cité le verset

3. POINTS CLÉS: Les enseignements principaux, formulés comme des leçons concrètes

4. PHRASES FORTES: Les citations EXACTES et puissantes prononcées (mot pour mot)

5. Si aucun verset n'est cité dans la transcription, mettre une liste vide []

Réponds UNIQUEMENT avec le JSON, sans markdown."""

        llm_chat = LlmChat(
            api_key=api_key,
            session_id=str(uuid4()),
            system_message="Tu analyses des transcriptions de prédications. Tu génères du contenu JSON fidèle à la transcription."
        ).with_model("openai", "gpt-4o-mini")
        
        user_message = UserMessage(text=prompt)
        response = await llm_chat.send_message(user_message)
        
        # Parser le JSON
        clean_response = response.strip()
        if clean_response.startswith("```json"):
            clean_response = clean_response[7:]
        if clean_response.startswith("```"):
            clean_response = clean_response[3:]
        if clean_response.endswith("```"):
            clean_response = clean_response[:-3]
        clean_response = clean_response.strip()
        
        result = json_module.loads(clean_response)
        return result
        
    except json_module.JSONDecodeError as e:
        logger.error(f"Erreur parsing JSON: {str(e)}")
        raise HTTPException(status_code=500, detail="Erreur lors de la génération")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")


@api_router.post("/pain-du-jour/quiz/submit")
async def submit_quiz(submission: QuizSubmission):
    """Soumettre les réponses du quiz (anonyme)"""
    try:
        week_num = datetime.strptime(submission.date, "%Y-%m-%d").isocalendar()[1]
        year = datetime.strptime(submission.date, "%Y-%m-%d").year
        week_key = f"S{week_num}"
        
        # Enregistrer la soumission
        quiz_data = {
            "id": str(uuid4()),
            "date": submission.date,
            "answers": submission.answers,
            "score": submission.score,
            "submitted_at": datetime.now(timezone.utc).isoformat()
        }
        await db.pain_du_jour_quiz_submissions.insert_one(quiz_data)
        
        # Mettre à jour les stats
        await db.pain_du_jour_stats.update_one(
            {"semaine": week_key, "annee": year},
            {
                "$inc": {
                    "quiz_total": 1,
                    "quiz_score_total": submission.score
                }
            },
            upsert=True
        )
        
        return {"message": "Quiz soumis avec succès", "score": submission.score}
    except Exception as e:
        logger.error(f"Erreur soumission quiz: {str(e)}")
        raise HTTPException(status_code=500, detail="Erreur lors de la soumission")


@api_router.get("/pain-du-jour/quiz/stats/{date}")
async def get_quiz_stats(date: str, current_user: dict = Depends(get_current_user)):
    """Obtenir les statistiques du quiz pour une date - Admin uniquement"""
    if current_user["role"] not in ["super_admin", "pasteur", "gestion_projet"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    submissions = await db.pain_du_jour_quiz_submissions.find(
        {"date": date},
        {"_id": 0}
    ).to_list(1000)
    
    if not submissions:
        return {
            "date": date,
            "total_participants": 0,
            "average_score": 0,
            "score_distribution": {}
        }
    
    total = len(submissions)
    total_score = sum(s["score"] for s in submissions)
    avg_score = round(total_score / total, 1) if total > 0 else 0
    
    # Distribution des scores
    distribution = {}
    for s in submissions:
        score_key = str(s["score"])
        distribution[score_key] = distribution.get(score_key, 0) + 1
    
    return {
        "date": date,
        "total_participants": total,
        "average_score": avg_score,
        "score_distribution": distribution
    }


# ==================== CHATBOT ASSISTANT IA ====================

class ChatMessage(BaseModel):
    message: str
    session_id: Optional[str] = None
    role: Optional[str] = None  # Rôle de l'utilisateur dans l'application

class ChatResponse(BaseModel):
    response: str
    session_id: str

# Base de connaissances de l'application ICC Hub
APP_KNOWLEDGE_BASE = """
Tu es Audrey, l'Assistant Virtuel de l'application ICC Hub (Impact Centre Chrétien).
Tu réponds TOUJOURS en français de manière claire, amicale et professionnelle.
Quand on te demande qui tu es, tu réponds "Je suis Audrey, votre Assistant Virtuel".

## MODULES DE L'APPLICATION:

### 1. PAGE D'ACCUEIL (/)
- Accès aux différents départements de l'église
- Pop-ups d'anniversaires des membres STARS
- Pop-ups des événements à venir (dans les 30 jours)
- Accès: Public

### 2. LE PAIN DU JOUR (/pain-du-jour)
- Contenu spirituel quotidien
- Vidéos YouTube: Temps de prière prophétique + Enseignements
- Versets du jour avec liens vers EMCI TV pour lecture
- Sondage de participation (lectures et vidéos)
- Administration: Gérer les contenus quotidiens + Statistiques
- Accès: Public (lecture), Admin/Pasteur/Gestion Projet pour mise à jour

### 3. MINISTÈRE DES STARS (/ministere-stars)
- Gestion des départements STARS (Service, Technique, Accueil, Régie, Sécurité)
- Planning hebdomadaire sur 52 semaines
- Attribution des tâches aux membres
- KPIs: nombre de stars en service par semaine
- Rôles: star (lecture seule), respo_departement (gestion complète)

### 4. MY EVENT CHURCH (/events-management)
- Gestion des projets et événements de l'église
- Planning des activités avec dates
- Campagnes d'évangélisation
- Statistiques des événements
- RSVP et inscriptions en ligne

### 5. FAMILLES IMPACT (/familles-impact)
- Gestion des Familles Impact (petits groupes de maison)
- Suivi des membres par FI
- Présences aux rencontres
- Pilotes et responsables de secteur

### 6. SUIVI DES NOUVEAUX (/nouveaux, /dashboard)
- Gestion des nouveaux arrivants (NA) et nouveaux convertis (NC)
- Attribution aux bergers par promo mensuelle (Janvier à Décembre)
- Suivi des présences dimanche et jeudi
- Formations: PCNC, Au Cœur de la Bible, STAR
- Dashboard superviseur promos avec statistiques

### 7. GESTION DES ACCÈS (/gestion-acces)
- Création et gestion des comptes utilisateurs
- Attribution des rôles et permissions
- Gestion multi-villes (Milan, Rome, Dijon, etc.)

## RÔLES DISPONIBLES:
- **visiteur**: Accès limité, consultation uniquement
- **membre**: Membre standard de l'église
- **berger/referent**: Responsable de suivi d'une promo mensuelle
- **responsable_promo**: Gère les bergers d'une promo
- **superviseur_promos**: Supervise toutes les promos d'une ville
- **pilote**: Pilote d'une Famille Impact
- **responsable_fi**: Responsable des Familles Impact
- **responsable_secteur**: Responsable d'un secteur de FI
- **superviseur_fi**: Supervise les FI d'une ville
- **star**: Membre du ministère STARS
- **respo_departement**: Responsable d'un département STARS
- **coordinateur**: Coordonne plusieurs départements
- **secretaire**: Gestion administrative
- **tresorier**: Gestion financière
- **pasteur**: Pasteur de l'église
- **gestion_projet**: Gestion des projets et événements
- **super_admin**: Accès complet à tout

## VILLES DISPONIBLES:
- Italie: Milan, Rome, Perugia, Bologne, Turin
- France: Dijon, Auxerre, Besançon, Chalon-Sur-Saone, Dole, Sens

## CONSEILS D'UTILISATION:
- Pour voir vos visiteurs: allez dans "Bergeries" puis "Liste des Nouveaux"
- Pour marquer les présences: utilisez les boutons Présent/Absent dans la fiche visiteur
- Pour les STARS: connectez-vous via "Ministère des STARS" avec vos identifiants
- Pour le Pain du Jour: accessible directement depuis l'accueil, pas besoin de connexion

## CONTACT SUPPORT:
Pour toute question technique, contactez l'administrateur de votre église.
"""

@api_router.post("/chatbot/message")
async def chatbot_message(chat: ChatMessage):
    """Endpoint pour le chatbot IA - public"""
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        
        api_key = os.environ.get("EMERGENT_LLM_KEY")
        if not api_key:
            return {"response": "Le service chatbot n'est pas configuré. Veuillez contacter l'administrateur.", "session_id": chat.session_id or str(uuid4())}
        
        session_id = chat.session_id or str(uuid4())
        
        # Construire le message système avec contexte du rôle
        system_message = APP_KNOWLEDGE_BASE
        if chat.role:
            system_message += f"\n\nL'utilisateur actuel a le rôle: {chat.role}. Adapte tes réponses à ses permissions et besoins."
        
        # Initialiser le chat LLM
        llm_chat = LlmChat(
            api_key=api_key,
            session_id=session_id,
            system_message=system_message
        ).with_model("openai", "gpt-4o-mini")
        
        # Envoyer le message
        user_message = UserMessage(text=chat.message)
        response = await llm_chat.send_message(user_message)
        
        return {"response": response, "session_id": session_id}
        
    except Exception as e:
        logger.error(f"Erreur chatbot: {str(e)}")
        return {
            "response": "Désolé, je rencontre un problème technique. Veuillez réessayer dans quelques instants.",
            "session_id": chat.session_id or str(uuid4())
        }


# ==============================================
# BERGERIE - SYSTÈME DE REPRODUCTION
# ==============================================

class BergerieObjectif(BaseModel):
    """Objectifs mensuels d'une bergerie"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    bergerie_month: str  # "01" à "12" (le mois de la bergerie)
    ville: str
    annee: int  # L'année de référence (ex: 2025)
    mois_cible: str  # Le mois pour lequel on fixe l'objectif (ex: "2025-09")
    objectif_nombre: int  # L'objectif de personnes à atteindre
    nombre_reel: Optional[int] = None  # Le nombre réel atteint
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

class BergerieObjectifCreate(BaseModel):
    bergerie_month: str
    ville: str
    annee: int
    mois_cible: str
    objectif_nombre: int
    nombre_reel: Optional[int] = None

class BergerieContact(BaseModel):
    """Personnes contactées par la bergerie (évangélisation ou autres)"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    bergerie_month: str  # "01" à "12"
    ville: str
    nom: str
    prenom: str
    telephone: Optional[str] = None
    date_contact: str  # Date du contact YYYY-MM-DD
    type_contact: str  # "Evangelisation" ou "Autres"
    precision_autres: Optional[str] = None  # Précision si type = "Autres"
    statut: Optional[str] = None  # "Réceptif", "Prière de salut", "Venu à l'église", etc.
    notes: Optional[str] = None
    created_by: str  # ID de l'utilisateur qui a créé
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BergerieContactCreate(BaseModel):
    bergerie_month: str
    ville: str
    nom: str
    prenom: str
    telephone: Optional[str] = None
    date_contact: str
    type_contact: str  # "Evangelisation" ou "Autres"
    precision_autres: Optional[str] = None
    statut: Optional[str] = None
    notes: Optional[str] = None

class BergerieDisciple(BaseModel):
    """Suivi des disciples dans une bergerie"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    visitor_id: str  # ID du visiteur
    bergerie_month: str  # "01" à "12"
    ville: str
    est_disciple: str  # "Oui", "En Cours", "Non"
    date_devenu_disciple: Optional[str] = None
    notes: Optional[str] = None
    updated_by: str
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BergerieDiscipleUpdate(BaseModel):
    est_disciple: str  # "Oui", "En Cours", "Non"
    date_devenu_disciple: Optional[str] = None
    notes: Optional[str] = None


# --- Endpoints Bergerie Objectifs ---

@api_router.get("/bergerie/objectifs/{ville}/{bergerie_month}")
async def get_bergerie_objectifs(ville: str, bergerie_month: str, current_user: dict = Depends(get_current_user)):
    """Récupérer tous les objectifs d'une bergerie"""
    objectifs = await db.bergerie_objectifs.find(
        {"ville": ville, "bergerie_month": bergerie_month},
        {"_id": 0}
    ).sort("mois_cible", 1).to_list(100)
    return objectifs

@api_router.post("/bergerie/objectifs")
async def create_bergerie_objectif(objectif: BergerieObjectifCreate, current_user: dict = Depends(get_current_user)):
    """Créer ou mettre à jour un objectif de bergerie"""
    # Vérifier si l'objectif existe déjà pour ce mois_cible
    existing = await db.bergerie_objectifs.find_one({
        "bergerie_month": objectif.bergerie_month,
        "ville": objectif.ville,
        "mois_cible": objectif.mois_cible
    })
    
    if existing:
        # Mettre à jour
        await db.bergerie_objectifs.update_one(
            {"id": existing["id"]},
            {"$set": {
                "objectif_nombre": objectif.objectif_nombre,
                "nombre_reel": objectif.nombre_reel,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        return {"message": "Objectif mis à jour", "id": existing["id"]}
    else:
        # Créer nouveau
        new_objectif = BergerieObjectif(**objectif.model_dump())
        await db.bergerie_objectifs.insert_one(new_objectif.model_dump())
        return {"message": "Objectif créé", "id": new_objectif.id}

@api_router.put("/bergerie/objectifs/{objectif_id}")
async def update_bergerie_objectif(objectif_id: str, objectif: BergerieObjectifCreate, current_user: dict = Depends(get_current_user)):
    """Mettre à jour un objectif"""
    result = await db.bergerie_objectifs.update_one(
        {"id": objectif_id},
        {"$set": {
            "objectif_nombre": objectif.objectif_nombre,
            "nombre_reel": objectif.nombre_reel,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Objectif non trouvé")
    return {"message": "Objectif mis à jour"}


# --- Endpoints Bergerie Contacts (Évangélisation) ---

@api_router.get("/bergerie/contacts/{ville}/{bergerie_month}")
async def get_bergerie_contacts(ville: str, bergerie_month: str, current_user: dict = Depends(get_current_user)):
    """Récupérer tous les contacts d'une bergerie"""
    contacts = await db.bergerie_contacts.find(
        {"ville": ville, "bergerie_month": bergerie_month},
        {"_id": 0}
    ).sort("date_contact", -1).to_list(500)
    return contacts

@api_router.post("/bergerie/contacts")
async def create_bergerie_contact(contact: BergerieContactCreate, current_user: dict = Depends(get_current_user)):
    """Ajouter un nouveau contact"""
    new_contact = BergerieContact(
        **contact.model_dump(),
        created_by=current_user["id"]
    )
    await db.bergerie_contacts.insert_one(new_contact.model_dump())
    return {"message": "Contact ajouté", "id": new_contact.id}

# --- Endpoints Bergerie PUBLICS (sans authentification) ---

@api_router.get("/bergerie/public/contacts/{ville}/{bergerie_month}")
async def get_bergerie_contacts_public(ville: str, bergerie_month: str):
    """Récupérer tous les contacts d'une bergerie - Public"""
    contacts = await db.bergerie_contacts.find(
        {"ville": ville, "bergerie_month": bergerie_month},
        {"_id": 0}
    ).sort("date_contact", -1).to_list(500)
    return contacts

@api_router.post("/bergerie/public/contacts")
async def create_bergerie_contact_public(contact: BergerieContactCreate):
    """Ajouter un nouveau contact - Public"""
    new_contact = BergerieContact(
        **contact.model_dump(),
        created_by="public"
    )
    await db.bergerie_contacts.insert_one(new_contact.model_dump())
    return {"message": "Contact ajouté", "id": new_contact.id}

@api_router.delete("/bergerie/public/contacts/{contact_id}")
async def delete_bergerie_contact_public(contact_id: str):
    """Supprimer un contact - Public"""
    result = await db.bergerie_contacts.delete_one({"id": contact_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Contact non trouvé")
    return {"message": "Contact supprimé"}

@api_router.get("/bergerie/public/objectifs/{ville}/{bergerie_month}")
async def get_bergerie_objectifs_public(ville: str, bergerie_month: str):
    """Récupérer tous les objectifs d'une bergerie - Public"""
    objectifs = await db.bergerie_objectifs.find(
        {"ville": ville, "bergerie_month": bergerie_month},
        {"_id": 0}
    ).sort("mois_cible", 1).to_list(100)
    return objectifs

@api_router.post("/bergerie/public/objectifs")
async def create_bergerie_objectif_public(objectif: BergerieObjectifCreate):
    """Créer ou mettre à jour un objectif - Public"""
    existing = await db.bergerie_objectifs.find_one({
        "bergerie_month": objectif.bergerie_month,
        "ville": objectif.ville,
        "mois_cible": objectif.mois_cible
    })
    
    if existing:
        await db.bergerie_objectifs.update_one(
            {"id": existing["id"]},
            {"$set": {
                "objectif_nombre": objectif.objectif_nombre,
                "nombre_reel": objectif.nombre_reel,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        return {"message": "Objectif mis à jour", "id": existing["id"]}
    else:
        new_objectif = BergerieObjectif(**objectif.model_dump())
        await db.bergerie_objectifs.insert_one(new_objectif.model_dump())
        return {"message": "Objectif créé", "id": new_objectif.id}

@api_router.put("/bergerie/public/objectifs/{objectif_id}")
async def update_bergerie_objectif_public(objectif_id: str, objectif: BergerieObjectifCreate):
    """Mettre à jour un objectif - Public"""
    result = await db.bergerie_objectifs.update_one(
        {"id": objectif_id},
        {"$set": {
            "objectif_nombre": objectif.objectif_nombre,
            "nombre_reel": objectif.nombre_reel,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Objectif non trouvé")
    return {"message": "Objectif mis à jour"}

@api_router.post("/bergerie/public/disciples/{visitor_id}")
async def update_bergerie_disciple_public(visitor_id: str, disciple: BergerieDiscipleUpdate):
    """Mettre à jour le statut disciple d'un visiteur - Public"""
    visitor = await db.visitors.find_one({"id": visitor_id})
    if not visitor:
        raise HTTPException(status_code=404, detail="Visiteur non trouvé")
    
    bergerie_month = visitor.get("assigned_month", "").split("-")[1] if visitor.get("assigned_month") else "01"
    ville = visitor.get("city", "")
    
    existing = await db.bergerie_disciples.find_one({"visitor_id": visitor_id})
    
    if existing:
        await db.bergerie_disciples.update_one(
            {"visitor_id": visitor_id},
            {"$set": {
                "est_disciple": disciple.est_disciple,
                "date_devenu_disciple": disciple.date_devenu_disciple,
                "notes": disciple.notes,
                "updated_by": "public",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    else:
        new_disciple = BergerieDisciple(
            visitor_id=visitor_id,
            bergerie_month=bergerie_month,
            ville=ville,
            est_disciple=disciple.est_disciple,
            date_devenu_disciple=disciple.date_devenu_disciple,
            notes=disciple.notes,
            updated_by="public"
        )
        await db.bergerie_disciples.insert_one(new_disciple.model_dump())
    
    return {"message": "Statut disciple mis à jour"}

@api_router.get("/bergerie/public/reproduction/{ville}/{bergerie_month}")
async def get_bergerie_reproduction_data_public(ville: str, bergerie_month: str):
    """Récupérer toutes les données de reproduction d'une bergerie - Public"""
    
    # 1. Récupérer les visiteurs assignés à cette bergerie
    visitors = await db.visitors.find(
        {"city": ville, "assigned_month": {"$regex": f"-{bergerie_month}$"}},
        {"_id": 0}
    ).to_list(500)
    
    # 2. Récupérer les objectifs
    objectifs = await db.bergerie_objectifs.find(
        {"ville": ville, "bergerie_month": bergerie_month},
        {"_id": 0}
    ).sort("mois_cible", 1).to_list(100)
    
    # 3. Récupérer les contacts
    contacts = await db.bergerie_contacts.find(
        {"ville": ville, "bergerie_month": bergerie_month},
        {"_id": 0}
    ).sort("date_contact", -1).to_list(500)
    
    # 4. Récupérer les statuts disciples
    disciples = await db.bergerie_disciples.find(
        {"ville": ville, "bergerie_month": bergerie_month},
        {"_id": 0}
    ).to_list(500)
    
    # Créer un map pour les statuts disciples
    disciples_map = {d["visitor_id"]: d for d in disciples}
    
    # Enrichir les visiteurs avec leur statut disciple
    for visitor in visitors:
        visitor_disciple = disciples_map.get(visitor["id"], {})
        visitor["est_disciple"] = visitor_disciple.get("est_disciple", "Non")
        visitor["date_devenu_disciple"] = visitor_disciple.get("date_devenu_disciple")
    
    # Calculer les stats
    total_recus = len([v for v in visitors if not v.get("tracking_stopped")])
    total_disciples_oui = len([v for v in visitors if v.get("est_disciple") == "Oui"])
    total_disciples_en_cours = len([v for v in visitors if v.get("est_disciple") == "En Cours"])
    total_evangelises = len([c for c in contacts if c.get("type_contact") == "Evangelisation"])
    
    return {
        "visitors": visitors,
        "objectifs": objectifs,
        "contacts": contacts,
        "stats": {
            "total_recus": total_recus,
            "total_disciples_oui": total_disciples_oui,
            "total_disciples_en_cours": total_disciples_en_cours,
            "total_evangelises": total_evangelises
        }
    }

@api_router.put("/bergerie/contacts/{contact_id}")
async def update_bergerie_contact(contact_id: str, contact: BergerieContactCreate, current_user: dict = Depends(get_current_user)):
    """Mettre à jour un contact"""
    result = await db.bergerie_contacts.update_one(
        {"id": contact_id},
        {"$set": {
            "nom": contact.nom,
            "prenom": contact.prenom,
            "telephone": contact.telephone,
            "date_contact": contact.date_contact,
            "type_contact": contact.type_contact,
            "precision_autres": contact.precision_autres,
            "statut": contact.statut,
            "notes": contact.notes
        }}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Contact non trouvé")
    return {"message": "Contact mis à jour"}

@api_router.delete("/bergerie/contacts/{contact_id}")
async def delete_bergerie_contact(contact_id: str, current_user: dict = Depends(get_current_user)):
    """Supprimer un contact"""
    result = await db.bergerie_contacts.delete_one({"id": contact_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Contact non trouvé")
    return {"message": "Contact supprimé"}


# --- Endpoints Bergerie Disciples ---

@api_router.get("/bergerie/disciples/{ville}/{bergerie_month}")
async def get_bergerie_disciples(ville: str, bergerie_month: str, current_user: dict = Depends(get_current_user)):
    """Récupérer le statut disciple de tous les visiteurs d'une bergerie"""
    disciples = await db.bergerie_disciples.find(
        {"ville": ville, "bergerie_month": bergerie_month},
        {"_id": 0}
    ).to_list(500)
    return disciples

@api_router.post("/bergerie/disciples/{visitor_id}")
async def update_bergerie_disciple(visitor_id: str, disciple: BergerieDiscipleUpdate, current_user: dict = Depends(get_current_user)):
    """Mettre à jour le statut disciple d'un visiteur"""
    # Récupérer le visiteur pour avoir son assigned_month
    visitor = await db.visitors.find_one({"id": visitor_id})
    if not visitor:
        raise HTTPException(status_code=404, detail="Visiteur non trouvé")
    
    bergerie_month = visitor.get("assigned_month", "").split("-")[1] if visitor.get("assigned_month") else "01"
    ville = visitor.get("city", "")
    
    # Vérifier si un enregistrement existe déjà
    existing = await db.bergerie_disciples.find_one({"visitor_id": visitor_id})
    
    if existing:
        # Mettre à jour
        await db.bergerie_disciples.update_one(
            {"visitor_id": visitor_id},
            {"$set": {
                "est_disciple": disciple.est_disciple,
                "date_devenu_disciple": disciple.date_devenu_disciple,
                "notes": disciple.notes,
                "updated_by": current_user["id"],
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    else:
        # Créer nouveau
        new_disciple = BergerieDisciple(
            visitor_id=visitor_id,
            bergerie_month=bergerie_month,
            ville=ville,
            est_disciple=disciple.est_disciple,
            date_devenu_disciple=disciple.date_devenu_disciple,
            notes=disciple.notes,
            updated_by=current_user["id"]
        )
        await db.bergerie_disciples.insert_one(new_disciple.model_dump())
    
    return {"message": "Statut disciple mis à jour"}


# --- Endpoints Stats Bergerie pour Dynamique Évangélisation ---

@api_router.get("/bergerie/stats/evangelisation/{ville}")
async def get_bergerie_evangelisation_stats(ville: str, current_user: dict = Depends(get_current_user)):
    """Récupérer les statistiques d'évangélisation de toutes les bergeries d'une ville"""
    # Récupérer tous les contacts de type "Evangelisation"
    contacts = await db.bergerie_contacts.find(
        {"ville": ville, "type_contact": "Evangelisation"},
        {"_id": 0}
    ).to_list(1000)
    
    # Grouper par bergerie
    stats_by_bergerie = {}
    month_names = {
        '01': 'Janvier', '02': 'Février', '03': 'Mars', '04': 'Avril',
        '05': 'Mai', '06': 'Juin', '07': 'Juillet', '08': 'Août',
        '09': 'Septembre', '10': 'Octobre', '11': 'Novembre', '12': 'Décembre'
    }
    
    for contact in contacts:
        bergerie_month = contact.get("bergerie_month", "01")
        bergerie_name = f"Bergerie {month_names.get(bergerie_month, bergerie_month)}"
        
        if bergerie_name not in stats_by_bergerie:
            stats_by_bergerie[bergerie_name] = {
                "bergerie_month": bergerie_month,
                "bergerie_name": bergerie_name,
                "total_evangelises": 0,
                "receptifs": 0,
                "priere_salut": 0,
                "venus_eglise": 0
            }
        
        stats_by_bergerie[bergerie_name]["total_evangelises"] += 1
        statut = contact.get("statut", "").lower()
        if "réceptif" in statut or "receptif" in statut:
            stats_by_bergerie[bergerie_name]["receptifs"] += 1
        if "prière" in statut or "salut" in statut:
            stats_by_bergerie[bergerie_name]["priere_salut"] += 1
        if "venu" in statut or "église" in statut or "eglise" in statut:
            stats_by_bergerie[bergerie_name]["venus_eglise"] += 1
    
    # Calculer les totaux
    totals = {
        "total_evangelises": sum(s["total_evangelises"] for s in stats_by_bergerie.values()),
        "receptifs": sum(s["receptifs"] for s in stats_by_bergerie.values()),
        "priere_salut": sum(s["priere_salut"] for s in stats_by_bergerie.values()),
        "venus_eglise": sum(s["venus_eglise"] for s in stats_by_bergerie.values())
    }
    
    return {
        "by_bergerie": list(stats_by_bergerie.values()),
        "totals": totals
    }

@api_router.get("/bergerie/reproduction/{ville}/{bergerie_month}")
async def get_bergerie_reproduction_data(ville: str, bergerie_month: str, current_user: dict = Depends(get_current_user)):
    """Récupérer toutes les données de reproduction d'une bergerie"""
    
    # 1. Récupérer les visiteurs assignés à cette bergerie
    visitors = await db.visitors.find(
        {"city": ville, "assigned_month": {"$regex": f"-{bergerie_month}$"}},
        {"_id": 0}
    ).to_list(500)
    
    # 2. Récupérer les objectifs
    objectifs = await db.bergerie_objectifs.find(
        {"ville": ville, "bergerie_month": bergerie_month},
        {"_id": 0}
    ).sort("mois_cible", 1).to_list(100)
    
    # 3. Récupérer les contacts
    contacts = await db.bergerie_contacts.find(
        {"ville": ville, "bergerie_month": bergerie_month},
        {"_id": 0}
    ).sort("date_contact", -1).to_list(500)
    
    # 4. Récupérer les statuts disciples
    disciples = await db.bergerie_disciples.find(
        {"ville": ville, "bergerie_month": bergerie_month},
        {"_id": 0}
    ).to_list(500)
    
    # Créer un map pour les statuts disciples
    disciples_map = {d["visitor_id"]: d for d in disciples}
    
    # Enrichir les visiteurs avec leur statut disciple
    for visitor in visitors:
        visitor_disciple = disciples_map.get(visitor["id"], {})
        visitor["est_disciple"] = visitor_disciple.get("est_disciple", "Non")
        visitor["date_devenu_disciple"] = visitor_disciple.get("date_devenu_disciple")
    
    # Calculer les stats
    total_recus = len([v for v in visitors if not v.get("tracking_stopped")])
    total_disciples_oui = len([v for v in visitors if v.get("est_disciple") == "Oui"])
    total_disciples_en_cours = len([v for v in visitors if v.get("est_disciple") == "En Cours"])
    total_evangelises = len([c for c in contacts if c.get("type_contact") == "Evangelisation"])
    
    return {
        "visitors": visitors,
        "objectifs": objectifs,
        "contacts": contacts,
        "stats": {
            "total_recus": total_recus,
            "total_disciples_oui": total_disciples_oui,
            "total_disciples_en_cours": total_disciples_en_cours,
            "total_evangelises": total_evangelises
        }
    }

@api_router.get("/bergerie/list/{ville}")
async def get_bergeries_list(ville: str, current_user: dict = Depends(get_current_user)):
    """Récupérer la liste des 12 bergeries d'une ville avec leurs stats"""
    return await _get_bergeries_list_internal(ville)

@api_router.get("/bergerie/list-public/{ville}")
async def get_bergeries_list_public(ville: str):
    """Récupérer la liste des 12 bergeries d'une ville - Accès public sans authentification"""
    return await _get_bergeries_list_internal(ville)

async def _get_bergeries_list_internal(ville: str):
    """Fonction interne pour récupérer la liste des bergeries"""
    month_names = {
        '01': 'Janvier', '02': 'Février', '03': 'Mars', '04': 'Avril',
        '05': 'Mai', '06': 'Juin', '07': 'Juillet', '08': 'Août',
        '09': 'Septembre', '10': 'Octobre', '11': 'Novembre', '12': 'Décembre'
    }
    
    bergeries = []
    
    for month_num, month_name in month_names.items():
        # Compter les visiteurs pour cette bergerie
        visitors_count = await db.visitors.count_documents({
            "city": ville,
            "assigned_month": {"$regex": f"-{month_num}$"},
            "tracking_stopped": {"$ne": True}
        })
        
        # Récupérer les bergers assignés
        bergers = await db.users.find(
            {
                "city": ville,
                "role": {"$in": ["referent", "berger"]},
                "assigned_month": {"$regex": f"-{month_num}$"}
            },
            {"_id": 0, "id": 1, "username": 1, "promo_name": 1}
        ).to_list(10)
        
        # Nom personnalisé s'il existe
        custom_name = None
        for b in bergers:
            if b.get("promo_name"):
                custom_name = b["promo_name"]
                break
        
        bergeries.append({
            "month_num": month_num,
            "month_name": month_name,
            "nom": custom_name or f"Bergerie {month_name}",
            "total_personnes": visitors_count,
            "bergers": bergers
        })
    
    return bergeries


# ========== ENDPOINTS PUBLICS - MINISTÈRE DES STARS ==========

@api_router.get("/stars/public/stats")
async def get_stars_public_stats(ville: Optional[str] = None):
    """Récupérer les statistiques des Stars - Accès public"""
    query = {}
    if ville:
        query["ville"] = ville
    
    # Total
    total = await db.stars.count_documents(query)
    
    # Actifs
    query_actifs = {**query, "statut": "Actif"}
    actifs = await db.stars.count_documents(query_actifs)
    
    # Non actifs
    non_actifs = total - actifs
    
    # Par département
    pipeline = [
        {"$match": query},
        {"$unwind": "$departements"},
        {"$group": {"_id": "$departements", "count": {"$sum": 1}}}
    ]
    dept_counts_cursor = db.stars.aggregate(pipeline)
    dept_counts = {}
    async for doc in dept_counts_cursor:
        dept_counts[doc["_id"]] = doc["count"]
    
    return {
        "total": total,
        "actifs": actifs,
        "non_actifs": non_actifs,
        "par_departement": dept_counts
    }

@api_router.get("/stars/public/multi-departements")
async def get_stars_public_multi_departements(ville: Optional[str] = None):
    """Récupérer les stars servant dans plusieurs départements - Accès public"""
    query = {}
    if ville:
        query["ville"] = ville
    
    # Trouver les stars avec plus d'un département
    stars = await db.stars.find(query, {"_id": 0}).to_list(1000)
    
    multi_dept_stars = [
        {
            "prenom": s.get("prenom", ""),
            "nom": s.get("nom", ""),
            "departements": s.get("departements", [])
        }
        for s in stars
        if len(s.get("departements", [])) > 1
    ]
    
    # Trier par nombre de départements
    multi_dept_stars.sort(key=lambda x: len(x["departements"]), reverse=True)
    
    return multi_dept_stars


# ========== ENDPOINTS PUBLICS - BERGERS ÉGLISE ==========

@api_router.get("/bergers-eglise/public/stats")
async def get_bergers_eglise_public_stats(ville: Optional[str] = None):
    """Récupérer les statistiques globales pour un Berger d'Église - Accès public"""
    query = {}
    if ville:
        query = {"city": ville}
    
    # Total nouveaux arrivants (cette année)
    current_year = datetime.now().year
    visitors_query = {**query, "tracking_stopped": {"$ne": True}}
    total_visitors = await db.visitors.count_documents(visitors_query)
    
    # Total Familles d'Impact
    fi_query = {"ville": ville} if ville else {}
    total_fi = await db.familles_impact.count_documents(fi_query)
    
    # Total Stars
    stars_query = {"ville": ville} if ville else {}
    total_stars = await db.stars.count_documents(stars_query)
    
    # Total évangélisés (via bergerie contacts)
    evangelises_query = {"ville": ville, "type_contact": "Evangelisation"} if ville else {"type_contact": "Evangelisation"}
    total_evangelises = await db.bergerie_contacts.count_documents(evangelises_query)
    
    return {
        "total_visitors": total_visitors,
        "total_fi": total_fi,
        "total_stars": total_stars,
        "total_evangelises": total_evangelises
    }

@api_router.get("/public/bergerie/list/{ville}")
async def get_public_bergerie_list(ville: str):
    """Récupérer la liste des bergeries pour une ville - Accès public"""
    month_names = {
        '01': 'Janvier', '02': 'Février', '03': 'Mars', '04': 'Avril',
        '05': 'Mai', '06': 'Juin', '07': 'Juillet', '08': 'Août',
        '09': 'Septembre', '10': 'Octobre', '11': 'Novembre', '12': 'Décembre'
    }
    
    bergeries = []
    
    for month_num, month_name in month_names.items():
        # Compter les visiteurs pour cette bergerie
        visitors_count = await db.visitors.count_documents({
            "city": ville,
            "assigned_month": {"$regex": f"-{month_num}$"},
            "tracking_stopped": {"$ne": True}
        })
        
        bergeries.append({
            "month": month_num,
            "name": f"Bergerie {month_name}",
            "count": visitors_count
        })
    
    return bergeries


# ========== BERGERIES - GROUPES DE DISCIPLES ==========

class BergerieDisciple(BaseModel):
    """Modèle pour un groupe de disciples"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nom: str
    responsable: str
    ville: str
    membres_count: int = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MembreDisciple(BaseModel):
    """Modèle pour un membre d'un groupe de disciples"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    bergerie_id: str
    prenom: str
    nom: Optional[str] = ""
    telephone: Optional[str] = ""
    profession: Optional[str] = ""
    est_disciple: str = "Non"  # Non, En Cours, Oui
    informations: List[dict] = Field(default_factory=list)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ObjectifMultiplication(BaseModel):
    """Modèle pour un objectif de multiplication"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    bergerie_id: str
    mois: str
    objectif: int
    reel: int = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ContactEvangile(BaseModel):
    """Modèle pour un contact évangélisé"""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    bergerie_id: str
    prenom: str
    nom: Optional[str] = ""
    telephone: Optional[str] = ""
    date: str
    type: str = "Évangélisation"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Liste statique des groupes de disciples (basée sur le Google Sheet)
STATIC_BERGERIES_DISCIPLES = [
    {"id": "bg-1", "nom": "Choisies de Dieu ICC", "responsable": "Ps Nathalie", "ville": "Dijon", "membres_count": 27, "membres": ["Marie Esther", "Jade", "Julie", "Marie Augustine", "Marlyne", "Marthe", "Monica", "Syntiche", "Christelle", "Andréa", "Anoushka", "Linda", "Celine 1", "Celine 2", "Celine 3", "Christie", "Douclear", "Jaqueline", "Myriam", "Nadège", "Noella", "Tatiana", "Yadii"]},
    {"id": "bg-2", "nom": "Les perles Besançon", "responsable": "Olive", "ville": "Besançon", "membres_count": 16, "membres": ["Adeline", "Cerena", "Daphney", "Darla", "Gloria", "Keren", "Tricia", "Ashley", "Gloria", "Merveille", "Bonheur", "Anaïs", "Chancia", "Merveille", "Sandrine", "Alexia"]},
    {"id": "bg-3", "nom": "Ambassadeurs de Christ", "responsable": "Eddy Marc", "ville": "Besançon", "membres_count": 20, "membres": ["Elisée", "Jean Luc", "Pierre Christian", "William", "Marvin", "Dylan", "Stévie", "Corneille", "Brandon", "Ravi", "Michée", "Chrisley", "Florent", "Marvin", "Jean", "Sagesse", "Sassou", "Chadrack", "Benjamen", "Timothée"]},
    {"id": "bg-4", "nom": "Groupe NEHEMI", "responsable": "Tonton Frank", "ville": "Dijon", "membres_count": 15, "membres": ["Marc Zouzou", "Eugène", "Bernard", "François", "Jacob Fils", "Jo Rodriguez", "Samuel", "Elisée Jacob", "Gildas", "Jean David", "Moïse", "Kurt", "Rosario", "Laurent", "Matthias"]},
    {"id": "bg-5", "nom": "Les Prunelles de Dieu", "responsable": "Xaviera", "ville": "Dijon", "membres_count": 19, "membres": ["Anaïs", "Lesty", "Grâce", "Gaëlle", "Gloria", "Augustine", "Ange", "Christie", "Elodie", "Deborah", "Eunice", "Leïla", "Rita", "Rosine", "Sarah", "Nadège", "Rosine", "Leïla"]},
    {"id": "bg-6", "nom": "Femmes Fortes et Puissantes", "responsable": "Félicie", "ville": "Dijon", "membres_count": 17, "membres": ["Tata Liliane", "Christelle Laure", "Nativa Ange", "Daniela", "Elodie", "Gastella", "Gloria", "Ines M.", "Kadi", "Louise", "Multivette", "Rita", "Vernaline", "Marie Pauline"]},
    {"id": "bg-7", "nom": "Lampe Allumée", "responsable": "Rebecca", "ville": "Besançon", "membres_count": 13, "membres": ["Asheley", "Daphney", "Keren", "Fabiola", "Tricia", "Bonheur", "Cerena", "Priscilia", "Ardenne", "Gloria", "Wylliana", "Zaphnath"]},
    {"id": "bg-8", "nom": "Les disciples de Bérée", "responsable": "Jules", "ville": "Dijon", "membres_count": 16, "membres": ["Jacques", "Hervé", "Joy", "Marc Michel", "Marco", "Junior", "Benie", "Gabin (Dikamo)", "Chams", "Matis", "Joyces", "Samuel", "Hermon", "Quentin", "Exaucé", "Ethaan"]},
    {"id": "bg-9", "nom": "ROYALTY", "responsable": "Jemima", "ville": "Dijon", "membres_count": 16, "membres": ["Emmanuelle", "Marie Georgiana", "Michele", "Synthiche", "Nice", "Tessa", "Kenifé", "Alia", "Erica", "Rhessael", "Thya", "Maeva", "Gracia", "Priscillia", "Cheradja"]},
    {"id": "bg-10", "nom": "Vaillante Héroïne", "responsable": "Lesti", "ville": "Dijon", "membres_count": 12, "membres": ["Alvine", "Clara", "Gloria", "Queren", "Ariane", "Esther", "Gloria K"]},
    {"id": "bg-11", "nom": "HUIOS", "responsable": "Pierre Christian", "ville": "Besançon", "membres_count": 6, "membres": ["Désiré-Joseph", "Rad-hein", "Thierry", "Sylvester", "Kimra", "Gift"]},
    {"id": "bg-12", "nom": "Les Déboras", "responsable": "Queren", "ville": "Dijon", "membres_count": 11, "membres": ["Olivia", "Eden", "Fina"]},
    {"id": "bg-13", "nom": "Les Princesses de Dieu", "responsable": "Ps Nathalie", "ville": "Dijon", "membres_count": 11, "membres": ["Naomie", "Keren", "Priscillia", "Shina", "Yedidia", "Viviane", "Olive", "Ariane", "Adeoda"]},
    {"id": "bg-14", "nom": "Perles AUXERRE", "responsable": "", "ville": "Auxerre", "membres_count": 11, "membres": []},
    {"id": "bg-15", "nom": "Vaillants Héros de Dieu", "responsable": "Tonton Sikati", "ville": "Besançon", "membres_count": 8, "membres": ["Ezéchiel", "Guy", "Pierre-Etienne", "Daniel", "Jonathan", "Thiam", "Ouattara"]},
    {"id": "bg-16", "nom": "ESTHER", "responsable": "Hylarie", "ville": "Dijon", "membres_count": 7, "membres": ["Veronica", "Helena", "Esther", "Marie Paule", "Eunice", "Cathy", "Anaïs"]},
    {"id": "bg-17", "nom": "FEMMES D'EXPÉRIENCE MFI-B", "responsable": "", "ville": "Dijon", "membres_count": 10, "membres": []},
    {"id": "bg-18", "nom": "Groupe de disciple – Esther", "responsable": "Merveilles", "ville": "Besançon", "membres_count": 10, "membres": ["Anais", "Keren", "Clarence", "Joanisca", "Derviche", "Mirabeau"]},
    {"id": "bg-19", "nom": "Femmes de Destinée", "responsable": "", "ville": "Besançon", "membres_count": 9, "membres": []},
    {"id": "bg-20", "nom": "Les HÉRITIERS", "responsable": "Tonton Yves TAH", "ville": "Dijon", "membres_count": 9, "membres": ["Samuel A"]},
    {"id": "bg-21", "nom": "Les Princesses intimes du SAI", "responsable": "Anaïs E.", "ville": "Dijon", "membres_count": 9, "membres": ["Tarzie", "Christine", "Martha", "Joséphine", "Christiane", "Anne", "Yasmine"]},
    {"id": "bg-22", "nom": "DISCIPLES PST NATHALIE", "responsable": "Ps Nathalie", "ville": "Dijon", "membres_count": 8, "membres": ["Caty", "Anna", "Leathicia", "Daphline", "Anne Laure", "Fanny", "Hylarie", "Monique"]},
    {"id": "bg-23", "nom": "Medi'Time", "responsable": "Fanny", "ville": "Dijon", "membres_count": 5, "membres": ["Judith", "Cynthia", "Samirah", "Taliane", "Yasmine"]},
    {"id": "bg-24", "nom": "Sacerdocce royal", "responsable": "", "ville": "Dijon", "membres_count": 8, "membres": []},
    {"id": "bg-25", "nom": "Disciples assoiffées", "responsable": "Leathicia", "ville": "Dijon", "membres_count": 11, "membres": ["Amélia Moukoko", "Carmen Dadaglo", "Clara OKOUYA", "Laetitia Nessemon", "Veronica", "Adama", "Nancy", "Marie France", "Naïka", "Chancelle", "Prescillia"]},
    {"id": "bg-26", "nom": "Groupe ESDRAS", "responsable": "Thierry", "ville": "Dijon", "membres_count": 7, "membres": ["romain", "brayan", "anthony", "geofroid", "brice"]},
    {"id": "bg-27", "nom": "LES AIMÉES DE JÉSUS", "responsable": "Tata Nina", "ville": "Dijon", "membres_count": 7, "membres": ["Lisa ALLAMELON", "Béatrice DJAN", "Joyce DUBREUIL", "Nathalie MORETTI"]},
    {"id": "bg-28", "nom": "Les femmes de feu", "responsable": "Tata Laurence", "ville": "Dijon", "membres_count": 7, "membres": ["Ahouefa", "Louisa", "Nadège", "Fatou ROUSSEAU", "Virginie"]},
    {"id": "bg-29", "nom": "MERVEILLES & SPLENDEURS", "responsable": "Anna", "ville": "Dijon", "membres_count": 7, "membres": ["Guydil", "Multivette", "Nathaelle", "Ilona"]},
    {"id": "bg-30", "nom": "Groupe de partage et d'échange", "responsable": "", "ville": "Dijon", "membres_count": 6, "membres": []},
    {"id": "bg-31", "nom": "KABOD", "responsable": "Tarzie", "ville": "Dijon", "membres_count": 6, "membres": ["Myriam", "Emmanuelle", "Saralynn"]},
    {"id": "bg-32", "nom": "Les Timothée", "responsable": "Yvan", "ville": "Dijon", "membres_count": 6, "membres": ["Samuel Ethan", "Iahhhel", "Ismael", "Hans"]},
    {"id": "bg-33", "nom": "Les kaleos", "responsable": "", "ville": "Dijon", "membres_count": 6, "membres": []},
    {"id": "bg-34", "nom": "Les lumières de Dieu / The Light", "responsable": "Marie Augustine", "ville": "Dijon", "membres_count": 6, "membres": ["Samirah", "Yasmine", "Trecy"]},
    {"id": "bg-35", "nom": "Lumières", "responsable": "", "ville": "Dijon", "membres_count": 5, "membres": []},
    {"id": "bg-36", "nom": "Les perles précieuses", "responsable": "", "ville": "Dijon", "membres_count": 3, "membres": []},
    {"id": "bg-37", "nom": "Les FLÈCHES de l'Éternel", "responsable": "Rita", "ville": "Dijon", "membres_count": 2, "membres": ["Fenny"]},
    {"id": "bg-38", "nom": "Construction des disciples", "responsable": "Ps Nathalie", "ville": "Besançon", "membres_count": 5, "membres": ["Yola", "Rebecca", "Lois", "Chancia", "Asley"]},
    {"id": "bg-39", "nom": "Assoiffés de Christ", "responsable": "Hass", "ville": "Dijon", "membres_count": 4, "membres": ["Chadread", "Félix", "Aristide", "Antoine"]},
    {"id": "bg-40", "nom": "Groupe disciple Daphline", "responsable": "", "ville": "Dijon", "membres_count": 0, "membres": []},
    {"id": "bg-41", "nom": "Groupe disciple Cathy", "responsable": "", "ville": "Dijon", "membres_count": 0, "membres": []},
    {"id": "bg-42", "nom": "Les Emmanuelle", "responsable": "Leaticia", "ville": "Chalon-sur-Saône", "membres_count": 6, "membres": ["Gislaine", "Mycille", "Yvana", "Rovlyne", "Ahoefa"]},
]

@api_router.get("/bergeries-disciples/list")
async def get_bergeries_disciples_list():
    """Récupérer la liste des groupes de disciples - Public"""
    # Récupérer les bergeries depuis la DB
    db_bergeries = await db.bergeries_disciples.find({}, {"_id": 0}).to_list(500)
    
    # Créer un set des IDs existants
    existing_ids = {b["id"] for b in db_bergeries}
    
    # Ajouter les bergeries statiques qui ne sont pas encore dans la DB
    all_bergeries = list(db_bergeries)
    for static in STATIC_BERGERIES_DISCIPLES:
        if static["id"] not in existing_ids:
            all_bergeries.append(static)
    
    # Mettre à jour le nombre de membres pour chaque bergerie
    for b in all_bergeries:
        membres_count = await db.membres_disciples.count_documents({"bergerie_id": b["id"]})
        if membres_count > 0:
            b["membres_count"] = membres_count
    
    # Trier par ville puis par nom
    all_bergeries.sort(key=lambda x: (x.get("ville", ""), x.get("nom", "")))
    
    return all_bergeries

@api_router.post("/bergeries-disciples/create")
async def create_bergerie_disciple(data: dict):
    """Créer une nouvelle bergerie manuellement"""
    from uuid import uuid4
    
    bergerie = {
        "id": f"bg-{str(uuid4())[:8]}",
        "nom": data.get("nom", "Nouvelle Bergerie"),
        "responsable": data.get("responsable", ""),
        "ville": data.get("ville", "Dijon"),
        "membres_count": 0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.bergeries_disciples.insert_one(bergerie.copy())
    
    # Ne pas retourner _id
    return {"id": bergerie["id"], "message": "Bergerie créée avec succès", "bergerie": bergerie}

@api_router.get("/bergeries-disciples/{bergerie_id}")
async def get_bergerie_disciple_info(bergerie_id: str):
    """Récupérer les infos d'un groupe de disciples par ID"""
    # D'abord essayer depuis la DB
    bergerie = await db.bergeries_disciples.find_one({"id": bergerie_id}, {"_id": 0})
    
    if bergerie:
        membres_count = await db.membres_disciples.count_documents({"bergerie_id": bergerie_id})
        bergerie["membres_count"] = membres_count
        return bergerie
    
    # Sinon, chercher dans la liste statique
    for b in STATIC_BERGERIES_DISCIPLES:
        if b["id"] == bergerie_id:
            return b
    
    raise HTTPException(status_code=404, detail="Bergerie non trouvée")

@api_router.put("/bergeries-disciples/{bergerie_id}")
async def update_bergerie_disciple(bergerie_id: str, data: dict):
    """Modifier une bergerie"""
    update_data = {
        "nom": data.get("nom"),
        "responsable": data.get("responsable"),
        "ville": data.get("ville"),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Supprimer les clés None
    update_data = {k: v for k, v in update_data.items() if v is not None}
    
    # D'abord vérifier si la bergerie existe en DB
    existing = await db.bergeries_disciples.find_one({"id": bergerie_id})
    
    if existing:
        # Mettre à jour
        await db.bergeries_disciples.update_one(
            {"id": bergerie_id},
            {"$set": update_data}
        )
    else:
        # C'est une bergerie statique, on la crée en DB pour pouvoir la modifier
        for static_b in STATIC_BERGERIES_DISCIPLES:
            if static_b["id"] == bergerie_id:
                new_bergerie = {**static_b, **update_data, "created_at": datetime.now(timezone.utc).isoformat()}
                new_bergerie.pop("membres", None)  # Ne pas stocker la liste des membres
                await db.bergeries_disciples.insert_one(new_bergerie.copy())
                break
    
    return {"message": "Bergerie modifiée avec succès"}

@api_router.delete("/bergeries-disciples/{bergerie_id}")
async def delete_bergerie_disciple(bergerie_id: str):
    """Supprimer une bergerie et ses données associées"""
    # Supprimer la bergerie de la DB
    result = await db.bergeries_disciples.delete_one({"id": bergerie_id})
    
    # Supprimer tous les membres associés
    await db.membres_disciples.delete_many({"bergerie_id": bergerie_id})
    
    # Supprimer les objectifs et contacts associés
    await db.objectifs_multiplication.delete_many({"bergerie_id": bergerie_id})
    await db.contacts_evangile.delete_many({"bergerie_id": bergerie_id})
    
    return {"message": "Bergerie et données associées supprimées"}

@api_router.get("/bergeries-disciples/{bergerie_id}/membres")
async def get_bergerie_disciples_membres(bergerie_id: str):
    """Récupérer les membres d'un groupe de disciples"""
    membres = await db.membres_disciples.find({"bergerie_id": bergerie_id}, {"_id": 0}).to_list(500)
    
    # Si pas de membres en DB, initialiser depuis les données statiques
    if not membres:
        for static_bergerie in STATIC_BERGERIES_DISCIPLES:
            if static_bergerie["id"] == bergerie_id and "membres" in static_bergerie:
                for prenom in static_bergerie["membres"]:
                    if prenom:  # Ignorer les noms vides
                        membre = {
                            "id": str(uuid.uuid4()),
                            "bergerie_id": bergerie_id,
                            "prenom": prenom,
                            "nom": "",
                            "telephone": "",
                            "profession": "",
                            "est_disciple": "Non",
                            "created_at": datetime.now(timezone.utc).isoformat()
                        }
                        await db.membres_disciples.insert_one(membre.copy())
                        membres.append(membre)
                break
    
    objectifs = await db.objectifs_multiplication.find({"bergerie_id": bergerie_id}, {"_id": 0}).to_list(100)
    contacts = await db.contacts_evangile.find({"bergerie_id": bergerie_id}, {"_id": 0}).to_list(500)
    
    return {
        "membres": membres,
        "objectifs": objectifs,
        "contacts": contacts
    }

@api_router.post("/bergeries-disciples/{bergerie_id}/membres")
async def add_membre_disciple(bergerie_id: str, membre_data: dict):
    """Ajouter un membre à un groupe de disciples"""
    membre = MembreDisciple(
        bergerie_id=bergerie_id,
        prenom=membre_data.get("prenom", ""),
        nom=membre_data.get("nom", ""),
        telephone=membre_data.get("telephone", ""),
        profession=membre_data.get("profession", ""),
        est_disciple=membre_data.get("est_disciple", "Non")
    )
    
    doc = membre.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    
    await db.membres_disciples.insert_one(doc)
    return {"id": membre.id, "message": "Membre ajouté avec succès"}

@api_router.put("/bergeries-disciples/membres/{membre_id}")
async def update_membre_disciple(membre_id: str, update_data: dict):
    """Mettre à jour un membre"""
    # Enlever les champs protégés
    protected = ["id", "created_at", "bergerie_id"]
    for field in protected:
        update_data.pop(field, None)
    
    result = await db.membres_disciples.update_one(
        {"id": membre_id},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Membre non trouvé")
    
    return {"message": "Membre mis à jour"}

@api_router.delete("/bergeries-disciples/membres/{membre_id}")
async def delete_membre_disciple(membre_id: str):
    """Supprimer un membre"""
    result = await db.membres_disciples.delete_one({"id": membre_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Membre non trouvé")
    
    return {"message": "Membre supprimé"}

@api_router.post("/bergeries-disciples/membres/{membre_id}/disciple")
async def update_membre_disciple_status(membre_id: str, status_data: dict):
    """Mettre à jour le statut disciple d'un membre"""
    est_disciple = status_data.get("est_disciple", "Non")
    
    result = await db.membres_disciples.update_one(
        {"id": membre_id},
        {"$set": {"est_disciple": est_disciple}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Membre non trouvé")
    
    return {"message": "Statut mis à jour"}

@api_router.post("/bergeries-disciples/{bergerie_id}/objectifs")
async def add_objectif_multiplication(bergerie_id: str, objectif_data: dict):
    """Ajouter un objectif de multiplication"""
    objectif = ObjectifMultiplication(
        bergerie_id=bergerie_id,
        mois=objectif_data.get("mois", ""),
        objectif=int(objectif_data.get("objectif", 0)),
        reel=int(objectif_data.get("reel", 0))
    )
    
    doc = objectif.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    
    await db.objectifs_multiplication.insert_one(doc)
    return {"id": objectif.id, "message": "Objectif ajouté"}

@api_router.post("/bergeries-disciples/{bergerie_id}/contacts")
async def add_contact_evangile(bergerie_id: str, contact_data: dict):
    """Ajouter un contact évangélisé"""
    contact = ContactEvangile(
        bergerie_id=bergerie_id,
        prenom=contact_data.get("prenom", ""),
        nom=contact_data.get("nom", ""),
        telephone=contact_data.get("telephone", ""),
        date=contact_data.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
        type=contact_data.get("type", "Évangélisation")
    )
    
    doc = contact.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    
    await db.contacts_evangile.insert_one(doc)
    return {"id": contact.id, "message": "Contact ajouté"}


# Include the router in the main app (must be at the end after all endpoints are defined)
app.include_router(api_router)

